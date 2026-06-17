from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date
import io
import zipfile

from app.database import get_db
from app.routers.correspondencia import _calcular_semaforo_row
from app.auth_utils import puede_escribir as _pw, registrar_log

_MOD = "backup"

router = APIRouter(prefix="/backup")
templates = make_templates(str(Path(__file__).parent.parent / "templates"))


def _v(val):
    """Sanitiza valor de celda: devuelve None si está vacío."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "None", "#VALUE!", "#N/A", "#REF!", "—"):
        return None
    return s


# ── Página principal ───────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def backup_home(request: Request, msg: str = ""):
    conn = get_db()
    total_base          = conn.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]
    total_digitales     = conn.execute("SELECT COUNT(*) FROM exp_digitales").fetchone()[0]
    total_sala          = conn.execute("SELECT COUNT(*) FROM sala_agenda").fetchone()[0]
    total_control_autos = conn.execute("SELECT COUNT(*) FROM control_autos_sustanciacion").fetchone()[0]
    total_corr          = conn.execute("SELECT COUNT(*) FROM correspondencia").fetchone()[0]
    total_sdqs          = conn.execute("SELECT COUNT(*) FROM sdqs").fetchone()[0]
    total_seguimiento   = conn.execute("SELECT COUNT(*) FROM seguimiento_mensual").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("backup.html", {
        "request": request,
        "msg": msg,
        "total_base": total_base,
        "total_digitales": total_digitales,
        "total_sala": total_sala,
        "total_control_autos": total_control_autos,
        "total_corr": total_corr,
        "total_sdqs": total_sdqs,
        "total_seguimiento": total_seguimiento,
    })


# ── Exportar Excel completo (3 hojas) ─────────────────────────────────────────

@router.get("/exportar")
async def backup_exportar():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/backup/?msg=error_openpyxl")

    conn = get_db()
    exps        = conn.execute("SELECT * FROM expedientes ORDER BY anio, n_expediente").fetchall()
    dig_exps    = conn.execute("SELECT * FROM exp_digitales ORDER BY anio DESC, n_expediente ASC").fetchall()
    dig_coms    = conn.execute(
        "SELECT * FROM exp_comunicaciones ORDER BY exp_digital_id ASC, fecha_envio ASC, id ASC"
    ).fetchall()
    dig_revs    = conn.execute(
        "SELECT exp_digital_id, MAX(fecha_revision) AS ultima_revision FROM exp_revisiones GROUP BY exp_digital_id"
    ).fetchall()
    sala        = conn.execute("SELECT * FROM sala_agenda ORDER BY fecha, franja").fetchall()
    ctrl_autos  = conn.execute(
        "SELECT * FROM control_autos_sustanciacion ORDER BY fecha_auto ASC, id ASC"
    ).fetchall()
    sdqs_rows   = conn.execute("SELECT * FROM sdqs ORDER BY fecha_asignacion, id").fetchall()
    corr_rows   = conn.execute("""
        SELECT c.*,
               GROUP_CONCAT(rs.radicado, ' | ') AS radicados_salida,
               GROUP_CONCAT(COALESCE(rs.url, ''), ' | ') AS radicados_urls
        FROM correspondencia c
        LEFT JOIN correspondencia_radicados_salida rs ON rs.correspondencia_id = c.id
        GROUP BY c.id
        ORDER BY c.fecha_ingreso DESC
    """).fetchall()
    seg_rows    = conn.execute("""
        SELECT s.anio, s.mes, e.n_expediente, e.anio AS exp_anio, e.abogado_asignado,
               s.descripcion, s.created_by, s.created_at
        FROM seguimiento_mensual s
        JOIN expedientes e ON e.id = s.expediente_id
        ORDER BY e.anio DESC, CAST(e.n_expediente AS INTEGER)
    """).fetchall()
    conn.close()

    ultima_rev = {r["exp_digital_id"]: r["ultima_revision"] for r in dig_revs}
    coms_por_exp: dict[int, list] = {}
    for c in dig_coms:
        coms_por_exp.setdefault(c["exp_digital_id"], []).append(dict(c))

    wb = openpyxl.Workbook()

    h_font    = Font(bold=True, color="FFFFFF", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", fgColor="EBF1F8")
    sub_fill  = PatternFill("solid", fgColor="dbeafe")
    fill_base = PatternFill("solid", fgColor="1B4F8A")
    fill_dig  = PatternFill("solid", fgColor="1e3a5f")
    fill_sala = PatternFill("solid", fgColor="065F46")

    # ── Hoja 1: Base Expedientes ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Base Expedientes"

    headers1 = [
        "N. EXPEDIENTE","AÑO","MES","MEDIO DE INGRESO","N. RADICADO",
        "FECHA RADICADO","ABOGADO ASIGNADO","ENTIDAD ORIGEN",
        "QUEJOSO","ASUNTO","IMPEDIMENTO","FECHA APERTURA EXPEDIENTE",
        "N. AUTO APERTURA INDAGACIÓN","FECHA AUTO APERTURA IND.","TIPO EXPEDIENTE",
        "TIPOLOGIA","RELACIONADO SINIESTRO","RESPONSABLE SINIESTRO",
        "RELACIONADO MALTRATO","RELACIONADO CORRUPCIÓN","VALORES INSTITUCIONALES",
        "FECHA HECHOS (OBSERVACIONES)","FECHA HECHOS",
        "F. ÚLTIMA ACTUACIÓN INDAGACIÓN","N. AUTO ÚLTIMA ACT. IND.",
        "F. APERTURA INVESTIGACIÓN","N. AUTO APERTURA INV.",
        "NOMBRE INVESTIGADO","CÉDULA","PERFIL INVESTIGADO","ÁREA ORIGEN INVESTIGADO",
        "FECHA PRÓRROGA","N. AUTO PRÓRROGA","TIEMPO PRÓRROGA",
        "F. ÚLTIMA ACTUACIÓN INVESTIGACIÓN","N. AUTO ÚLTIMA ACT. INV.",
        "N. AUTO TRASLADO","F. AUTO TRASLADO",
        "N. AUTO ACUMULACIÓN","F. AUTO ACUMULACIÓN","EXPEDIENTE ACUMULA",
        "F. AUTO ARCHIVO","N. AUTO ARCHIVO",
        "F. AUTO PLIEGO CARGOS","N. AUTO PLIEGO CARGOS",
        "ETAPA ACTUAL","ESTADO DEL PROCESO","OBSERVACIONES",
        "CREADO POR","FECHA CREACIÓN","ÚLTIMA ACTUALIZACIÓN",
    ]
    campos1 = [
        "n_expediente","anio","mes","medio_ingreso","n_radicado",
        "fecha_radicado","abogado_asignado","entidad_origen",
        "quejoso","asunto","impedimento","fecha_apertura_expediente",
        "numero_auto_apertura_ind","fecha_auto_apertura_ind","tipo_expediente",
        "tipologia","relacionado_siniestro","responsable_siniestro",
        "relacionado_maltrato","relacionado_corrupcion","valores_institucionales",
        "fecha_hechos_obs","fecha_hechos",
        "fecha_ultima_act_indagacion","numero_auto_ultima_act_ind",
        "fecha_apertura_investigacion","numero_auto_apertura_inv",
        "nombre_investigado","cedula","perfil_investigado","area_origen_investigado",
        "fecha_prorroga","numero_auto_prorroga","tiempo_prorroga",
        "fecha_ultima_act_investigacion","numero_auto_ultima_act_inv",
        "numero_auto_traslado","fecha_auto_traslado",
        "numero_auto_acumulacion","fecha_auto_acumulacion","expediente_acumula",
        "fecha_auto_archivo","numero_auto_archivo",
        "fecha_auto_pliego_cargos","numero_auto_pliego_cargos",
        "etapa_actual","estado_proceso","observaciones",
        "created_by","created_at","updated_at",
    ]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_base
        cell.font = h_font
        cell.alignment = center
    ws1.row_dimensions[1].height = 40

    for row_idx, row in enumerate(exps, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, campo in enumerate(campos1, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=d.get(campo))
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    for col in ws1.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    ws1.freeze_panes = "A2"

    # ── Hoja 2: Exp. Digitales ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Exp. Digitales")

    headers2 = [
        "N° Expediente", "Año", "Abogado", "Etapa", "Queja Inicial",
        "Radicado Auto", "Nombre Auto", "Fecha Auto",
        "Obs. Generales", "Última Revisión",
        "Radicado Comunicación", "Dependencia", "Fecha Envío",
        "Fecha Seguimiento", "Radicado Respuesta", "Fecha Respuesta",
        "Responsable", "Observaciones Com.",
    ]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_dig
        cell.font = h_font
        cell.alignment = center
    ws2.row_dimensions[1].height = 30

    for exp in dig_exps:
        ed = dict(exp)
        exp_coms = coms_por_exp.get(ed["id"], [])
        primera = exp_coms[0] if exp_coms else {}
        row = [
            ed.get("n_expediente"), ed.get("anio"), ed.get("abogado"), ed.get("etapa"),
            ed.get("queja_inicial"), ed.get("radicado_auto"), ed.get("nombre_auto"), ed.get("fecha_auto"),
            ed.get("observaciones"), ultima_rev.get(ed["id"]),
            primera.get("radicado_comunicacion"), primera.get("dependencia"),
            primera.get("fecha_envio"), primera.get("fecha_seguimiento"),
            primera.get("radicado_respuesta"), primera.get("fecha_respuesta"),
            primera.get("responsable"), primera.get("observaciones"),
        ]
        ws2.append(row)
        for com in exp_coms[1:]:
            sub_row = [None] * 10 + [
                com.get("radicado_comunicacion"), com.get("dependencia"),
                com.get("fecha_envio"), com.get("fecha_seguimiento"),
                com.get("radicado_respuesta"), com.get("fecha_respuesta"),
                com.get("responsable"), com.get("observaciones"),
            ]
            ws2.append(sub_row)
            for cell in ws2[ws2.max_row]:
                cell.fill = sub_fill

    col_widths2 = [15, 6, 22, 20, 14, 20, 30, 14, 40, 22, 22, 25, 14, 16, 22, 14, 20, 40]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Hoja 3: Sala de Audiencias ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Sala de Audiencias")

    headers3 = ["Fecha", "Franja", "Título", "Descripción", "Estado", "Responsable", "Fecha Creación"]
    campos3   = ["fecha", "franja", "titulo", "descripcion", "estado", "responsable", "created_at"]

    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_sala
        cell.font = h_font
        cell.alignment = center
    ws3.row_dimensions[1].height = 30

    for row_idx, row in enumerate(sala, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, campo in enumerate(campos3, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=d.get(campo))
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    col_widths3 = [14, 16, 30, 40, 12, 25, 20]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ── Hoja 4: Control de Autos de Sustanciación ─────────────────────────────
    ws4 = wb.create_sheet("Control Autos")
    fill_autos = PatternFill("solid", fgColor="2E7D32")

    headers4 = ["EXPEDIENTE", "NÚMERO DEL AUTO", "FECHA DEL AUTO", "ASUNTO AUTO", "ABOGADO RESPONSABLE", "OBSERVACIONES",
                "CREADO POR", "FECHA CREACIÓN", "ÚLTIMA ACTUALIZACIÓN"]
    campos4   = ["expediente", "numero_auto", "fecha_auto", "asunto_auto", "abogado_responsable", "observaciones",
                "created_by", "created_at", "updated_at"]

    for col_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_autos
        cell.font = h_font
        cell.alignment = center
    ws4.row_dimensions[1].height = 30

    for row_idx, row in enumerate(ctrl_autos, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, campo in enumerate(campos4, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=d.get(campo))
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    col_widths4 = [18, 16, 16, 48, 22, 30, 20, 20, 20]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    # ── Hoja 5: SDQS ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("SDQS")
    fill_sdqs = PatternFill("solid", fgColor="7B3F00")
    link_font = Font(color="0563C1", underline="single")

    headers5 = [
        "MES", "FECHA ASIGNACION", "SDQS", "URL SDQS", "FECHA VENCIMIENTO",
        "QUEJOSO", "CORREO", "TEMA", "COMPETENCIA OCDI", "BPM", "RESPONSABLE",
        "RAD SALIDA", "URL RAD SALIDA", "FECHA RESPUESTA", "OBSERVACIONES",
        "ESTADO PROCESO", "HECHO CORRUPTO", "VALOR INSTITUCIONAL", "TIPOLOGIA",
        "CREADO POR", "FECHA CREACIÓN", "ÚLTIMA ACTUALIZACIÓN",
    ]
    campos5 = [
        "mes", "fecha_asignacion", "sdqs", "url_sdqs", "fecha_vencimiento",
        "quejoso", "correo", "tema", "competencia_ocdi", "bpm", "responsable",
        "rad_salida", "url_rad_salida", "fecha_respuesta", "observaciones",
        "estado_proceso", "hecho_corrupto", "valor_institucional", "tipologia",
        "created_by", "created_at", "updated_at",
    ]
    for col_idx, h in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_sdqs
        cell.font = h_font
        cell.alignment = center
    ws5.row_dimensions[1].height = 30

    for row_idx, row in enumerate(sdqs_rows, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, campo in enumerate(campos5, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=d.get(campo))
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
        if d.get("url_sdqs") and d.get("sdqs"):
            c = ws5.cell(row_idx, 3)
            c.hyperlink = d["url_sdqs"]
            c.font = link_font
        if d.get("url_rad_salida") and d.get("rad_salida"):
            c = ws5.cell(row_idx, 12)
            c.hyperlink = d["url_rad_salida"]
            c.font = link_font

    col_widths5 = [10, 16, 14, 40, 16, 28, 28, 50, 14, 14, 28, 16, 40, 16, 40, 22, 22, 22, 20, 20, 20, 20]
    for i, w in enumerate(col_widths5, 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws5.freeze_panes = "A2"

    # ── Hoja 6: Correspondencia ───────────────────────────────────────────────
    ws6 = wb.create_sheet("Correspondencia")
    fill_corr6 = PatternFill("solid", fgColor="1B4F8A")

    headers6 = [
        "AÑO", "MES", "FECHA INGRESO", "N. RADICADO", "ENTIDAD",
        "CORREO REMITENTE", "ASUNTO", "SINPROC PERSONERÍA",
        "TIPO REQUERIMIENTO", "TÉRMINO (DÍAS)", "TIPO DOCUMENTO",
        "RESPONSABLE", "CASO BMP",
        "N. RADICADOS SALIDA", "URLS RAD SALIDA", "FECHA RAD SALIDA",
        "TIPO RESPUESTA", "TRÁMITE DE SALIDA",
        "FECHA CREACIÓN", "ÚLTIMA ACTUALIZACIÓN",
    ]
    for col_idx, h in enumerate(headers6, 1):
        cell = ws6.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_corr6
        cell.font = h_font
        cell.alignment = center
    ws6.row_dimensions[1].height = 30

    for row_idx, row in enumerate(corr_rows, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        vals6 = [
            d.get("anio"), d.get("mes"),
            d.get("fecha_ingreso")[:10] if d.get("fecha_ingreso") else None,
            d.get("n_radicado"), d.get("origen"), d.get("correo_remitente"), d.get("asunto"),
            d.get("sinproc_personeria"), d.get("tipo_requerimiento"), d.get("termino_dias"),
            d.get("tipo_documento"), d.get("responsable"), d.get("caso_bmp"),
            d.get("radicados_salida"),
            d.get("radicados_urls"),
            d.get("fecha_radicado_salida"),
            d.get("tipo_respuesta"), d.get("tramite_salida"),
            d.get("created_at")[:16] if d.get("created_at") else None,
            d.get("updated_at")[:16] if d.get("updated_at") else None,
        ]
        for col_idx, v in enumerate(vals6, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (5, 7)))
            if fill:
                cell.fill = fill

    col_widths6 = [6, 12, 20, 18, 30, 30, 40, 20, 40, 10, 18, 28, 10, 30, 50, 20, 25, 30, 20, 20]
    for i, w in enumerate(col_widths6, 1):
        ws6.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws6.freeze_panes = "A2"

    # ── Hoja 7: Seguimiento Mensual ───────────────────────────────────────────
    ws7 = wb.create_sheet("Seguimiento Mensual")
    fill_seg7 = PatternFill("solid", fgColor="0D3060")

    # "AÑO EXPEDIENTE" es indispensable para reimportar correctamente: los
    # números de expediente se reinician cada año (ver H2/H11), así que
    # buscar solo por N. EXPEDIENTE al reimportar puede enlazar el
    # seguimiento al expediente equivocado si el número se repite en otro año.
    headers7 = ["AÑO", "MES", "N. EXPEDIENTE", "AÑO EXPEDIENTE", "ABOGADO", "DESCRIPCIÓN ACTUACIÓN", "REGISTRADO POR", "FECHA REGISTRO"]
    for col_idx, h in enumerate(headers7, 1):
        cell = ws7.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_seg7
        cell.font = h_font
        cell.alignment = center
    ws7.row_dimensions[1].height = 30

    for row_idx, row in enumerate(seg_rows, 2):
        d = dict(row)
        fill = alt_fill if row_idx % 2 == 0 else None
        vals7 = [d.get("anio"), d.get("mes"), d.get("n_expediente"), d.get("exp_anio"),
                 d.get("abogado_asignado"), d.get("descripcion"),
                 d.get("created_by"), d.get("created_at")]
        for col_idx, v in enumerate(vals7, 1):
            cell = ws7.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 6))
            if fill:
                cell.fill = fill

    col_widths7 = [10, 14, 16, 12, 28, 60, 24, 20]
    for i, w in enumerate(col_widths7, 1):
        ws7.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws7.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = date.today().strftime("%Y%m%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=OCDI_Respaldo_Completo_{hoy}.xlsx"},
    )


# ── Importar Excel completo (reemplaza todo) ───────────────────────────────────

@router.post("/importar")
async def backup_importar(request: Request, archivo: UploadFile = File(...)):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/backup/?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/backup/?msg=error_openpyxl", status_code=303)

    contenido = await archivo.read()
    if not contenido:
        return RedirectResponse("/backup/?msg=error_vacio", status_code=303)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        return RedirectResponse("/backup/?msg=error_archivo", status_code=303)

    conn = get_db()
    stats = {"base": 0, "digitales": 0, "coms": 0, "sala": 0, "autos": 0, "sdqs": 0, "corr": 0, "seg": 0, "sdqs_omitidos": 0}

    try:
        # ── Hoja 1: Base Expedientes ───────────────────────────────────────────
        if "Base Expedientes" in wb.sheetnames:
            ws1 = wb["Base Expedientes"]
            campos1 = [
                "n_expediente","anio","mes","medio_ingreso","n_radicado",
                "fecha_radicado","abogado_asignado","entidad_origen",
                "quejoso","asunto","impedimento","fecha_apertura_expediente",
                "numero_auto_apertura_ind","fecha_auto_apertura_ind","tipo_expediente",
                "tipologia","relacionado_siniestro","responsable_siniestro",
                "relacionado_maltrato","relacionado_corrupcion","valores_institucionales",
                "fecha_hechos_obs","fecha_hechos",
                "fecha_ultima_act_indagacion","numero_auto_ultima_act_ind",
                "fecha_apertura_investigacion","numero_auto_apertura_inv",
                "nombre_investigado","cedula","perfil_investigado","area_origen_investigado",
                "fecha_prorroga","numero_auto_prorroga","tiempo_prorroga",
                "fecha_ultima_act_investigacion","numero_auto_ultima_act_inv",
                "numero_auto_traslado","fecha_auto_traslado",
                "numero_auto_acumulacion","fecha_auto_acumulacion","expediente_acumula",
                "fecha_auto_archivo","numero_auto_archivo",
                "fecha_auto_pliego_cargos","numero_auto_pliego_cargos",
                "etapa_actual","estado_proceso","observaciones",
                "created_by","created_at","updated_at",
            ]
            conn.execute("DELETE FROM expedientes")
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                n_exp = _v(row[0]) if len(row) > 0 else None
                if not n_exp:
                    continue
                vals = [_v(row[i]) if i < len(row) else None for i in range(len(campos1))]
                conn.execute(
                    f"INSERT INTO expedientes ({', '.join(campos1)}) VALUES ({', '.join(['?']*len(campos1))})",
                    vals,
                )
                stats["base"] += 1

        # ── Hoja 2: Exp. Digitales ─────────────────────────────────────────────
        if "Exp. Digitales" in wb.sheetnames:
            ws2 = wb["Exp. Digitales"]
            # Borrar todo (CASCADE elimina comunicaciones y revisiones)
            conn.execute("DELETE FROM exp_revisiones")
            conn.execute("DELETE FROM exp_comunicaciones")
            conn.execute("DELETE FROM exp_digitales")

            current_exp_id = None
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                n_exp = _v(row[0]) if len(row) > 0 else None
                if n_exp:
                    # Fila de expediente
                    cur = conn.execute(
                        """INSERT INTO exp_digitales
                           (n_expediente, anio, abogado, etapa, queja_inicial,
                            radicado_auto, nombre_auto, fecha_auto, observaciones)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        [
                            n_exp,
                            _v(row[1]) if len(row) > 1 else None,
                            _v(row[2]) if len(row) > 2 else None,
                            _v(row[3]) if len(row) > 3 else None,
                            _v(row[4]) if len(row) > 4 else None,
                            _v(row[5]) if len(row) > 5 else None,
                            _v(row[6]) if len(row) > 6 else None,
                            _v(row[7]) if len(row) > 7 else None,
                            _v(row[8]) if len(row) > 8 else None,
                        ],
                    )
                    current_exp_id = cur.lastrowid
                    stats["digitales"] += 1
                    # Restaurar última revisión si existe
                    ultima_rev_val = _v(row[9]) if len(row) > 9 else None
                    if ultima_rev_val:
                        conn.execute(
                            "INSERT INTO exp_revisiones (exp_digital_id, fecha_revision) VALUES (?, ?)",
                            (current_exp_id, ultima_rev_val),
                        )

                # Comunicación (fila principal o sub-fila)
                if current_exp_id:
                    rad_com    = _v(row[10]) if len(row) > 10 else None
                    dependencia = _v(row[11]) if len(row) > 11 else None
                    if rad_com or dependencia:
                        conn.execute(
                            """INSERT INTO exp_comunicaciones
                               (exp_digital_id, radicado_comunicacion, dependencia,
                                fecha_envio, fecha_seguimiento, radicado_respuesta,
                                fecha_respuesta, responsable, observaciones)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            [
                                current_exp_id,
                                rad_com,
                                dependencia,
                                _v(row[12]) if len(row) > 12 else None,
                                _v(row[13]) if len(row) > 13 else None,
                                _v(row[14]) if len(row) > 14 else None,
                                _v(row[15]) if len(row) > 15 else None,
                                _v(row[16]) if len(row) > 16 else None,
                                _v(row[17]) if len(row) > 17 else None,
                            ],
                        )
                        stats["coms"] += 1

        # ── Hoja 3: Sala de Audiencias ─────────────────────────────────────────
        if "Sala de Audiencias" in wb.sheetnames:
            ws3 = wb["Sala de Audiencias"]
            conn.execute("DELETE FROM sala_agenda")
            for row in ws3.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                fecha  = _v(row[0]) if len(row) > 0 else None
                franja = _v(row[1]) if len(row) > 1 else None
                if not fecha or not franja:
                    continue
                conn.execute(
                    """INSERT INTO sala_agenda (fecha, franja, titulo, descripcion, estado, responsable, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, COALESCE(?, datetime('now','localtime')))""",
                    [
                        fecha, franja,
                        _v(row[2]) if len(row) > 2 else None,
                        _v(row[3]) if len(row) > 3 else None,
                        _v(row[4]) if len(row) > 4 else "Ocupado",
                        _v(row[5]) if len(row) > 5 else None,
                        _v(row[6]) if len(row) > 6 else None,
                    ],
                )
                stats["sala"] += 1

        # ── Hoja 4: Control de Autos de Sustanciación ─────────────────────────
        if "Control Autos" in wb.sheetnames:
            ws4 = wb["Control Autos"]
            conn.execute("DELETE FROM control_autos_sustanciacion")
            for row in ws4.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                expediente  = _v(row[0]) if len(row) > 0 else None
                numero_auto = _v(row[1]) if len(row) > 1 else None
                fecha_raw   = row[2] if len(row) > 2 else None
                from datetime import datetime as _dt
                if fecha_raw and hasattr(fecha_raw, 'strftime'):
                    fecha_auto = fecha_raw.strftime("%Y-%m-%d")
                elif fecha_raw:
                    s = str(fecha_raw).strip()
                    fecha_auto = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            fecha_auto = _dt.strptime(s, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            pass
                    if not fecha_auto:
                        fecha_auto = s
                else:
                    fecha_auto = None
                asunto_auto = _v(row[3]) if len(row) > 3 else None
                abogado     = _v(row[4]) if len(row) > 4 else None
                obs         = _v(row[5]) if len(row) > 5 else None
                creado_por  = _v(row[6]) if len(row) > 6 else None
                creado_en   = _v(row[7]) if len(row) > 7 else None
                actualiz_en = _v(row[8]) if len(row) > 8 else None
                if not any([expediente, numero_auto, fecha_auto, asunto_auto, abogado]):
                    continue
                # Saltar filas de descripción del pie del formato
                if numero_auto and len(numero_auto) > 20:
                    continue
                conn.execute(
                    """INSERT INTO control_autos_sustanciacion
                       (expediente, numero_auto, fecha_auto, asunto_auto, abogado_responsable, observaciones,
                        created_by, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?,
                               COALESCE(?, datetime('now','localtime')),
                               COALESCE(?, datetime('now','localtime')))""",
                    [expediente, numero_auto, fecha_auto, asunto_auto, abogado, obs,
                     creado_por, creado_en, actualiz_en],
                )
                stats["autos"] += 1

        # ── Hoja 5: SDQS ───────────────────────────────────────────────────────
        if "SDQS" in wb.sheetnames:
            ws5 = wb["SDQS"]
            campos_sdqs = [
                "mes", "fecha_asignacion", "sdqs", "url_sdqs", "fecha_vencimiento",
                "quejoso", "correo", "tema", "competencia_ocdi", "bpm", "responsable",
                "rad_salida", "url_rad_salida", "fecha_respuesta", "observaciones",
                "estado_proceso", "hecho_corrupto", "valor_institucional", "tipologia",
                "created_by", "created_at", "updated_at",
            ]
            # Columnas NOT NULL de sdqs (además de sdqs_num, ya validado abajo).
            # Si vienen vacías en el Excel, _v() las convierte a None y el
            # INSERT viola la restricción NOT NULL — con "OR IGNORE" eso
            # significa perder el registro completo en silencio (confirmado
            # con datos reales: 5 registros con quejoso/fecha_asignacion
            # vacíos desaparecían en cada ciclo de respaldo-restauración).
            # Se coacciona a '' igual que hace el importador propio de SDQS
            # (app/routers/sdqs.py), que sí tolera estos campos vacíos.
            _idx_not_null_sdqs = {0, 1, 7, 8}  # mes, fecha_asignacion, tema, competencia_ocdi
            _idx_quejoso = 5
            conn.execute("DELETE FROM sdqs")
            sdqs_omitidos = 0
            for row in ws5.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                sdqs_num = _v(row[2]) if len(row) > 2 else None
                if not sdqs_num:
                    continue
                vals = [_v(row[i]) if i < len(row) else None for i in range(len(campos_sdqs))]
                for idx in _idx_not_null_sdqs | {_idx_quejoso}:
                    if vals[idx] is None:
                        vals[idx] = ""
                cur = conn.execute(
                    f"INSERT OR IGNORE INTO sdqs ({', '.join(campos_sdqs)}) VALUES ({', '.join(['?']*len(campos_sdqs))})",
                    vals,
                )
                if cur.rowcount:
                    stats["sdqs"] += 1
                else:
                    sdqs_omitidos += 1
            stats["sdqs_omitidos"] = sdqs_omitidos

        # ── Hoja 6: Correspondencia ────────────────────────────────────────────
        if "Correspondencia" in wb.sheetnames:
            ws6 = wb["Correspondencia"]
            conn.execute("DELETE FROM correspondencia_radicados_salida")
            conn.execute("DELETE FROM correspondencia")
            for row in ws6.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                n_rad = _v(row[3]) if len(row) > 3 else None
                if not n_rad:
                    continue
                termino = None
                termino_raw = _v(row[9]) if len(row) > 9 else None
                if termino_raw:
                    try:
                        termino = int(float(termino_raw))
                    except (ValueError, TypeError):
                        pass
                cur = conn.execute("""
                    INSERT INTO correspondencia
                    (anio, mes, fecha_ingreso, n_radicado, origen, correo_remitente, asunto,
                     sinproc_personeria, tipo_requerimiento, termino_dias, tipo_documento,
                     responsable, caso_bmp, fecha_radicado_salida, tipo_respuesta, tramite_salida,
                     created_at, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, [
                    _v(row[0]), _v(row[1]),
                    _v(row[2]), n_rad, _v(row[4]), _v(row[5]), _v(row[6]),
                    _v(row[7]), _v(row[8]), termino, _v(row[10]),
                    _v(row[11]), _v(row[12]),
                    _v(row[15]),   # fecha_radicado_salida
                    _v(row[16]),   # tipo_respuesta
                    _v(row[17]),   # tramite_salida / observaciones
                    _v(row[18]),   # created_at
                    _v(row[19]),   # updated_at
                ])
                corr_id = cur.lastrowid
                radicados_str = _v(row[13]) or ""
                urls_str = _v(row[14]) or ""
                if radicados_str:
                    rads = [r.strip() for r in radicados_str.split(" | ") if r.strip()]
                    urls = [u.strip() for u in urls_str.split(" | ")] if urls_str else []
                    for i, rad in enumerate(rads):
                        u = urls[i] if i < len(urls) and urls[i] else None
                        conn.execute(
                            "INSERT INTO correspondencia_radicados_salida (correspondencia_id, radicado, url) VALUES (?,?,?)",
                            (corr_id, rad, u),
                        )
                stats["corr"] += 1

        # ── Hoja 7: Seguimiento Mensual ────────────────────────────────────────
        if "Seguimiento Mensual" in wb.sheetnames:
            ws7 = wb["Seguimiento Mensual"]
            # Detectar si el archivo tiene la columna "AÑO EXPEDIENTE" (formato
            # nuevo) leyendo el encabezado real, para seguir aceptando backups
            # exportados antes de este fix sin romper la importación.
            header_row7 = [str(c.value or "").strip().upper() for c in next(ws7.iter_rows(min_row=1, max_row=1))]
            tiene_anio_exp = "AÑO EXPEDIENTE" in header_row7
            i_desc = 5 if tiene_anio_exp else 4
            i_creador = 6 if tiene_anio_exp else 5

            conn.execute("DELETE FROM seguimiento_mensual")
            for row in ws7.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                n_exp = _v(row[2]) if len(row) > 2 else None
                descripcion = _v(row[i_desc]) if len(row) > i_desc else None
                if not n_exp or not descripcion:
                    continue
                if tiene_anio_exp:
                    exp_anio = _v(row[3]) if len(row) > 3 else None
                    # Clave correcta: (n_expediente, anio) — los números de
                    # expediente se reinician cada año (ver H2/H11), así que
                    # buscar solo por n_expediente puede enlazar al expediente
                    # equivocado si el número se repite en otro año.
                    exp_row = conn.execute(
                        "SELECT id FROM expedientes WHERE n_expediente = ? AND anio = ?",
                        (n_exp, exp_anio),
                    ).fetchone()
                else:
                    # Backup exportado antes de este fix: no incluye el año del
                    # expediente — se mantiene el comportamiento anterior (con
                    # el mismo riesgo de ambigüedad si el número se repite).
                    exp_row = conn.execute(
                        "SELECT id FROM expedientes WHERE n_expediente = ?", (n_exp,)
                    ).fetchone()
                if not exp_row:
                    continue
                conn.execute("""
                    INSERT OR IGNORE INTO seguimiento_mensual (expediente_id, anio, mes, descripcion, created_by)
                    VALUES (?,?,?,?,?)
                """, [
                    exp_row[0],
                    _v(row[0]),
                    _v(row[1]),
                    descripcion,
                    _v(row[i_creador]) if len(row) > i_creador else None,
                ])
                stats["seg"] += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return RedirectResponse("/backup/?msg=error_import", status_code=303)

    conn.close()
    msg = f"ok_{stats['base']}_{stats['digitales']}_{stats['coms']}_{stats['sala']}_{stats['autos']}_{stats['sdqs']}_{stats['corr']}_{stats['seg']}_{stats['sdqs_omitidos']}"
    return RedirectResponse(f"/backup/?msg={msg}", status_code=303)


# ── Backup ZIP completo (4 carpetas, 4 Excel) ─────────────────────────────────

@router.get("/zip")
async def backup_zip():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/backup/?msg=error_openpyxl")

    hoy = date.today().strftime("%Y%m%d")
    conn = get_db()

    # ── Cargar todos los datos ────────────────────────────────────────────────
    exps      = conn.execute("SELECT * FROM expedientes ORDER BY anio, n_expediente").fetchall()
    dig_exps  = conn.execute("SELECT * FROM exp_digitales ORDER BY anio DESC, n_expediente ASC").fetchall()
    dig_coms  = conn.execute(
        "SELECT * FROM exp_comunicaciones ORDER BY exp_digital_id ASC, fecha_envio ASC, id ASC"
    ).fetchall()
    dig_revs  = conn.execute(
        "SELECT exp_digital_id, MAX(fecha_revision) AS ultima_revision FROM exp_revisiones GROUP BY exp_digital_id"
    ).fetchall()
    sala         = conn.execute("SELECT * FROM sala_agenda ORDER BY fecha, franja").fetchall()
    ctrl_autos_z = conn.execute(
        "SELECT * FROM control_autos_sustanciacion ORDER BY fecha_auto ASC, id ASC"
    ).fetchall()
    corr_rows_raw = conn.execute("""
        SELECT c.*,
               GROUP_CONCAT(rs.radicado, ' | ') AS radicados_salida,
               GROUP_CONCAT(COALESCE(rs.url, ''), ' | ') AS radicados_urls
        FROM correspondencia c
        LEFT JOIN correspondencia_radicados_salida rs ON rs.correspondencia_id = c.id
        GROUP BY c.id
        ORDER BY c.fecha_ingreso DESC
    """).fetchall()
    corr_rows = [_calcular_semaforo_row(dict(r)) for r in corr_rows_raw]
    sdqs_rows    = conn.execute("SELECT * FROM sdqs ORDER BY fecha_asignacion, id").fetchall()
    seg_rows     = conn.execute("""
        SELECT s.anio, s.mes, e.n_expediente, e.anio AS exp_anio, e.abogado_asignado,
               s.descripcion, s.created_by, s.created_at
        FROM seguimiento_mensual s
        JOIN expedientes e ON e.id = s.expediente_id
        ORDER BY e.anio DESC, CAST(e.n_expediente AS INTEGER), s.anio DESC,
                 CASE s.mes
                   WHEN 'ENERO' THEN 1 WHEN 'FEBRERO' THEN 2 WHEN 'MARZO' THEN 3
                   WHEN 'ABRIL' THEN 4 WHEN 'MAYO' THEN 5 WHEN 'JUNIO' THEN 6
                   WHEN 'JULIO' THEN 7 WHEN 'AGOSTO' THEN 8 WHEN 'SEPTIEMBRE' THEN 9
                   WHEN 'OCTUBRE' THEN 10 WHEN 'NOVIEMBRE' THEN 11 WHEN 'DICIEMBRE' THEN 12
                   ELSE 13 END
    """).fetchall()
    conn.close()

    ultima_rev = {r["exp_digital_id"]: r["ultima_revision"] for r in dig_revs}
    coms_por_exp: dict[int, list] = {}
    for c in dig_coms:
        coms_por_exp.setdefault(c["exp_digital_id"], []).append(dict(c))

    h_font   = Font(bold=True, color="FFFFFF", size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="EBF1F8")
    sub_fill = PatternFill("solid", fgColor="dbeafe")

    def make_wb_base() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Base Expedientes"
        fill = PatternFill("solid", fgColor="1B4F8A")
        headers = [
            "N. EXPEDIENTE","AÑO","MES","MEDIO DE INGRESO","N. RADICADO",
            "FECHA RADICADO","ABOGADO ASIGNADO","ENTIDAD ORIGEN","QUEJOSO","ASUNTO",
            "IMPEDIMENTO","FECHA APERTURA EXPEDIENTE","N. AUTO APERTURA IND.",
            "FECHA AUTO APERTURA IND.","TIPO EXPEDIENTE","TIPOLOGÍA",
            "RELACIONADO SINIESTRO","RESPONSABLE SINIESTRO","RELACIONADO MALTRATO",
            "RELACIONADO CORRUPCIÓN","VALORES INSTITUCIONALES",
            "FECHA HECHOS (OBS)","FECHA HECHOS","F. ÚLTIMA ACT. INDAGACIÓN",
            "N. AUTO ÚLTIMA ACT. IND.","F. APERTURA INVESTIGACIÓN","N. AUTO APERTURA INV.",
            "NOMBRE INVESTIGADO","CÉDULA","PERFIL INVESTIGADO","ÁREA ORIGEN INVESTIGADO",
            "FECHA PRÓRROGA","N. AUTO PRÓRROGA","TIEMPO PRÓRROGA",
            "F. ÚLTIMA ACT. INVESTIGACIÓN","N. AUTO ÚLTIMA ACT. INV.",
            "N. AUTO TRASLADO","F. AUTO TRASLADO","N. AUTO ACUMULACIÓN","F. AUTO ACUMULACIÓN",
            "EXPEDIENTE ACUMULA","F. AUTO ARCHIVO","N. AUTO ARCHIVO",
            "F. AUTO PLIEGO CARGOS","N. AUTO PLIEGO CARGOS",
            "ETAPA ACTUAL","ESTADO DEL PROCESO","OBSERVACIONES",
            "CREADO POR","FECHA CREACIÓN","ÚLTIMA ACTUALIZACIÓN",
        ]
        campos = [
            "n_expediente","anio","mes","medio_ingreso","n_radicado",
            "fecha_radicado","abogado_asignado","entidad_origen","quejoso","asunto",
            "impedimento","fecha_apertura_expediente","numero_auto_apertura_ind",
            "fecha_auto_apertura_ind","tipo_expediente","tipologia",
            "relacionado_siniestro","responsable_siniestro","relacionado_maltrato",
            "relacionado_corrupcion","valores_institucionales",
            "fecha_hechos_obs","fecha_hechos","fecha_ultima_act_indagacion",
            "numero_auto_ultima_act_ind","fecha_apertura_investigacion","numero_auto_apertura_inv",
            "nombre_investigado","cedula","perfil_investigado","area_origen_investigado",
            "fecha_prorroga","numero_auto_prorroga","tiempo_prorroga",
            "fecha_ultima_act_investigacion","numero_auto_ultima_act_inv",
            "numero_auto_traslado","fecha_auto_traslado","numero_auto_acumulacion","fecha_auto_acumulacion",
            "expediente_acumula","fecha_auto_archivo","numero_auto_archivo",
            "fecha_auto_pliego_cargos","numero_auto_pliego_cargos",
            "etapa_actual","estado_proceso","observaciones",
            "created_by","created_at","updated_at",
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = fill; cell.font = h_font; cell.alignment = center
        ws.row_dimensions[1].height = 40
        for ri, row in enumerate(exps, 2):
            d = dict(row)
            rf = alt_fill if ri % 2 == 0 else None
            for ci, campo in enumerate(campos, 1):
                cell = ws.cell(row=ri, column=ci, value=d.get(campo))
                cell.alignment = Alignment(vertical="center")
                if rf: cell.fill = rf
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)
        ws.freeze_panes = "A2"

        # Hoja 2: Seguimiento Mensual (pertenece a Base Expedientes)
        ws_seg = wb.create_sheet("Seguimiento Mensual")
        fill_seg = PatternFill("solid", fgColor="0D3060")
        seg_headers = ["AÑO", "MES", "N. EXPEDIENTE", "AÑO EXPEDIENTE", "ABOGADO", "DESCRIPCIÓN ACTUACIÓN", "REGISTRADO POR", "FECHA REGISTRO"]
        for ci, h in enumerate(seg_headers, 1):
            cell = ws_seg.cell(row=1, column=ci, value=h)
            cell.fill = fill_seg; cell.font = h_font; cell.alignment = center
        ws_seg.row_dimensions[1].height = 30
        for ri, row in enumerate(seg_rows, 2):
            d = dict(row)
            rf = alt_fill if ri % 2 == 0 else None
            vals = [d.get("anio"), d.get("mes"), d.get("n_expediente"), d.get("exp_anio"),
                    d.get("abogado_asignado"), d.get("descripcion"),
                    d.get("created_by"), d.get("created_at")]
            for ci, v in enumerate(vals, 1):
                cell = ws_seg.cell(row=ri, column=ci, value=v)
                cell.alignment = Alignment(vertical="center", wrap_text=(ci == 6))
                if rf: cell.fill = rf
        for i, w in enumerate([10, 14, 16, 12, 28, 60, 24, 20], 1):
            ws_seg.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws_seg.freeze_panes = "A2"

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_wb_sdqs() -> io.BytesIO:
        from app.routers.sdqs import _calcular_semaforo_sdqs
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SDQS"
        fill = PatternFill("solid", fgColor="7B3F00")
        link_font_sdqs = Font(color="0563C1", underline="single", size=10)
        headers = [
            "MES","FECHA ASIGNACION","SDQS","URL SDQS","FECHA VENCIMIENTO","ESTADO DIAS",
            "QUEJOSO","CORREO","TEMA",
            "COMPETENCIA OCDI","BPM","RESPONSABLE","RAD SALIDA","URL RAD SALIDA","FECHA RESPUESTA",
            "OBSERVACIONES","ESTADO PROCESO","HECHO CORRUPTO","VALOR INSTITUCIONAL",
            "TIPOLOGIA","CREADO POR","FECHA CREACIÓN","ÚLTIMA ACTUALIZACIÓN",
        ]
        campos = [
            "mes","fecha_asignacion","sdqs","url_sdqs","fecha_vencimiento","estado_dias",
            "quejoso","correo","tema",
            "competencia_ocdi","bpm","responsable","rad_salida","url_rad_salida","fecha_respuesta",
            "observaciones","estado_proceso","hecho_corrupto","valor_institucional",
            "tipologia","created_by","created_at","updated_at",
        ]
        SEM_COLORS = {"verde": "D4EDDA", "amarillo": "FFF3CD", "rojo": "F8D7DA"}
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = fill; cell.font = h_font; cell.alignment = center
        ws.row_dimensions[1].height = 30
        for ri, row in enumerate(sdqs_rows, 2):
            d = _calcular_semaforo_sdqs(dict(row))
            rf = alt_fill if ri % 2 == 0 else None
            for ci, campo in enumerate(campos, 1):
                cell = ws.cell(row=ri, column=ci, value=d.get(campo))
                cell.alignment = Alignment(vertical="center", wrap_text=(ci == 9))
                if rf: cell.fill = rf
            sem = d.get("semaforo_sdqs")
            if sem and sem in SEM_COLORS:
                ws.cell(ri, 6).fill = PatternFill("solid", fgColor=SEM_COLORS[sem])
            url_sdqs_val = d.get("url_sdqs")
            if url_sdqs_val and d.get("sdqs"):
                c = ws.cell(ri, 3)
                c.hyperlink = url_sdqs_val
                c.font = link_font_sdqs
            url = d.get("url_rad_salida")
            if url and d.get("rad_salida"):
                c = ws.cell(ri, 13)
                c.hyperlink = url
                c.font = link_font_sdqs
        col_widths = [10, 16, 14, 40, 16, 12, 28, 28, 50, 14, 14, 28, 16, 40, 16, 40, 22, 22, 22, 20, 20, 20, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_wb_correspondencia() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CORRESPONDENCIA"
        fill = PatternFill("solid", fgColor="1B4F8A")
        link_font = Font(color="0563C1", underline="single", size=10)
        headers = [
            "AÑO", "MES", "FECHA INGRESO DE OFICIO", "N. RADICADOS",
            "ENTIDAD", "CORREO REMITENTE", "ASUNTO", "NUMERO SINPROC PERSONERIA",
            "TIPO DE REQUERIMIENTO", "TERMINO (DIAS)", "TIPO DE DOCUMENTO",
            "RESPONSABLE", "CASO BMP", "N RADICADO SALIDA",
            "FECHA RADICADO DE SALIDA", "TIPO DE RESPUESTA", "TRÁMITE DE SALIDA",
            "FECHA DE VENCIMIENTO LEGAL",
            "FECHA REVISIÓN SUGERIDA (−2 días hábiles)",
            "DÍAS TRANSCURRIDOS",
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = fill; cell.font = h_font; cell.alignment = center
        ws.row_dimensions[1].height = 36
        for ri, d in enumerate(corr_rows, 2):
            rf = alt_fill if ri % 2 == 0 else None
            urls_str = d.get("radicados_urls") or ""
            urls_list = [u.strip() for u in urls_str.split(" | ")] if urls_str.strip() else []
            first_url = next((u for u in urls_list if u), None)
            vals = [
                d.get("anio"), d.get("mes"),
                d.get("fecha_ingreso")[:10] if d.get("fecha_ingreso") else None,
                d.get("n_radicado"), d.get("origen"), d.get("correo_remitente"), d.get("asunto"),
                d.get("sinproc_personeria"), d.get("tipo_requerimiento"), d.get("termino_dias"),
                d.get("tipo_documento"), d.get("responsable"), d.get("caso_bmp"),
                d.get("radicados_salida"),           # col 14 — N RADICADO SALIDA
                d.get("fecha_radicado_salida")[:10] if d.get("fecha_radicado_salida") else None,
                d.get("tipo_respuesta"), d.get("tramite_salida"),
                d.get("fecha_vencimiento"),           # col 18 — plazo legal real
                d.get("fecha_termino_respuesta"),     # col 19 — fecha revisión sugerida
                d.get("dias_transcurridos"),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = Alignment(vertical="center", wrap_text=ci in (5, 7))
                if rf: cell.fill = rf
            if first_url and d.get("radicados_salida"):
                rad_cell = ws.cell(row=ri, column=14)
                rad_cell.hyperlink = first_url
                rad_cell.font = link_font
        col_widths = [6, 12, 20, 18, 30, 30, 40, 20, 40, 10, 18, 28, 10, 22, 20, 25, 30, 20, 28, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_wb_digitales() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EXP DIGIT 2025-2026"
        fill = PatternFill("solid", fgColor="1e3a5f")
        headers = [
            "N° Expediente","Año","Abogado","Etapa","Queja Inicial",
            "Radicado Auto","Nombre Auto","Fecha Auto",
            "Obs. Generales","Última Revisión",
            "Radicado Comunicación","Dependencia","Fecha Envío",
            "Fecha Seguimiento","Radicado Respuesta","Fecha Respuesta",
            "Responsable","Observaciones",
        ]
        ws.append(["SEGUIMIENTO EXPEDIENTES DIGITALES"])
        ws.append(headers)
        for cell in ws[2]:
            cell.font = h_font; cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for exp in dig_exps:
            ed = dict(exp)
            exp_coms = coms_por_exp.get(ed["id"], [])
            primera = exp_coms[0] if exp_coms else {}
            ws.append([
                ed.get("n_expediente"), ed.get("anio"), ed.get("abogado"), ed.get("etapa"),
                ed.get("queja_inicial"), ed.get("radicado_auto"), ed.get("nombre_auto"), ed.get("fecha_auto"),
                ed.get("observaciones"), ultima_rev.get(ed["id"]),
                primera.get("radicado_comunicacion"), primera.get("dependencia"),
                primera.get("fecha_envio"), primera.get("fecha_seguimiento"),
                primera.get("radicado_respuesta"), primera.get("fecha_respuesta"),
                primera.get("responsable"), primera.get("observaciones"),
            ])
            for com in exp_coms[1:]:
                sr = [None]*10 + [
                    com.get("radicado_comunicacion"), com.get("dependencia"),
                    com.get("fecha_envio"), com.get("fecha_seguimiento"),
                    com.get("radicado_respuesta"), com.get("fecha_respuesta"),
                    com.get("responsable"), com.get("observaciones"),
                ]
                ws.append(sr)
                for cell in ws[ws.max_row]: cell.fill = sub_fill
        col_widths = [15, 6, 22, 20, 14, 20, 30, 14, 40, 22, 22, 25, 14, 16, 22, 14, 20, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_wb_sala() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sala de Audiencias"
        fill = PatternFill("solid", fgColor="065F46")
        headers = ["Fecha","Franja","Título","Descripción","Estado","Responsable","Fecha Creación"]
        campos  = ["fecha","franja","titulo","descripcion","estado","responsable","created_at"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = fill; cell.font = h_font; cell.alignment = center
        ws.row_dimensions[1].height = 30
        for ri, row in enumerate(sala, 2):
            d = dict(row)
            rf = alt_fill if ri % 2 == 0 else None
            for ci, campo in enumerate(campos, 1):
                cell = ws.cell(row=ri, column=ci, value=d.get(campo))
                cell.alignment = Alignment(vertical="center")
                if rf: cell.fill = rf
        col_widths = [14, 16, 30, 40, 12, 25, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    def make_wb_control_autos() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Control Autos"
        fill = PatternFill("solid", fgColor="2E7D32")
        headers = ["EXPEDIENTE", "NÚMERO DEL AUTO", "FECHA DEL AUTO", "ASUNTO AUTO", "ABOGADO RESPONSABLE", "OBSERVACIONES",
                   "CREADO POR", "FECHA CREACIÓN", "ÚLTIMA ACTUALIZACIÓN"]
        campos  = ["expediente", "numero_auto", "fecha_auto", "asunto_auto", "abogado_responsable", "observaciones",
                   "created_by", "created_at", "updated_at"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = fill; cell.font = h_font; cell.alignment = center
        ws.row_dimensions[1].height = 30
        alt_autos = PatternFill("solid", fgColor="F1F8E9")
        for ri, row in enumerate(ctrl_autos_z, 2):
            d = dict(row)
            rf = alt_autos if ri % 2 == 0 else None
            for ci, campo in enumerate(campos, 1):
                cell = ws.cell(row=ri, column=ci, value=d.get(campo))
                cell.alignment = Alignment(vertical="center")
                if rf: cell.fill = rf
        col_widths = [18, 16, 16, 48, 22, 30, 20, 20, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    # ── Construir ZIP ─────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            f"OCDI/01_Base_Expedientes/Base_Expedientes_{hoy}.xlsx",
            make_wb_base().read(),
        )
        zf.writestr(
            f"OCDI/02_Lista_Reparto_Abogados/Correspondencia_{hoy}.xlsx",
            make_wb_correspondencia().read(),
        )
        zf.writestr(
            f"OCDI/03_Control_Autos_Sustanciacion/SDS-CDO-FT-001_Control_Autos_{hoy}.xlsx",
            make_wb_control_autos().read(),
        )
        zf.writestr(
            f"OCDI/04_SDQS/SDQS_{hoy}.xlsx",
            make_wb_sdqs().read(),
        )
        zf.writestr(
            f"OCDI/05_Expedientes_Digitales/Exp_Digitales_{hoy}.xlsx",
            make_wb_digitales().read(),
        )
        zf.writestr(
            f"OCDI/06_Sala_Audiencias/Sala_Audiencias_{hoy}.xlsx",
            make_wb_sala().read(),
        )
    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=OCDI_Backup_Completo_{hoy}.zip"},
    )
