from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from datetime import date
import io

from urllib.parse import quote_plus as _quote_plus
from app.database import get_db
from app.auth_utils import puede_escribir as _pw, registrar_log

_MOD = "digitales"

router = APIRouter(prefix="/digitales")
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))
templates.env.filters["quote_plus"] = _quote_plus


# ── Helpers ────────────────────────────────────────────────────────────────────

def _texto(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "#VALUE!", "#N/A", "#REF!"):
        return None
    return s


def _clase_alerta(dias) -> str | None:
    """Retorna clase CSS de alerta según días transcurridos sin respuesta."""
    if dias is None:
        return None
    try:
        d = int(dias)
    except (TypeError, ValueError):
        return None
    if d >= 14:
        return "roja"
    if d >= 13:
        return "amarilla"
    if d >= 8:
        return "azul"
    return None


def _fecha(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s in ("", "nan", "None", "#VALUE!", "#N/A"):
        return None
    if len(s) == 10 and s[4] == "-":
        return s
    if len(s) == 10 and s[2] == "/":
        parts = s.split("/")
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return None


# ── Lista ──────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def lista(
    request: Request,
    q: str = "",
    abogado: str = "",
    etapa: str = "",
    anio: str = "",
    sin_respuesta: str = "",
    queja: str = "",
    alerta: str = "",
    msg: str = "",
    page: int = 1,
    por_pagina: int = 20,
):
    conn = get_db()

    filtros = ["1=1"]
    params: list = []

    if q.strip():
        filtros.append("(e.n_expediente LIKE ? OR e.abogado LIKE ? OR e.radicado_auto LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if abogado.strip():
        filtros.append("e.abogado = ?")
        params.append(abogado.strip())
    if etapa.strip():
        filtros.append("e.etapa = ?")
        params.append(etapa.strip())
    if anio.strip():
        filtros.append("e.anio = ?")
        params.append(int(anio.strip()))
    if sin_respuesta == "1":
        filtros.append("""e.id IN (
            SELECT DISTINCT exp_digital_id FROM exp_comunicaciones
            WHERE fecha_respuesta IS NULL OR fecha_respuesta = ''
        )""")
    if queja == "si":
        filtros.append("(e.queja_inicial = 'Sí' OR e.queja_inicial = 'Si' OR e.queja_inicial = 'SI' OR e.queja_inicial = 'sí')")
    if alerta == "roja":
        filtros.append("""e.id IN (
            SELECT DISTINCT exp_digital_id FROM exp_comunicaciones
            WHERE (fecha_respuesta IS NULL OR fecha_respuesta = '')
            AND fecha_envio IS NOT NULL AND fecha_envio != ''
            AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER) >= 14)""")
    elif alerta == "amarilla":
        filtros.append("""e.id IN (
            SELECT DISTINCT exp_digital_id FROM exp_comunicaciones
            WHERE (fecha_respuesta IS NULL OR fecha_respuesta = '')
            AND fecha_envio IS NOT NULL AND fecha_envio != ''
            AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER) = 13)""")
    elif alerta == "azul":
        filtros.append("""e.id IN (
            SELECT DISTINCT exp_digital_id FROM exp_comunicaciones
            WHERE (fecha_respuesta IS NULL OR fecha_respuesta = '')
            AND fecha_envio IS NOT NULL AND fecha_envio != ''
            AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER) >= 8
            AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER) < 13)""")

    where = " AND ".join(filtros)

    total = conn.execute(f"SELECT COUNT(*) FROM exp_digitales e WHERE {where}", params).fetchone()[0]
    offset = (page - 1) * por_pagina
    rows = conn.execute(
        f"""SELECT e.*,
            (SELECT COUNT(*) FROM exp_comunicaciones WHERE exp_digital_id = e.id) AS num_coms,
            (SELECT COUNT(*) FROM exp_comunicaciones
             WHERE exp_digital_id = e.id AND (fecha_respuesta IS NULL OR fecha_respuesta = '')) AS coms_sin_resp,
            (SELECT MAX(CAST(julianday('now') - julianday(fecha_envio) AS INTEGER))
             FROM exp_comunicaciones
             WHERE exp_digital_id = e.id
             AND (fecha_respuesta IS NULL OR fecha_respuesta = '')
             AND fecha_envio IS NOT NULL AND fecha_envio != '') AS max_dias_pendiente,
            (SELECT MAX(fecha_revision) FROM exp_revisiones WHERE exp_digital_id = e.id) AS ultima_revision
            FROM exp_digitales e
            WHERE {where}
            ORDER BY e.anio DESC, CAST(e.n_expediente AS INTEGER) ASC, e.n_expediente ASC LIMIT ? OFFSET ?""",
        params + [por_pagina, offset],
    ).fetchall()

    abogados = [r[0] for r in conn.execute(
        "SELECT DISTINCT abogado FROM exp_digitales WHERE abogado IS NOT NULL ORDER BY abogado"
    ).fetchall()]
    etapas_list = [r[0] for r in conn.execute(
        "SELECT DISTINCT etapa FROM exp_digitales WHERE etapa IS NOT NULL ORDER BY etapa"
    ).fetchall()]
    anios = [r[0] for r in conn.execute(
        "SELECT DISTINCT anio FROM exp_digitales WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()]

    conn.close()

    total_pages = max(1, (total + por_pagina - 1) // por_pagina)

    return templates.TemplateResponse("digitales_lista.html", {
        "request": request,
        "active": "digitales_lista",
        "rows": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "total_pages": total_pages,
        "por_pagina": por_pagina,
        "q": q,
        "abogado": abogado,
        "etapa": etapa,
        "anio": anio,
        "sin_respuesta": sin_respuesta,
        "queja": queja,
        "alerta": alerta,
        "abogados": abogados,
        "etapas": etapas_list,
        "anios": anios,
        "msg": msg,
        "clase_alerta": _clase_alerta,
        "back_url": request.url.path + ("?" + str(request.url.query) if request.url.query else ""),
    })


# ── Dashboard ──────────────────────────────────────────────────────────────────

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    conn = get_db()

    total = conn.execute("SELECT COUNT(*) FROM exp_digitales").fetchone()[0]

    por_etapa = conn.execute("""
        SELECT etapa, COUNT(*) cant FROM exp_digitales
        WHERE etapa IS NOT NULL GROUP BY etapa ORDER BY cant DESC
    """).fetchall()

    por_abogado = conn.execute("""
        SELECT abogado, COUNT(*) cant FROM exp_digitales
        WHERE abogado IS NOT NULL GROUP BY abogado ORDER BY cant DESC
    """).fetchall()

    sin_respuesta = conn.execute("""
        SELECT COUNT(*) FROM exp_comunicaciones
        WHERE fecha_respuesta IS NULL OR fecha_respuesta = ''
    """).fetchone()[0]

    total_coms = conn.execute("SELECT COUNT(*) FROM exp_comunicaciones").fetchone()[0]

    queja_si = conn.execute(
        "SELECT COUNT(*) FROM exp_digitales WHERE queja_inicial = 'Sí' OR queja_inicial = 'Si' OR queja_inicial = 'SI'"
    ).fetchone()[0]

    por_anio = conn.execute("""
        SELECT anio, COUNT(*) cant FROM exp_digitales
        WHERE anio IS NOT NULL GROUP BY anio ORDER BY anio DESC
    """).fetchall()

    _dias_base = """
        (fecha_respuesta IS NULL OR fecha_respuesta = '')
        AND fecha_envio IS NOT NULL AND fecha_envio != ''
        AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER)
    """
    alerta_azul = conn.execute(f"""
        SELECT COUNT(*) FROM exp_comunicaciones
        WHERE {_dias_base} >= 8 AND CAST(julianday('now') - julianday(fecha_envio) AS INTEGER) < 13
    """).fetchone()[0]
    alerta_amarilla = conn.execute(f"""
        SELECT COUNT(*) FROM exp_comunicaciones
        WHERE {_dias_base} = 13
    """).fetchone()[0]
    alerta_roja = conn.execute(f"""
        SELECT COUNT(*) FROM exp_comunicaciones
        WHERE {_dias_base} >= 14
    """).fetchone()[0]

    conn.close()

    return templates.TemplateResponse("digitales_dashboard.html", {
        "request": request,
        "active": "digitales_dash",
        "total": total,
        "por_etapa": [dict(r) for r in por_etapa],
        "por_abogado": [dict(r) for r in por_abogado],
        "sin_respuesta": sin_respuesta,
        "total_coms": total_coms,
        "queja_si": queja_si,
        "por_anio": [dict(r) for r in por_anio],
        "alerta_azul": alerta_azul,
        "alerta_amarilla": alerta_amarilla,
        "alerta_roja": alerta_roja,
    })


# ── Nuevo ──────────────────────────────────────────────────────────────────────

def _get_abogados(conn) -> list[str]:
    """Retorna lista de nombres desde el catálogo, sincronizando primero con exp_digitales."""
    conn.execute("""
        INSERT OR IGNORE INTO abogados_digitales (nombre)
        SELECT DISTINCT abogado FROM exp_digitales
        WHERE abogado IS NOT NULL AND TRIM(abogado) != ''
    """)
    conn.commit()
    return [r[0] for r in conn.execute(
        "SELECT nombre FROM abogados_digitales ORDER BY nombre"
    ).fetchall()]


@router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_form(request: Request):
    conn = get_db()
    abogados = _get_abogados(conn)
    conn.close()
    return templates.TemplateResponse("digitales_form.html", {
        "request": request,
        "active": "digitales_lista",
        "exp": {},
        "comunicaciones": [],
        "modo": "nuevo",
        "abogados": abogados,
    })


@router.post("/nuevo")
async def nuevo_post(
    request: Request,
    n_expediente: str = Form(""),
    anio: str = Form(""),
    abogado: str = Form(""),
    etapa: str = Form(""),
    queja_inicial: str = Form("No"),
    radicado_auto: str = Form(""),
    nombre_auto: str = Form(""),
    fecha_auto: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("""
        INSERT INTO exp_digitales (n_expediente, anio, abogado, etapa, queja_inicial,
            radicado_auto, nombre_auto, fecha_auto)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        _texto(n_expediente), int(anio) if anio.strip() else None,
        _texto(abogado), _texto(etapa), queja_inicial or "No",
        _texto(radicado_auto), _texto(nombre_auto), _fecha(fecha_auto),
    ))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return RedirectResponse(f"/digitales/{new_id}?msg=creado", status_code=303)


# ── Importar Excel  ← DEBE IR ANTES QUE /{exp_id} ────────────────────────────

@router.get("/importar", response_class=HTMLResponse)
async def importar_form(request: Request, msg: str = ""):
    return templates.TemplateResponse("digitales_importar.html", {
        "request": request,
        "active": "digitales_importar",
        "msg": msg,
        "resultado": None,
    })


@router.post("/importar", response_class=HTMLResponse)
async def importar_post(request: Request, archivo: UploadFile = File(...)):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/importar?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return templates.TemplateResponse("digitales_importar.html", {
            "request": request,
            "active": "digitales_importar",
            "msg": "error_openpyxl",
            "resultado": None,
        })

    contenido = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

    # Seleccionar hoja
    hoja = None
    for nombre in wb.sheetnames:
        if "EXP DIGIT" in nombre.upper() or "EXPEDIENTE" in nombre.upper() or "DIGIT" in nombre.upper():
            hoja = wb[nombre]
            break
    if hoja is None:
        hoja = wb.active

    conn = get_db()
    exp_insertados = 0
    exp_omitidos = 0
    coms_insertadas = 0
    ultimo_exp_id: int | None = None

    for idx, row in enumerate(hoja.iter_rows(values_only=True)):
        if idx < 2:  # fila 0 = título, fila 1 = cabeceras
            continue

        col = list(row)
        while len(col) < 16:
            col.append(None)

        n_exp = _texto(col[0])
        radicado_com = _texto(col[8])

        if n_exp:
            # Fila padre — expediente
            anio_val = col[1]
            try:
                anio_int = int(anio_val) if anio_val else None
            except (ValueError, TypeError):
                anio_int = None

            existing = conn.execute(
                "SELECT id FROM exp_digitales WHERE n_expediente = ? AND anio = ?",
                (n_exp, anio_int)
            ).fetchone()

            if existing:
                ultimo_exp_id = existing[0]
                exp_omitidos += 1
            else:
                conn.execute("""
                    INSERT INTO exp_digitales
                        (n_expediente, anio, abogado, etapa, queja_inicial,
                         radicado_auto, nombre_auto, fecha_auto)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    n_exp, anio_int,
                    _texto(col[2]), _texto(col[3]),
                    _texto(col[4]) or "No",
                    _texto(col[5]), _texto(col[6]), _fecha(col[7]),
                ))
                ultimo_exp_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                exp_insertados += 1

            # Si la fila padre también tiene comunicación
            radicado_com_padre = _texto(col[8])
            if radicado_com_padre and ultimo_exp_id:
                conn.execute("""
                    INSERT INTO exp_comunicaciones
                        (exp_digital_id, radicado_comunicacion, dependencia, fecha_envio,
                         fecha_seguimiento, radicado_respuesta, fecha_respuesta, responsable, observaciones)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    ultimo_exp_id,
                    radicado_com_padre,
                    _texto(col[9]), _fecha(col[10]), _fecha(col[11]),
                    _texto(col[12]), _fecha(col[13]), _texto(col[14]), _texto(col[15]),
                ))
                coms_insertadas += 1

        elif radicado_com and ultimo_exp_id:
            # Fila hija — comunicación
            conn.execute("""
                INSERT INTO exp_comunicaciones
                    (exp_digital_id, radicado_comunicacion, dependencia, fecha_envio,
                     fecha_seguimiento, radicado_respuesta, fecha_respuesta, responsable, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                ultimo_exp_id,
                radicado_com,
                _texto(col[9]), _fecha(col[10]), _fecha(col[11]),
                _texto(col[12]), _fecha(col[13]), _texto(col[14]), _texto(col[15]),
            ))
            coms_insertadas += 1

    conn.commit()
    conn.close()

    resultado = {
        "exp_insertados": exp_insertados,
        "exp_omitidos": exp_omitidos,
        "coms_insertadas": coms_insertadas,
    }

    return templates.TemplateResponse("digitales_importar.html", {
        "request": request,
        "active": "digitales_importar",
        "msg": "ok",
        "resultado": resultado,
    })


# ── Exportar Excel  ← DEBE IR ANTES QUE /{exp_id} ────────────────────────────

@router.get("/exportar")
async def exportar():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/digitales/?msg=error_openpyxl")

    conn = get_db()
    exps = conn.execute("SELECT * FROM exp_digitales ORDER BY anio DESC, n_expediente ASC").fetchall()
    coms = conn.execute(
        "SELECT * FROM exp_comunicaciones ORDER BY exp_digital_id ASC, fecha_envio ASC, id ASC"
    ).fetchall()
    revs = conn.execute(
        "SELECT exp_digital_id, MAX(fecha_revision) AS ultima_revision FROM exp_revisiones GROUP BY exp_digital_id"
    ).fetchall()
    conn.close()

    ultima_rev_por_exp: dict[int, str] = {r["exp_digital_id"]: r["ultima_revision"] for r in revs}

    coms_por_exp: dict[int, list] = {}
    for c in coms:
        eid = c["exp_digital_id"]
        coms_por_exp.setdefault(eid, []).append(dict(c))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EXP DIGIT 2025-2026"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="dbeafe")

    cabeceras = [
        "N° Expediente", "Año", "Abogado", "Etapa", "Queja Inicial",
        "Radicado Auto", "Nombre Auto", "Fecha Auto",
        "Obs. Generales", "Última Revisión",
        "Radicado Comunicación", "Dependencia", "Fecha Envío",
        "Fecha Seguimiento", "Radicado Respuesta", "Fecha Respuesta",
        "Responsable", "Observaciones",
    ]
    ws.append(["SEGUIMIENTO EXPEDIENTES DIGITALES"])
    ws.append(cabeceras)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for exp in exps:
        ed = dict(exp)
        exp_coms = coms_por_exp.get(ed["id"], [])
        primera_com = exp_coms[0] if exp_coms else {}

        row = [
            ed.get("n_expediente"), ed.get("anio"), ed.get("abogado"), ed.get("etapa"),
            ed.get("queja_inicial"), ed.get("radicado_auto"), ed.get("nombre_auto"), ed.get("fecha_auto"),
            ed.get("observaciones"), ultima_rev_por_exp.get(ed["id"]),
            primera_com.get("radicado_comunicacion"), primera_com.get("dependencia"),
            primera_com.get("fecha_envio"), primera_com.get("fecha_seguimiento"),
            primera_com.get("radicado_respuesta"), primera_com.get("fecha_respuesta"),
            primera_com.get("responsable"), primera_com.get("observaciones"),
        ]
        ws.append(row)

        for com in exp_coms[1:]:
            sub_row = [
                None, None, None, None, None, None, None, None,
                None, None,
                com.get("radicado_comunicacion"), com.get("dependencia"),
                com.get("fecha_envio"), com.get("fecha_seguimiento"),
                com.get("radicado_respuesta"), com.get("fecha_respuesta"),
                com.get("responsable"), com.get("observaciones"),
            ]
            ws.append(sub_row)
            for cell in ws[ws.max_row]:
                cell.fill = sub_fill

    col_widths = [15, 6, 22, 20, 14, 20, 30, 14, 40, 22, 22, 25, 14, 16, 22, 14, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = date.today().strftime("%Y%m%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=exp_digitales_{hoy}.xlsx"},
    )


# ── Vista global de comunicaciones  ← ANTES DE /{exp_id} ─────────────────────

@router.get("/comunicaciones", response_class=HTMLResponse)
async def comunicaciones_lista(
    request: Request,
    sin_respuesta: str = "",
    alerta: str = "",
    abogado: str = "",
    q: str = "",
):
    conn = get_db()

    filtros = ["1=1"]
    params: list = []

    if sin_respuesta == "1":
        filtros.append("(c.fecha_respuesta IS NULL OR c.fecha_respuesta = '')")
    if alerta == "roja":
        filtros.append("(c.fecha_respuesta IS NULL OR c.fecha_respuesta = '')")
        filtros.append("c.fecha_envio IS NOT NULL AND c.fecha_envio != ''")
        filtros.append("CAST(julianday('now') - julianday(c.fecha_envio) AS INTEGER) >= 14")
    elif alerta == "amarilla":
        filtros.append("(c.fecha_respuesta IS NULL OR c.fecha_respuesta = '')")
        filtros.append("c.fecha_envio IS NOT NULL AND c.fecha_envio != ''")
        filtros.append("CAST(julianday('now') - julianday(c.fecha_envio) AS INTEGER) = 13")
    elif alerta == "azul":
        filtros.append("(c.fecha_respuesta IS NULL OR c.fecha_respuesta = '')")
        filtros.append("c.fecha_envio IS NOT NULL AND c.fecha_envio != ''")
        filtros.append("CAST(julianday('now') - julianday(c.fecha_envio) AS INTEGER) >= 8")
        filtros.append("CAST(julianday('now') - julianday(c.fecha_envio) AS INTEGER) < 13")
    if abogado.strip():
        filtros.append("e.abogado = ?")
        params.append(abogado.strip())
    if q.strip():
        filtros.append("(c.radicado_comunicacion LIKE ? OR c.dependencia LIKE ? OR e.n_expediente LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    where = " AND ".join(filtros)

    rows = conn.execute(f"""
        SELECT c.*,
               e.n_expediente, e.abogado, e.anio, e.etapa,
               e.id AS exp_id,
               CASE
                 WHEN (c.fecha_respuesta IS NULL OR c.fecha_respuesta = '')
                      AND c.fecha_envio IS NOT NULL AND c.fecha_envio != ''
                 THEN CAST(julianday('now') - julianday(c.fecha_envio) AS INTEGER)
                 ELSE NULL
               END AS dias_transcurridos
        FROM exp_comunicaciones c
        JOIN exp_digitales e ON c.exp_digital_id = e.id
        WHERE {where}
        ORDER BY e.abogado ASC, e.n_expediente ASC, c.fecha_envio ASC, c.id ASC
    """, params).fetchall()

    abogados = [r[0] for r in conn.execute(
        "SELECT DISTINCT abogado FROM exp_digitales WHERE abogado IS NOT NULL ORDER BY abogado"
    ).fetchall()]

    conn.close()

    rows_list = [dict(r) for r in rows]
    # Agregar clase de alerta a cada fila
    for r in rows_list:
        r["clase_alerta"] = _clase_alerta(r.get("dias_transcurridos"))

    return templates.TemplateResponse("digitales_comunicaciones.html", {
        "request": request,
        "active": "digitales_lista",
        "rows": rows_list,
        "total": len(rows_list),
        "sin_respuesta": sin_respuesta,
        "alerta": alerta,
        "abogado": abogado,
        "q": q,
        "abogados": abogados,
    })


# ── Comunicaciones CRUD (rutas sin {exp_id} al inicio) ────────────────────────

@router.post("/comunicacion/{com_id}/editar")
async def com_editar(
    request: Request,
    com_id: int,
    radicado_comunicacion: str = Form(""),
    dependencia: str = Form(""),
    fecha_envio: str = Form(""),
    fecha_seguimiento: str = Form(""),
    radicado_respuesta: str = Form(""),
    fecha_respuesta: str = Form(""),
    responsable: str = Form(""),
    observaciones: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute("SELECT exp_digital_id FROM exp_comunicaciones WHERE id = ?", (com_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/digitales/", status_code=303)
    exp_id = row[0]
    conn.execute("""
        UPDATE exp_comunicaciones SET
            radicado_comunicacion=?, dependencia=?, fecha_envio=?,
            fecha_seguimiento=?, radicado_respuesta=?, fecha_respuesta=?,
            responsable=?, observaciones=?
        WHERE id=?
    """, (
        _texto(radicado_comunicacion), _texto(dependencia),
        _fecha(fecha_envio), _fecha(fecha_seguimiento),
        _texto(radicado_respuesta), _fecha(fecha_respuesta),
        _texto(responsable), _texto(observaciones),
        com_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/digitales/{exp_id}/editar?msg=com_actualizada", status_code=303)


@router.post("/comunicacion/{com_id}/eliminar")
async def com_eliminar(request: Request, com_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute("SELECT exp_digital_id FROM exp_comunicaciones WHERE id = ?", (com_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/digitales/", status_code=303)
    exp_id = row[0]
    conn.execute("DELETE FROM exp_comunicaciones WHERE id = ?", (com_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/digitales/{exp_id}/editar?msg=com_eliminada", status_code=303)


# ── CRUD Abogados  ← ANTES DE /{exp_id} ──────────────────────────────────────

@router.get("/abogados", response_class=HTMLResponse)
async def abogados_lista(request: Request, msg: str = ""):
    conn = get_db()
    _get_abogados(conn)  # sincroniza exp_digitales → abogados_digitales
    abogados = conn.execute(
        "SELECT id, nombre FROM abogados_digitales ORDER BY nombre"
    ).fetchall()
    conn.close()
    return templates.TemplateResponse("digitales_abogados.html", {
        "request": request,
        "active": "digitales_abogados",
        "abogados": [dict(a) for a in abogados],
        "msg": msg,
    })


@router.post("/abogados/nuevo")
async def abogado_crear(request: Request, nombre: str = Form("")):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/abogados?msg=sin_permiso", status_code=303)
    nombre = (nombre or "").strip()
    if nombre:
        conn = get_db()
        try:
            conn.execute("INSERT INTO abogados_digitales (nombre) VALUES (?)", (nombre,))
            conn.commit()
            msg = "ab_creado"
        except Exception:
            msg = "ab_duplicado"
        finally:
            conn.close()
    else:
        msg = "ab_vacio"
    return RedirectResponse(f"/digitales/abogados?msg={msg}", status_code=303)


@router.post("/abogados/{ab_id}/editar")
async def abogado_editar(request: Request, ab_id: int, nombre: str = Form("")):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/abogados?msg=sin_permiso", status_code=303)
    nombre = (nombre or "").strip()
    if nombre:
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT nombre FROM abogados_digitales WHERE id=?", (ab_id,)
            ).fetchone()
            if row:
                nombre_viejo = row[0]
                # Comprobar si el nombre destino ya existe (fusión de duplicados)
                destino = conn.execute(
                    "SELECT id FROM abogados_digitales WHERE nombre=? AND id!=?", (nombre, ab_id)
                ).fetchone()
                # Reasignar expedientes al nombre destino (nuevo o ya existente)
                conn.execute(
                    "UPDATE exp_digitales SET abogado=? WHERE abogado=?",
                    (nombre, nombre_viejo)
                )
                if destino:
                    # Fusión: el nombre ya existe → solo eliminar la entrada duplicada
                    conn.execute("DELETE FROM abogados_digitales WHERE id=?", (ab_id,))
                    msg = "ab_fusionado"
                else:
                    conn.execute("UPDATE abogados_digitales SET nombre=? WHERE id=?", (nombre, ab_id))
                    msg = "ab_actualizado"
            conn.commit()
        except Exception:
            msg = "ab_duplicado"
        finally:
            conn.close()
    else:
        msg = "ab_vacio"
    return RedirectResponse(f"/digitales/abogados?msg={msg}", status_code=303)


@router.post("/abogados/{ab_id}/eliminar")
async def abogado_eliminar(request: Request, ab_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/digitales/abogados?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute(
        "SELECT nombre FROM abogados_digitales WHERE id=?", (ab_id,)
    ).fetchone()
    if row:
        # Cascada: limpiar el campo en exp_digitales para que no reaparezca al sincronizar
        conn.execute(
            "UPDATE exp_digitales SET abogado=NULL WHERE abogado=?", (row[0],)
        )
    conn.execute("DELETE FROM abogados_digitales WHERE id=?", (ab_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/digitales/abogados?msg=ab_eliminado", status_code=303)


# ── Detalle  ← /{exp_id} siempre AL FINAL ─────────────────────────────────────

@router.get("/{exp_id}", response_class=HTMLResponse)
async def detalle(request: Request, exp_id: int, msg: str = "", back: str = ""):
    conn = get_db()
    exp = conn.execute("SELECT * FROM exp_digitales WHERE id = ?", (exp_id,)).fetchone()
    if not exp:
        conn.close()
        return RedirectResponse("/digitales/?msg=no_encontrado")
    comunicaciones = conn.execute(
        "SELECT * FROM exp_comunicaciones WHERE exp_digital_id = ? ORDER BY fecha_envio ASC, id ASC",
        (exp_id,)
    ).fetchall()
    revisiones = conn.execute(
        "SELECT * FROM exp_revisiones WHERE exp_digital_id = ? ORDER BY fecha_revision DESC",
        (exp_id,)
    ).fetchall()
    conn.close()
    return templates.TemplateResponse("digitales_detalle.html", {
        "request": request,
        "active": "digitales_lista",
        "exp": dict(exp),
        "comunicaciones": [dict(c) for c in comunicaciones],
        "revisiones": [dict(r) for r in revisiones],
        "msg": msg,
        "back": back,
    })


@router.post("/{exp_id}/revisar")
async def marcar_revisado(request: Request, exp_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/digitales/{exp_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute(
        "INSERT INTO exp_revisiones (exp_digital_id) VALUES (?)", (exp_id,)
    )
    conn.commit()
    conn.close()
    return RedirectResponse(f"/digitales/{exp_id}?msg=revisado", status_code=303)


@router.get("/{exp_id}/editar", response_class=HTMLResponse)
async def editar_form(request: Request, exp_id: int, msg: str = "", back: str = ""):
    conn = get_db()
    exp = conn.execute("SELECT * FROM exp_digitales WHERE id = ?", (exp_id,)).fetchone()
    if not exp:
        conn.close()
        return RedirectResponse("/digitales/?msg=no_encontrado")
    comunicaciones = conn.execute(
        "SELECT * FROM exp_comunicaciones WHERE exp_digital_id = ? ORDER BY fecha_envio ASC, id ASC",
        (exp_id,)
    ).fetchall()
    revisiones = conn.execute(
        "SELECT * FROM exp_revisiones WHERE exp_digital_id = ? ORDER BY fecha_revision DESC LIMIT 1",
        (exp_id,)
    ).fetchall()
    abogados = _get_abogados(conn)
    conn.close()
    return templates.TemplateResponse("digitales_form.html", {
        "request": request,
        "active": "digitales_lista",
        "exp": dict(exp),
        "comunicaciones": [dict(c) for c in comunicaciones],
        "revisiones": [dict(r) for r in revisiones],
        "modo": "editar",
        "abogados": abogados,
        "msg": msg,
        "back": back,
    })


@router.post("/{exp_id}/editar")
async def editar_post(
    request: Request,
    exp_id: int,
    n_expediente: str = Form(""),
    anio: str = Form(""),
    abogado: str = Form(""),
    etapa: str = Form(""),
    queja_inicial: str = Form("No"),
    radicado_auto: str = Form(""),
    nombre_auto: str = Form(""),
    fecha_auto: str = Form(""),
    observaciones: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/digitales/{exp_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("""
        UPDATE exp_digitales SET n_expediente=?, anio=?, abogado=?, etapa=?, queja_inicial=?,
            radicado_auto=?, nombre_auto=?, fecha_auto=?, observaciones=?,
            updated_at=datetime('now','localtime')
        WHERE id=?
    """, (
        _texto(n_expediente), int(anio) if anio.strip() else None,
        _texto(abogado), _texto(etapa), queja_inicial or "No",
        _texto(radicado_auto), _texto(nombre_auto), _fecha(fecha_auto),
        _texto(observaciones), exp_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/digitales/{exp_id}?msg=actualizado", status_code=303)


@router.post("/{exp_id}/eliminar")
async def eliminar(request: Request, exp_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/digitales/?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute("SELECT n_expediente FROM exp_digitales WHERE id = ?", (exp_id,)).fetchone()
    if row:
        n = row[0] or str(exp_id)
        conn.execute("DELETE FROM exp_digitales WHERE id = ?", (exp_id,))
        conn.commit()
        conn.close()
        return RedirectResponse(f"/digitales/?msg=eliminado_{n}", status_code=303)
    conn.close()
    return RedirectResponse("/digitales/?msg=no_encontrado", status_code=303)


@router.post("/{exp_id}/comunicacion/nueva")
async def com_nueva(
    request: Request,
    exp_id: int,
    radicado_comunicacion: str = Form(""),
    dependencia: str = Form(""),
    fecha_envio: str = Form(""),
    fecha_seguimiento: str = Form(""),
    radicado_respuesta: str = Form(""),
    fecha_respuesta: str = Form(""),
    responsable: str = Form(""),
    observaciones: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/digitales/{exp_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("""
        INSERT INTO exp_comunicaciones
            (exp_digital_id, radicado_comunicacion, dependencia, fecha_envio,
             fecha_seguimiento, radicado_respuesta, fecha_respuesta, responsable, observaciones)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        exp_id,
        _texto(radicado_comunicacion), _texto(dependencia),
        _fecha(fecha_envio), _fecha(fecha_seguimiento),
        _texto(radicado_respuesta), _fecha(fecha_respuesta),
        _texto(responsable), _texto(observaciones),
    ))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/digitales/{exp_id}/editar?msg=com_creada", status_code=303)
