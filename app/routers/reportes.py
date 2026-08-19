"""
Reporte de vencimientos críticos — Sistema OCDI
================================================
Genera un Excel con dos hojas:
  Hoja 1 — Correspondencia: registros amarillos, rojos y sin plazo definido
  Hoja 2 — SDQS: registros amarillos, rojos y sin fecha de vencimiento

Pensado para revisión los martes y jueves.
"""

from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from pathlib import Path
from datetime import date, datetime
import io

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.routers.correspondencia import _calcular_semaforo_row
from app.routers.sdqs import _calcular_semaforo_sdqs

router = APIRouter(prefix="/reportes")


def _row_to_dict(row) -> dict:
    return dict(row)


# ── Colores ───────────────────────────────────────────────────────────────────
_ROJO_BG    = PatternFill("solid", fgColor="FFCCCC")
_AMARILLO_BG = PatternFill("solid", fgColor="FFF8CC")
_SINPLAZO_BG = PatternFill("solid", fgColor="FCE7F3")
_HEADER_BG   = PatternFill("solid", fgColor="0D3060")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT  = Font(bold=True, size=12, color="0D3060")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_na = lambda v: v if (v is not None and str(v).strip() != "") else "N/A"


def _fmt_fecha(s):
    if not s:
        return "N/A"
    try:
        return date.fromisoformat(str(s)[:10]).strftime("%d/%m/%Y")
    except Exception:
        return str(s)


def _fmt_ts(s):
    """Formatea timestamp 'YYYY-MM-DD HH:MM:SS' → 'DD/MM/YYYY HH:MM'"""
    if not s:
        return "N/A"
    try:
        dt = datetime.fromisoformat(str(s)[:16])
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(s)


def _set_header_row(ws, fila: int, cols: list[str]):
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=fila, column=ci, value=h)
        c.fill = _HEADER_BG
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _THIN_BORDER
    ws.row_dimensions[fila].height = 32


def _set_data_cell(ws, ri, ci, val, fill=None, align=None):
    c = ws.cell(row=ri, column=ci, value=val)
    if fill:
        c.fill = fill
    c.alignment = align or _LEFT
    c.border = _THIN_BORDER
    c.font = Font(size=10)
    return c


# ── Hoja 1: Correspondencia ───────────────────────────────────────────────────

def _build_hoja_correspondencia(wb, hoy: date):
    ws = wb.active
    ws.title = "Correspondencia"

    # Título
    ws.merge_cells("A1:R1")
    t = ws.cell(row=1, column=1,
        value=f"REPORTE VENCIMIENTOS — CORRESPONDENCIA — {hoy.strftime('%d/%m/%Y')}")
    t.font = Font(bold=True, size=13, color="0D3060")
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:R2")
    sub = ws.cell(row=2, column=1,
        value="Incluye: 🔴 Vencidos · 🟡 Por vencer · ⚠️ Sin plazo definido — pendientes sin respuesta")
    sub.font = Font(size=10, italic=True, color="555555")
    sub.alignment = _CENTER
    ws.row_dimensions[2].height = 18

    COLS = [
        "AÑO", "MES", "FECHA INGRESO", "N. RADICADO",
        "ENTIDAD / ORIGEN", "ASUNTO", "TIPO REQUERIMIENTO",
        "TÉRMINO (DÍAS HAB.)", "PLAZO DEFINIDO",
        "RESPONSABLE", "CASO BMP",
        "ESTADO SEMÁFORO", "DÍAS TRANSCURRIDOS",
        "FECHA VENCIMIENTO LEGAL", "DÍAS RESTANTES",
        "TIPO RESPUESTA", "FECHA CREACIÓN", "ÚLTIMA MODIFICACIÓN",
    ]
    _set_header_row(ws, 3, COLS)

    conn = get_db()
    rows_raw = conn.execute("""
        SELECT c.*,
               GROUP_CONCAT(DISTINCT rs.radicado ORDER BY rs.id) AS radicados_concat
        FROM correspondencia c
        LEFT JOIN correspondencia_radicados_salida rs ON rs.correspondencia_id = c.id
        GROUP BY c.id
        ORDER BY c.fecha_ingreso ASC
    """).fetchall()
    conn.close()

    ri = 4
    total = 0
    for raw in rows_raw:
        d = _calcular_semaforo_row(_row_to_dict(raw))
        sem = d.get("semaforo")
        pendiente = not bool(d.get("fecha_radicado_salida"))
        sin_plazo = pendiente and not d.get("termino_dias")

        # Incluir: amarilla, roja, o pendiente sin plazo
        if sem not in ("amarilla", "roja") and not sin_plazo:
            continue

        if sem == "roja":
            fill = _ROJO_BG
            estado_label = "🔴 VENCIDO"
        elif sem == "amarilla":
            fill = _AMARILLO_BG
            estado_label = "🟡 POR VENCER"
        elif sin_plazo:
            fill = _SINPLAZO_BG
            estado_label = "⚠️ SIN PLAZO"
        else:
            continue

        dias_r = d.get("dias_restantes")
        dias_r_str = str(dias_r) if dias_r is not None else "N/A"
        n_rad = d.get("radicados_concat") or d.get("n_radicado") or "N/A"

        vals = [
            _na(d.get("anio")),
            _na(d.get("mes")),
            _fmt_fecha(d.get("fecha_ingreso")),
            _na(n_rad),
            _na(d.get("origen")),
            _na(d.get("asunto")),
            _na(d.get("tipo_requerimiento")),
            d.get("termino_dias") if d.get("termino_dias") is not None else "N/A",
            "SÍ" if d.get("termino_dias") else "NO ⚠️",
            _na(d.get("responsable")),
            _na(d.get("caso_bmp")),
            estado_label,
            _na(d.get("dias_transcurridos")),
            _fmt_fecha(d.get("fecha_vencimiento")),
            dias_r_str,
            _na(d.get("tipo_respuesta")) if d.get("tipo_respuesta") else "PENDIENTE",
            _fmt_ts(d.get("created_at")),
            _fmt_ts(d.get("updated_at")),
        ]
        for ci, v in enumerate(vals, 1):
            _set_data_cell(ws, ri, ci, v, fill=fill,
                           align=_CENTER if ci in (1, 2, 8, 9, 12, 13, 15) else _LEFT)
        ws.row_dimensions[ri].height = 16
        ri += 1
        total += 1

    # Anchos de columna
    anchos = [7, 10, 14, 22, 28, 36, 22, 10, 10, 20, 14, 14, 10, 16, 10, 16, 18, 18]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return total


# ── Hoja 2: SDQS ─────────────────────────────────────────────────────────────

def _build_hoja_sdqs(wb, hoy: date):
    ws = wb.create_sheet("SDQS")

    ws.merge_cells("A1:N1")
    t = ws.cell(row=1, column=1,
        value=f"REPORTE VENCIMIENTOS — SDQS — {hoy.strftime('%d/%m/%Y')}")
    t.font = Font(bold=True, size=13, color="0D3060")
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    sub = ws.cell(row=2, column=1,
        value="Incluye: 🔴 Vencidos (≤ 2 días) · 🟡 Segunda mitad del plazo · ⚠️ Sin fecha de vencimiento — activos")
    sub.font = Font(size=10, italic=True, color="555555")
    sub.alignment = _CENTER
    ws.row_dimensions[2].height = 18

    COLS = [
        "MES", "SDQS", "QUEJOSO", "TEMA",
        "FECHA ASIGNACIÓN", "FECHA VENCIMIENTO", "PLAZO DEFINIDO",
        "ESTADO SEMÁFORO", "DÍAS RESTANTES",
        "COMPETENCIA OCDI", "BPM", "RESPONSABLE",
        "FECHA CREACIÓN", "ÚLTIMA MODIFICACIÓN",
    ]
    _set_header_row(ws, 3, COLS)

    conn = get_db()
    rows_raw = conn.execute("""
        SELECT * FROM sdqs ORDER BY fecha_asignacion ASC
    """).fetchall()
    conn.close()

    ri = 4
    total = 0
    for raw in rows_raw:
        d = _calcular_semaforo_sdqs(_row_to_dict(raw))
        sem = d.get("semaforo_sdqs")
        activo = not bool((d.get("rad_salida") or "").strip())
        sin_vencimiento = activo and not (d.get("fecha_vencimiento") or "").strip()

        if sem not in ("amarillo", "rojo") and not sin_vencimiento:
            continue

        if sem == "rojo":
            fill = _ROJO_BG
            estado_label = "🔴 VENCIDO"
        elif sem == "amarillo":
            fill = _AMARILLO_BG
            estado_label = "🟡 POR VENCER"
        elif sin_vencimiento:
            fill = _SINPLAZO_BG
            estado_label = "⚠️ SIN VENCIMIENTO"
        else:
            continue

        fa = d.get("fecha_asignacion") or ""
        fv = d.get("fecha_vencimiento") or ""
        dias_rest = "N/A"
        if fv:
            try:
                dias_rest = str((date.fromisoformat(fv[:10]) - hoy).days)
            except Exception:
                pass

        vals = [
            _na(d.get("mes")),
            _na(d.get("sdqs")),
            _na(d.get("quejoso")),
            _na(d.get("tema")),
            _fmt_fecha(fa),
            _fmt_fecha(fv) if fv else "N/A ⚠️",
            "SÍ" if fv else "NO ⚠️",
            estado_label,
            dias_rest,
            _na(d.get("competencia_ocdi")),
            _na(d.get("bpm")),
            _na(d.get("responsable")),
            _fmt_ts(d.get("created_at")),
            _fmt_ts(d.get("updated_at")),
        ]
        for ci, v in enumerate(vals, 1):
            _set_data_cell(ws, ri, ci, v, fill=fill,
                           align=_CENTER if ci in (5, 6, 7, 8, 9, 10) else _LEFT)
        ws.row_dimensions[ri].height = 16
        ri += 1
        total += 1

    anchos = [10, 18, 22, 32, 14, 14, 10, 14, 10, 10, 14, 20, 18, 18]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return total


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.get("/vencimientos")
async def reporte_vencimientos(request: Request):
    hoy = date.today()
    dia_semana = hoy.strftime("%A")  # para el nombre del archivo

    wb = openpyxl.Workbook()
    total_corr = _build_hoja_correspondencia(wb, hoy)
    total_sdqs = _build_hoja_sdqs(wb, hoy)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"Reporte_Vencimientos_{hoy.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )
