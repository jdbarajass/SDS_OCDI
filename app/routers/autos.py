from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from datetime import date
import io

from app.database import get_db, row_to_dict

router = APIRouter()
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))

MESES_ORD = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
             "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

# Tipos de auto y el campo de fecha correspondiente en la BD
TIPOS_AUTO = [
    ("APERTURA INDAGACIÓN PREVIA",    "fecha_auto_apertura_ind"),
    ("TRASLADO INDAGACIÓN PREVIA",    "fecha_auto_traslado_ind"),
    ("ARCHIVO INDAGACIÓN PREVIA",     "fecha_auto_archivo_ind"),
    ("APERTURA INVESTIGACIÓN DISC.",  "fecha_auto_apertura_inv"),
    ("TRASLADO INVESTIGACIÓN DISC.",  "fecha_auto_traslado_inv"),
    ("ARCHIVO INVESTIGACIÓN DISC.",   "fecha_auto_archivo_inv"),
]


def _mes_de_fecha(fecha_str: str | None) -> str | None:
    """Devuelve el nombre del mes a partir de una fecha ISO."""
    if not fecha_str:
        return None
    try:
        m = int(fecha_str[5:7])
        return MESES_ORD[m - 1]
    except (ValueError, IndexError):
        return None


@router.get("/autos", response_class=HTMLResponse)
async def control_autos(request: Request, anio: str = ""):
    anio_sel = anio or str(date.today().year)

    conn = get_db()

    anios = [str(r[0]) for r in conn.execute(
        "SELECT DISTINCT anio FROM expedientes WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()]

    # Construir tabla: tipo_auto × mes → cantidad
    tabla = {}  # {tipo_auto: {mes: count}}
    tabla_abogado = {}  # {abogado: {tipo_auto: count}}

    for tipo, campo_fecha in TIPOS_AUTO:
        tabla[tipo] = {m: 0 for m in MESES_ORD}
        filas = conn.execute(
            f"SELECT {campo_fecha}, nombre_abogado FROM expedientes "
            f"WHERE {campo_fecha} IS NOT NULL AND {campo_fecha} LIKE '{anio_sel}-%'"
        ).fetchall()
        for fila in filas:
            mes = _mes_de_fecha(fila[0])
            abogado = fila[1] or "Sin asignar"
            if mes:
                tabla[tipo][mes] += 1
            # Acumular por abogado
            if abogado not in tabla_abogado:
                tabla_abogado[abogado] = {t: 0 for t, _ in TIPOS_AUTO}
            tabla_abogado[abogado][tipo] += 1

    # Totales por mes (suma de todos los tipos)
    totales_mes = {m: sum(tabla[t][m] for t, _ in TIPOS_AUTO) for m in MESES_ORD}

    # Totales por tipo
    totales_tipo = {tipo: sum(tabla[tipo].values()) for tipo, _ in TIPOS_AUTO}

    # Total general
    total_general = sum(totales_tipo.values())

    conn.close()

    return templates.TemplateResponse("autos.html", {
        "request": request,
        "active": "autos",
        "anio_sel": anio_sel,
        "anios": anios,
        "tipos_auto": [t for t, _ in TIPOS_AUTO],
        "meses": MESES_ORD,
        "tabla": tabla,
        "tabla_abogado": tabla_abogado,
        "totales_mes": totales_mes,
        "totales_tipo": totales_tipo,
        "total_general": total_general,
    })


@router.get("/autos/exportar")
async def exportar_autos(request: Request, anio: str = ""):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    anio_sel = anio or str(date.today().year)
    conn = get_db()

    tabla = {}
    tabla_abogado = {}
    for tipo, campo_fecha in TIPOS_AUTO:
        tabla[tipo] = {m: 0 for m in MESES_ORD}
        filas = conn.execute(
            f"SELECT {campo_fecha}, nombre_abogado FROM expedientes "
            f"WHERE {campo_fecha} IS NOT NULL AND {campo_fecha} LIKE '{anio_sel}-%'"
        ).fetchall()
        for fila in filas:
            mes = _mes_de_fecha(fila[0])
            abogado = fila[1] or "Sin asignar"
            if mes:
                tabla[tipo][mes] += 1
            if abogado not in tabla_abogado:
                tabla_abogado[abogado] = {t: 0 for t, _ in TIPOS_AUTO}
            tabla_abogado[abogado][tipo] += 1
    conn.close()

    wb = openpyxl.Workbook()
    h_fill = PatternFill("solid", fgColor="1B4F8A")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    ctr = Alignment(horizontal="center", vertical="center")

    # Hoja 1: Por mes
    ws1 = wb.active
    ws1.title = f"Autos por Mes {anio_sel}"
    headers = ["TIPO DE AUTO"] + MESES_ORD + ["TOTAL"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = h_fill; c.font = h_font; c.alignment = ctr
    ws1.row_dimensions[1].height = 30

    alt = PatternFill("solid", fgColor="EBF3FD")
    for ri, (tipo, _) in enumerate(TIPOS_AUTO, 2):
        ws1.cell(row=ri, column=1, value=tipo).font = Font(bold=True, size=9)
        total_tipo = 0
        for ci, mes in enumerate(MESES_ORD, 2):
            v = tabla[tipo][mes]
            ws1.cell(row=ri, column=ci, value=v if v > 0 else "").alignment = ctr
            total_tipo += v
            if ri % 2 == 0:
                ws1.cell(row=ri, column=ci).fill = alt
        ws1.cell(row=ri, column=len(MESES_ORD)+2, value=total_tipo).font = Font(bold=True)
        if ri % 2 == 0:
            ws1.cell(row=ri, column=1).fill = alt

    # Fila de totales
    tr = len(TIPOS_AUTO) + 2
    ws1.cell(row=tr, column=1, value="TOTAL GENERAL").font = Font(bold=True, color="FFFFFF")
    ws1.cell(row=tr, column=1).fill = h_fill
    grand = 0
    for ci, mes in enumerate(MESES_ORD, 2):
        tot = sum(tabla[t][mes] for t, _ in TIPOS_AUTO)
        ws1.cell(row=tr, column=ci, value=tot if tot > 0 else "")
        ws1.cell(row=tr, column=ci).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=tr, column=ci).fill = h_fill
        ws1.cell(row=tr, column=ci).alignment = ctr
        grand += tot
    ws1.cell(row=tr, column=len(MESES_ORD)+2, value=grand).font = Font(bold=True, color="FFFFFF")
    ws1.cell(row=tr, column=len(MESES_ORD)+2).fill = h_fill
    ws1.column_dimensions["A"].width = 36
    for col in ws1.columns:
        if col[0].column_letter != "A":
            ws1.column_dimensions[col[0].column_letter].width = 10

    # Hoja 2: Por abogado
    ws2 = wb.create_sheet(f"Autos por Abogado {anio_sel}")
    h2 = ["ABOGADO"] + [t for t, _ in TIPOS_AUTO] + ["TOTAL"]
    for ci, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = h_fill; c.font = h_font; c.alignment = ctr
    ws2.row_dimensions[1].height = 50

    for ri, (abogado, tipos) in enumerate(sorted(tabla_abogado.items()), 2):
        ws2.cell(row=ri, column=1, value=abogado)
        total_ab = 0
        for ci, (tipo, _) in enumerate(TIPOS_AUTO, 2):
            v = tipos.get(tipo, 0)
            ws2.cell(row=ri, column=ci, value=v if v > 0 else "").alignment = ctr
            total_ab += v
            if ri % 2 == 0: ws2.cell(row=ri, column=ci).fill = alt
        ws2.cell(row=ri, column=len(TIPOS_AUTO)+2, value=total_ab).font = Font(bold=True)
        if ri % 2 == 0: ws2.cell(row=ri, column=1).fill = alt
    ws2.column_dimensions["A"].width = 32
    for col in ws2.columns:
        if col[0].column_letter != "A":
            ws2.column_dimensions[col[0].column_letter].width = 22

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"OCDI_ControlAutos_{anio_sel}_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
