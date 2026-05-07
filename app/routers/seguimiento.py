import io
from datetime import date
from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from app.database import get_db
from app.auth_utils import tpl, puede_escribir

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

MESES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

_MOD = "expedientes"


@router.get("/seguimiento", response_class=HTMLResponse)
async def seguimiento_get(request: Request, anio: int = -1, abogado: str = "", q: str = ""):
    user = request.state.user
    if not user:
        return RedirectResponse("/login", status_code=302)

    anio_actual = date.today().year

    conn = get_db()

    # Años disponibles (de expedientes en BD)
    anios_rows = conn.execute(
        "SELECT DISTINCT anio FROM expedientes WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()
    anios = [r[0] for r in anios_rows]
    if anio_actual not in anios:
        anios.insert(0, anio_actual)

    # anio=-1 (sin especificar) → el más reciente con expedientes
    # anio=0  → todos los años
    # anio>0  → año concreto
    if anio == -1:
        anio = anios_rows[0][0] if anios_rows else anio_actual

    # Abogados disponibles
    ab_rows = conn.execute(
        "SELECT DISTINCT abogado_asignado FROM expedientes "
        "WHERE abogado_asignado IS NOT NULL AND abogado_asignado != '' "
        "ORDER BY abogado_asignado"
    ).fetchall()
    abogados = [r[0] for r in ab_rows]

    # Expedientes filtrados
    sql = (
        "SELECT id, n_expediente, etapa_actual AS etapa, abogado_asignado AS nombre_abogado, anio "
        "FROM expedientes WHERE 1=1"
    )
    params: list = []

    if anio != 0:
        sql += " AND anio = ?"
        params.append(anio)

    if abogado:
        sql += " AND abogado_asignado = ?"
        params.append(abogado)

    if q.strip():
        sql += " AND (n_expediente LIKE ? OR UPPER(nombre_investigado) LIKE ?)"
        like = f"%{q.strip().upper()}%"
        params += [like, like]

    sql += " ORDER BY anio DESC, CAST(n_expediente AS INTEGER), n_expediente"
    exp_rows = conn.execute(sql, params).fetchall()
    expedientes = [dict(r) for r in exp_rows]

    # Actuaciones para esos expedientes
    acts_map: dict = {}
    if expedientes:
        ids = [e["id"] for e in expedientes]
        placeholders = ",".join("?" * len(ids))
        if anio != 0:
            act_rows = conn.execute(
                f"SELECT expediente_id, mes, descripcion, created_by "
                f"FROM seguimiento_mensual WHERE anio = ? AND expediente_id IN ({placeholders})",
                [anio] + ids,
            ).fetchall()
        else:
            act_rows = conn.execute(
                f"SELECT expediente_id, mes, descripcion, created_by "
                f"FROM seguimiento_mensual WHERE expediente_id IN ({placeholders})",
                ids,
            ).fetchall()
        for r in act_rows:
            acts_map[(r["expediente_id"], r["mes"])] = dict(r)

    conn.close()

    return templates.TemplateResponse("seguimiento.html", {
        "request": request,
        "active": "seguimiento",
        "expedientes": expedientes,
        "anios": anios,
        "anio_sel": anio,
        "abogados": abogados,
        "abogado_sel": abogado,
        "q": q,
        "meses": MESES,
        "acts_map": acts_map,
    })


@router.post("/seguimiento/guardar")
async def seguimiento_guardar(request: Request):
    user = request.state.user
    if not user:
        return RedirectResponse("/login", status_code=302)

    if not puede_escribir(user, _MOD):
        return RedirectResponse("/seguimiento?msg=sin_permiso", status_code=302)

    form = await request.form()
    expediente_id = int(form.get("expediente_id", 0))
    anio = int(form.get("anio", date.today().year))
    mes = (form.get("mes") or "").strip().upper()
    descripcion = (form.get("descripcion") or "").strip()
    created_by = (form.get("created_by") or user.get("nombre_completo") or "").strip()

    conn = get_db()
    if descripcion:
        conn.execute(
            """INSERT INTO seguimiento_mensual
                   (expediente_id, anio, mes, descripcion, created_by, updated_at)
               VALUES (?,?,?,?,?, datetime('now','localtime'))
               ON CONFLICT(expediente_id, anio, mes) DO UPDATE SET
                   descripcion = excluded.descripcion,
                   created_by  = excluded.created_by,
                   updated_at  = datetime('now','localtime')""",
            (expediente_id, anio, mes, descripcion, created_by),
        )
    else:
        conn.execute(
            "DELETE FROM seguimiento_mensual WHERE expediente_id=? AND anio=? AND mes=?",
            (expediente_id, anio, mes),
        )
    conn.commit()
    conn.close()

    return RedirectResponse(f"/seguimiento?anio={anio}", status_code=302)


@router.get("/seguimiento/exportar")
async def seguimiento_exportar(request: Request, anio: int = 0, abogado: str = "", q: str = ""):
    user = request.state.user
    if not user:
        return RedirectResponse("/login", status_code=302)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return RedirectResponse("/seguimiento?msg=error_openpyxl")

    conn = get_db()

    sql = (
        "SELECT id, n_expediente, anio, etapa_actual AS etapa, abogado_asignado AS nombre_abogado "
        "FROM expedientes WHERE 1=1"
    )
    params: list = []
    if anio != 0:
        sql += " AND anio = ?"
        params.append(anio)
    if abogado:
        sql += " AND abogado_asignado = ?"
        params.append(abogado)
    if q.strip():
        sql += " AND (n_expediente LIKE ? OR UPPER(nombre_investigado) LIKE ?)"
        like = f"%{q.strip().upper()}%"
        params += [like, like]
    sql += " ORDER BY anio DESC, CAST(n_expediente AS INTEGER), n_expediente"

    exp_rows = conn.execute(sql, params).fetchall()
    expedientes = [dict(r) for r in exp_rows]

    acts_map: dict = {}
    if expedientes:
        ids = [e["id"] for e in expedientes]
        placeholders = ",".join("?" * len(ids))
        if anio != 0:
            act_rows = conn.execute(
                f"SELECT expediente_id, mes, descripcion FROM seguimiento_mensual "
                f"WHERE anio = ? AND expediente_id IN ({placeholders})",
                [anio] + ids,
            ).fetchall()
        else:
            act_rows = conn.execute(
                f"SELECT expediente_id, mes, descripcion FROM seguimiento_mensual "
                f"WHERE expediente_id IN ({placeholders})",
                ids,
            ).fetchall()
        for r in act_rows:
            acts_map[(r["expediente_id"], r["mes"])] = r["descripcion"] or ""
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    titulo = f"Seguimiento {anio}" if anio else "Seguimiento Todos"
    ws.title = titulo

    # Estilos
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="0D3060")
    mes_fill   = PatternFill("solid", fgColor="1B4F8A")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap       = Alignment(wrap_text=True, vertical="top")
    thin       = Side(style="thin", color="CBD5E1")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    meses_abrev = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]
    encabezados = ["N. EXPEDIENTE", "AÑO", "ABOGADO", "ETAPA"] + meses_abrev

    ws.append(encabezados)
    for col_idx, hdr in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill if col_idx <= 4 else mes_fill
        cell.alignment = center
        cell.border = border

    for exp in expedientes:
        row_data = [
            exp["n_expediente"],
            exp.get("anio", ""),
            exp["nombre_abogado"] or "",
            exp["etapa"] or "",
        ]
        for mes_completo in MESES:
            row_data.append(acts_map.get((exp["id"], mes_completo), ""))
        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx in range(1, len(encabezados) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = wrap if col_idx > 4 else Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    # Anchos de columna
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    for i, col_letter in enumerate(["E","F","G","H","I","J","K","L","M","N","O","P"], 0):
        ws.column_dimensions[col_letter].width = 22
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "E2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    hoy_str = date.today().strftime("%Y%m%d")
    nombre = f"SeguimientoMensual_{anio if anio else 'Todos'}_{hoy_str}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )
