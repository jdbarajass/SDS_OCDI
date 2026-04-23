from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date, datetime
import io

from app.database import get_db
from app.auth_utils import tpl, puede_escribir as _pw, registrar_log

_MOD = "control_autos"

router = APIRouter(prefix="/control-autos")
templates = make_templates(str(Path(__file__).parent.parent / "templates"))

POR_PAGINA = 25

ABOGADOS_RESPONSABLES = [
    "ANDRES EDUARDO SANDOVAL MAYORGA",
    "CARLOS ALFONSO PARRA MALAVER",
    "CESAR IVAN RODRIGUEZ DAMIAN",
    "DAVID FELIPE MORALES NOGUERA",
    "JANIK HERNANDO DE LA HOZ RIOS",
    "JOSE DE JESUS BARAJAS SOTELO",
    "LUNA GICELL GUZMAN YATE",
    "MABEL GICELLA HURTADO SANCHEZ",
    "MAGDA XIMENA PAREDES LIEVANO",
    "MARA LUCIA UCROS MERLANO",
    "MARTHA PATRICIA AÑEZ MAESTRE",
]

ASUNTOS_COMUNES = [
    "AUTO DE APERTURA DE INDAGACIÓN PREVIA",
    "AUTO DE TERMINACIÓN Y ARCHIVO DEFINITIVO DE LA INVESTIGACIÓN DISCIPLINARIA",
    "AUTO DE TERMINACIÓN Y ARCHIVO DEFINITIVO DE UNA INDAGACION PREVIA",
    "AUTO INHIBITORIO",
    "AUTO POR EL CUAL SE NIEGA EL RECURSO DE QUEJA",
    "AUTO POR EL CUAL SE ASUME LA COMPETENCIA DE UN EXPEDIENTE DISCIPLINARIO",
    "AUTO POR EL CUAL SE CONCEDE EL RECURSO DE APELACIÓN",
    "AUTO POR EL CUAL SE CONCEDE EL RECURSO DE QUEJA",
    "AUTO POR EL CUAL SE CORRIGE UN ERROR ARITMÉTICO EN INDAGACIÓN PREVIA/ INVESTIGACIÓN DISCIPLINARIA",
    "AUTO POR EL CUAL SE DECLARA CERRADA LA INVESTIGACIÓN DISCIPLINARIA",
    "AUTO POR EL CUAL SE DECLARA DESIERTO UN RECURSO (QUEJA)",
    "AUTO POR EL CUAL SE DECLARA LA SUSPENSIÓN PROVISIONAL DE UN SERVIDOR PÚBLICO.",
    "AUTO POR EL CUAL SE FORMULA PLIEGO DE CARGOS",
    "AUTO POR EL CUAL SE HACE UNA ACUMULACIÓN DEL EXPEDIENTE",
    "AUTO POR EL CUAL SE HACE UNA INCORPORACION DE (DOCUMENTOS, CORREO ELECTRONICO Y DEMAS ACTUACIONES) EN LA INDAGACIÓN PREVIA / INVESTIGACION DISCIPLINARIA..",
    "AUTO POR EL CUAL SE NIEGA EL RECURSO DE APELACIÓN",
    "AUTO POR EL CUAL SE ORDENA DECRETAR PRUEBAS EN INVESTIGACION DSICIPLINARIA A SOLICITUD DE PARTE O DE OFICIO",
    "AUTO POR EL CUAL SE ORDENA LA APERTURA DE INVESTIGACIÓN DISCIPLINARIA",
    "AUTO POR EL CUAL SE ORDENA UN DESGLOSE",
    "AUTO POR EL CUAL SE PRÓRROGA EL TÉRMINO DE UNA INVESTIGACIÓN DISCIPLINARIA",
    "AUTO POR EL CUAL SE RECHAZA EL RECURSO DE QUEJA POR EXTEMPORANEO",
    "AUTO POR EL CUAL SE REMITE POR COMPETENCIA UN EXPEDIENTE",
    "AUTO POR EL CUAL SE RESUELVE NULIDAD A PETICIÓN DE PARTE O DE OFICIO EN INDAGACION PREVIA O INVESTIGACION DSICIPLINARIA",
    "AUTO POR MEDIO DEL CUAL SE RECONOCE PERSONERÍA JURÍDICA",
    "AUTO QUE RECONOCE LA FALSEDAD O TEMERIDAD DE UNA QUEJA",
    "AUTO DE PRUEBAS INDAGACIÓN PREVIA A SOLICITUD DE PARTE O DE OFICIO",
]


def _v(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "None", "#VALUE!", "#N/A", "#REF!", "—"):
        return None
    return s


def _fecha(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s if s else None


# ── Lista ──────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def ca_lista(
    request: Request,
    q: str = "",
    abogado: str = "",
    anio: str = "",
    mes: str = "",
    page: int = 1,
    msg: str = "",
):
    conn = get_db()

    where, params = [], []
    if q:
        where.append("(expediente LIKE ? OR numero_auto LIKE ? OR asunto_auto LIKE ? OR abogado_responsable LIKE ?)")
        params += [f"%{q}%"] * 4
    if abogado:
        where.append("abogado_responsable = ?")
        params.append(abogado)
    if anio:
        where.append("strftime('%Y', fecha_auto) = ?")
        params.append(anio)
    if mes:
        where.append("strftime('%m', fecha_auto) = ?")
        params.append(mes.zfill(2))

    cond = ("WHERE " + " AND ".join(where)) if where else ""
    total = conn.execute(f"SELECT COUNT(*) FROM control_autos_sustanciacion {cond}", params).fetchone()[0]
    offset = (page - 1) * POR_PAGINA
    rows = conn.execute(
        f"SELECT * FROM control_autos_sustanciacion {cond} ORDER BY fecha_auto ASC, id ASC LIMIT ? OFFSET ?",
        params + [POR_PAGINA, offset],
    ).fetchall()

    anios = conn.execute(
        "SELECT DISTINCT strftime('%Y', fecha_auto) AS a FROM control_autos_sustanciacion WHERE fecha_auto IS NOT NULL ORDER BY a DESC"
    ).fetchall()
    conn.close()

    total_pages = max(1, (total + POR_PAGINA - 1) // POR_PAGINA)
    return templates.TemplateResponse("ca_lista.html", tpl(request, _MOD,
        rows=[dict(r) for r in rows], total=total, page=page,
        total_pages=total_pages, q=q, abogado=abogado, anio=anio, mes=mes,
        abogados=ABOGADOS_RESPONSABLES, anios=[r[0] for r in anios if r[0]],
        msg=msg, active="ca_lista",
    ))


# ── Nuevo ──────────────────────────────────────────────────────────────────────

@router.get("/nuevo", response_class=HTMLResponse)
async def ca_nuevo_form(request: Request):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/control-autos/?msg=sin_permiso", status_code=303)
    return templates.TemplateResponse("ca_form.html", tpl(request, _MOD,
        reg=None, abogados=ABOGADOS_RESPONSABLES,
        asuntos=ASUNTOS_COMUNES, active="ca_nuevo",
    ))


@router.post("/nuevo")
async def ca_nuevo_post(
    request: Request,
    expediente: str = Form(""),
    numero_auto: str = Form(""),
    fecha_auto: str = Form(""),
    asunto_auto: str = Form(""),
    abogado_responsable: str = Form(""),
    observaciones: str = Form(""),
    created_by: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/control-autos/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute(
        """INSERT INTO control_autos_sustanciacion
           (expediente, numero_auto, fecha_auto, asunto_auto, abogado_responsable, observaciones, created_by)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        [
            expediente.strip() or None,
            numero_auto.strip() or None,
            fecha_auto or None,
            asunto_auto.strip() or None,
            abogado_responsable.strip() or None,
            observaciones.strip() or None,
            created_by.strip() or None,
        ],
    )
    conn.commit()
    conn.close()
    registrar_log(user, "crear", _MOD, f"Auto #{numero_auto} — {expediente}",
                  request.client.host if request.client else None)
    return RedirectResponse("/control-autos/?msg=creado", status_code=303)


# ── Exportar ───────────────────────────────────────────────────────────────────

@router.get("/exportar")
async def ca_exportar():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return RedirectResponse("/control-autos/?msg=error_openpyxl")

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM control_autos_sustanciacion ORDER BY fecha_auto ASC, id ASC"
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONTROL AUTOS"

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 30

    thin = Side(style="thin")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_v = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_inst  = PatternFill("solid", fgColor="1B4F8A")
    fill_title = PatternFill("solid", fgColor="1B4F8A")
    fill_head  = PatternFill("solid", fgColor="2E7D32")
    fill_meta  = PatternFill("solid", fgColor="E8F5E9")
    alt_fill   = PatternFill("solid", fgColor="F1F8E9")

    font_w  = Font(bold=True, color="FFFFFF", size=10)
    font_wb = Font(bold=True, color="FFFFFF", size=12)
    font_b  = Font(bold=True, size=9)
    font_s  = Font(size=9)

    # Fila 1 — logo / institución
    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value = "CONTROL DISCIPLINARIO · OFICINA DE CONTROL DISCIPLINARIO INTERNO · SISTEMA DE GESTIÓN · CONTROL DOCUMENTAL"
    c.fill = fill_inst; c.font = font_w; c.alignment = center; c.border = borde
    ws.row_dimensions[1].height = 28

    # Fila 2 — título
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "CONTROL DE AUTOS DE SUSTANCIACIÓN Y/O TRÁMITES"
    c.fill = fill_title; c.font = font_wb; c.alignment = center; c.border = borde
    ws.row_dimensions[2].height = 24

    # Fila 3 — código y versión
    ws["B3"].value = "Código:"; ws["B3"].font = font_b; ws["B3"].alignment = center; ws["B3"].border = borde; ws["B3"].fill = fill_meta
    ws.merge_cells("C3:E3")
    ws["C3"].value = "SDS-CDO-FT-001"; ws["C3"].font = font_b; ws["C3"].alignment = center; ws["C3"].border = borde; ws["C3"].fill = fill_meta
    ws["F3"].value = "Versión"; ws["F3"].font = font_b; ws["F3"].alignment = center; ws["F3"].border = borde; ws["F3"].fill = fill_meta
    ws["G3"].value = 4; ws["G3"].font = font_b; ws["G3"].alignment = center; ws["G3"].border = borde; ws["G3"].fill = fill_meta
    ws.row_dimensions[3].height = 18

    # Fila 4 — elaborado/revisado/aprobado
    ws.merge_cells("B4:G4")
    c = ws["B4"]
    c.value = ("Elaborado por: Maricela Aldana Caicedo  /  "
               "Revisado por: Rodolfo Carrillo Quintero  /  "
               "Aprobado por: Martha Patricia Añez Maestre")
    c.font = Font(size=8, italic=True); c.alignment = left_v; c.border = borde; c.fill = fill_meta
    ws.row_dimensions[4].height = 16

    # Fila 5 — vacía de separación
    ws.row_dimensions[5].height = 6

    # Fila 6 — encabezados de columna
    headers = ["EXPEDIENTE", "NÚMERO DEL AUTO", "FECHA DEL AUTO", "ASUNTO AUTO", "ABOGADO RESPONSABLE", "OBSERVACIONES"]
    for ci, h in enumerate(headers, 2):
        cell = ws.cell(row=6, column=ci, value=h)
        cell.fill = fill_head; cell.font = font_w; cell.alignment = center; cell.border = borde
    ws.row_dimensions[6].height = 30
    ws.freeze_panes = "B7"

    # Datos desde fila 7
    for ri, row in enumerate(rows, 7):
        d = dict(row)
        fill = alt_fill if ri % 2 == 0 else None
        vals = [
            d.get("expediente"),
            d.get("numero_auto"),
            d.get("fecha_auto"),
            d.get("asunto_auto"),
            d.get("abogado_responsable"),
            d.get("observaciones"),
        ]
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = font_s
            cell.alignment = center if ci != 6 else left_v
            cell.border = borde
            if fill:
                cell.fill = fill
        ws.row_dimensions[ri].height = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = date.today().strftime("%Y%m%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=SDS-CDO-FT-001_Control_Autos_{hoy}.xlsx"},
    )


# ── Importar ───────────────────────────────────────────────────────────────────

@router.get("/importar", response_class=HTMLResponse)
async def ca_importar_form(request: Request, msg: str = ""):
    return templates.TemplateResponse("ca_importar.html", {
        "request": request,
        "msg": msg,
        "active": "ca_importar",
    })


@router.post("/importar")
async def ca_importar_post(request: Request, archivo: UploadFile = File(...)):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/control-autos/importar?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/control-autos/importar?msg=error_openpyxl", status_code=303)

    contenido = await archivo.read()
    if not contenido:
        return RedirectResponse("/control-autos/importar?msg=error_vacio", status_code=303)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        return RedirectResponse("/control-autos/importar?msg=error_archivo", status_code=303)

    # Detectar formato:
    # Formato original: hoja "NUEVO", encabezados fila 7 (col B-G), datos fila 8+
    # Formato exportado: cualquier hoja, encabezados fila 1 (col B-G), datos fila 2+
    ws = None
    data_start_row = 2
    col_offset = 2  # columna B = índice 2

    if "NUEVO" in wb.sheetnames:
        ws = wb["NUEVO"]
        data_start_row = 8
        col_offset = 2
    elif "CONTROL AUTOS" in wb.sheetnames:
        ws = wb["CONTROL AUTOS"]
        data_start_row = 7
        col_offset = 2
    else:
        ws = wb[wb.sheetnames[0]]
        data_start_row = 2
        col_offset = 1  # si columna A tiene datos

    conn = get_db()
    count = 0
    try:
        conn.execute("DELETE FROM control_autos_sustanciacion")
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not any(v for v in row if v is not None):
                continue
            # Ajustar según col_offset: si offset=2, B está en índice 1 (0-based)
            idx = col_offset - 1
            def get(i):
                return _v(row[idx + i]) if len(row) > idx + i else None

            expediente  = get(0)
            numero_auto = get(1)
            fecha_auto  = _fecha(row[idx + 2]) if len(row) > idx + 2 else None
            asunto_auto = get(3)
            abogado     = get(4)
            obs         = get(5)

            if not any([expediente, numero_auto, fecha_auto, asunto_auto, abogado]):
                continue
            # Saltar filas de descripción del pie del formato (numero_auto largo = es texto descriptivo)
            if numero_auto and len(numero_auto) > 20:
                continue

            conn.execute(
                """INSERT INTO control_autos_sustanciacion
                   (expediente, numero_auto, fecha_auto, asunto_auto, abogado_responsable, observaciones)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                [expediente, numero_auto, fecha_auto, asunto_auto, abogado, obs],
            )
            count += 1

        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return RedirectResponse("/control-autos/importar?msg=error_import", status_code=303)

    conn.close()
    registrar_log(user, "importar", _MOD, f"{count} registros importados",
                  request.client.host if request.client else None)
    return RedirectResponse(f"/control-autos/importar?msg=ok_{count}", status_code=303)


# ── Detalle ────────────────────────────────────────────────────────────────────

@router.get("/{reg_id}", response_class=HTMLResponse)
async def ca_detalle(request: Request, reg_id: int, msg: str = ""):
    conn = get_db()
    reg = conn.execute(
        "SELECT * FROM control_autos_sustanciacion WHERE id = ?", (reg_id,)
    ).fetchone()
    conn.close()
    if not reg:
        return RedirectResponse("/control-autos/?msg=no_encontrado")
    return templates.TemplateResponse("ca_detalle.html", tpl(request, _MOD,
        reg=dict(reg), msg=msg, active="ca_lista",
    ))


# ── Editar ─────────────────────────────────────────────────────────────────────

@router.get("/{reg_id}/editar", response_class=HTMLResponse)
async def ca_editar_form(request: Request, reg_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/control-autos/{reg_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    reg = conn.execute(
        "SELECT * FROM control_autos_sustanciacion WHERE id = ?", (reg_id,)
    ).fetchone()
    conn.close()
    if not reg:
        return RedirectResponse("/control-autos/?msg=no_encontrado")
    return templates.TemplateResponse("ca_form.html", tpl(request, _MOD,
        reg=dict(reg), abogados=ABOGADOS_RESPONSABLES,
        asuntos=ASUNTOS_COMUNES, active="ca_lista",
    ))


@router.post("/{reg_id}/editar")
async def ca_editar_post(
    request: Request,
    reg_id: int,
    expediente: str = Form(""),
    numero_auto: str = Form(""),
    fecha_auto: str = Form(""),
    asunto_auto: str = Form(""),
    abogado_responsable: str = Form(""),
    observaciones: str = Form(""),
    created_by: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/control-autos/{reg_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute(
        """UPDATE control_autos_sustanciacion
           SET expediente=?, numero_auto=?, fecha_auto=?, asunto_auto=?,
               abogado_responsable=?, observaciones=?,
               updated_at=datetime('now','localtime')
           WHERE id=?""",
        [
            expediente.strip() or None,
            numero_auto.strip() or None,
            fecha_auto or None,
            asunto_auto.strip() or None,
            abogado_responsable.strip() or None,
            observaciones.strip() or None,
            reg_id,
        ],
    )
    conn.commit()
    conn.close()
    return RedirectResponse(f"/control-autos/{reg_id}?msg=actualizado", status_code=303)


# ── Eliminar ───────────────────────────────────────────────────────────────────

@router.post("/{reg_id}/eliminar")
async def ca_eliminar(request: Request, reg_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/control-autos/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("DELETE FROM control_autos_sustanciacion WHERE id = ?", (reg_id,))
    conn.commit()
    conn.close()
    registrar_log(user, "eliminar", _MOD, f"Auto #{reg_id}",
                  request.client.host if request.client else None)
    return RedirectResponse("/control-autos/?msg=eliminado", status_code=303)
