from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date, datetime
import io

from app.database import get_db, row_to_dict, get_personal_oficina
from app.auth_utils import tpl, puede_escribir as _pw, puede_importar as _pi, registrar_log

_MOD = "equipos"

router = APIRouter(prefix="/equipos")
templates = make_templates(str(Path(__file__).parent.parent / "templates"))

ESTADOS_PRESTAMO = ["Prestado", "Devuelto"]

_PAGE_SIZE = 25


def _v(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s.upper() in ("", "NAN", "NONE", "#VALUE!", "#N/A", "#REF!", "—") else s


def _fecha(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.upper() in ("NAN", "NONE"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s[:10]


def _equipo_label(b: dict) -> str:
    """Construye una descripción legible de un bien para mostrar/seleccionar en préstamos."""
    partes = [b.get("descripcion_elemento") or "EQUIPO"]
    marca_modelo = " ".join(p for p in [b.get("marca"), b.get("modelo")] if p)
    if marca_modelo:
        partes.append(marca_modelo)
    if b.get("numero_placa_fisica"):
        partes.append(f"Placa {b['numero_placa_fisica']}")
    return " — ".join(partes)


# ═══════════════════════ PRÉSTAMOS DE EQUIPOS ════════════════════════════════

@router.get("/", response_class=HTMLResponse)
async def lista(
    request: Request,
    estado: str = "",
    funcionario: str = "",
    q: str = "",
    page: int = 1,
    msg: str = "",
):
    conn = get_db()
    where, params = ["1=1"], []
    if estado:
        where.append("estado = ?")
        params.append(estado)
    if funcionario:
        where.append("funcionario = ?")
        params.append(funcionario)
    if q:
        where.append("(UPPER(equipo_descripcion) LIKE ? OR UPPER(funcionario) LIKE ? OR UPPER(observaciones) LIKE ?)")
        like = f"%{q.upper()}%"
        params += [like, like, like]

    cond = " AND ".join(where)
    total = conn.execute(f"SELECT COUNT(*) FROM prestamos_equipos WHERE {cond}", params).fetchone()[0]
    total_pages = max(1, (total + _PAGE_SIZE - 1) // _PAGE_SIZE)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * _PAGE_SIZE
    rows = conn.execute(
        f"""SELECT * FROM prestamos_equipos WHERE {cond}
            ORDER BY (estado = 'Prestado') DESC, fecha_prestamo DESC, id DESC
            LIMIT ? OFFSET ?""",
        params + [_PAGE_SIZE, offset],
    ).fetchall()

    funcionarios_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT funcionario FROM prestamos_equipos WHERE funcionario IS NOT NULL ORDER BY funcionario"
    ).fetchall()]
    total_prestados = conn.execute(
        "SELECT COUNT(*) FROM prestamos_equipos WHERE estado = 'Prestado'"
    ).fetchone()[0]
    conn.close()

    return templates.TemplateResponse("equipos_lista.html", tpl(request, _MOD,
        rows=[dict(r) for r in rows], total=total, page=page, total_pages=total_pages,
        estado=estado, funcionario=funcionario, q=q,
        funcionarios_bd=funcionarios_bd, estados=ESTADOS_PRESTAMO,
        total_prestados=total_prestados,
        msg=msg, active="equipos_lista",
    ))


@router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_form(request: Request):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/equipos/?msg=sin_permiso", status_code=303)
    conn = get_db()
    personal = get_personal_oficina(conn)
    bienes = conn.execute(
        "SELECT * FROM bienes_muebles ORDER BY descripcion_elemento, marca"
    ).fetchall()
    conn.close()
    bienes_opts = [{"id": b["id"], "label": _equipo_label(dict(b))} for b in bienes]
    return templates.TemplateResponse("equipos_form.html", tpl(request, _MOD,
        modo="nuevo", reg={}, personal=personal, bienes=bienes_opts,
        estados=ESTADOS_PRESTAMO, active="equipos_nuevo",
    ))


@router.post("/nuevo")
async def nuevo_post(
    request: Request,
    bien_id: str = Form(""),
    equipo_descripcion: str = Form(""),
    funcionario: str = Form(""),
    entregado_por: str = Form(""),
    fecha_prestamo: str = Form(...),
    hora_prestamo: str = Form(""),
    fecha_devolucion_estimada: str = Form(""),
    observaciones: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/equipos/?msg=sin_permiso", status_code=303)

    funcionario = funcionario.strip()
    equipo_descripcion = equipo_descripcion.strip()
    if not funcionario or not equipo_descripcion:
        return RedirectResponse("/equipos/nuevo?msg=error_obligatorios", status_code=303)

    bien_id_val = int(bien_id) if bien_id.strip().isdigit() else None

    conn = get_db()
    conn.execute(
        """INSERT INTO prestamos_equipos
           (bien_id, equipo_descripcion, funcionario, entregado_por, fecha_prestamo,
            hora_prestamo, fecha_devolucion_estimada, estado, observaciones, created_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'Prestado', ?, ?)""",
        (bien_id_val, equipo_descripcion, funcionario, entregado_por.strip() or None,
         fecha_prestamo, hora_prestamo or None, fecha_devolucion_estimada or None,
         observaciones.strip() or None,
         user.get("nombre_completo") if user else None),
    )
    conn.commit()
    conn.close()
    registrar_log(user, "crear", _MOD, f"Préstamo de {equipo_descripcion} a {funcionario}")
    return RedirectResponse("/equipos/?msg=creado", status_code=303)


@router.get("/{reg_id}", response_class=HTMLResponse)
async def detalle(request: Request, reg_id: int, msg: str = ""):
    conn = get_db()
    reg = conn.execute("SELECT * FROM prestamos_equipos WHERE id = ?", (reg_id,)).fetchone()
    conn.close()
    if not reg:
        return RedirectResponse("/equipos/?msg=no_encontrado", status_code=303)
    return templates.TemplateResponse("equipos_detalle.html", tpl(request, _MOD,
        reg=dict(reg), msg=msg, active="equipos_lista",
    ))


@router.get("/{reg_id}/editar", response_class=HTMLResponse)
async def editar_form(request: Request, reg_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/equipos/{reg_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    reg = conn.execute("SELECT * FROM prestamos_equipos WHERE id = ?", (reg_id,)).fetchone()
    personal = get_personal_oficina(conn)
    bienes = conn.execute(
        "SELECT * FROM bienes_muebles ORDER BY descripcion_elemento, marca"
    ).fetchall()
    conn.close()
    if not reg:
        return RedirectResponse("/equipos/?msg=no_encontrado", status_code=303)
    bienes_opts = [{"id": b["id"], "label": _equipo_label(dict(b))} for b in bienes]
    return templates.TemplateResponse("equipos_form.html", tpl(request, _MOD,
        modo="editar", reg=dict(reg), personal=personal, bienes=bienes_opts,
        estados=ESTADOS_PRESTAMO, active="equipos_lista",
    ))


@router.post("/{reg_id}/editar")
async def editar_post(
    request: Request,
    reg_id: int,
    bien_id: str = Form(""),
    equipo_descripcion: str = Form(""),
    funcionario: str = Form(""),
    entregado_por: str = Form(""),
    fecha_prestamo: str = Form(...),
    hora_prestamo: str = Form(""),
    fecha_devolucion_estimada: str = Form(""),
    fecha_devolucion_real: str = Form(""),
    estado: str = Form("Prestado"),
    observaciones: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/equipos/{reg_id}?msg=sin_permiso", status_code=303)

    funcionario = funcionario.strip()
    equipo_descripcion = equipo_descripcion.strip()
    if not funcionario or not equipo_descripcion:
        return RedirectResponse(f"/equipos/{reg_id}/editar?msg=error_obligatorios", status_code=303)

    bien_id_val = int(bien_id) if bien_id.strip().isdigit() else None
    # Si se marca como Devuelto y no hay fecha de devolución, se asume hoy.
    if estado == "Devuelto" and not fecha_devolucion_real.strip():
        fecha_devolucion_real = date.today().isoformat()
    if estado == "Prestado":
        fecha_devolucion_real = ""

    conn = get_db()
    conn.execute(
        """UPDATE prestamos_equipos SET
           bien_id=?, equipo_descripcion=?, funcionario=?, entregado_por=?,
           fecha_prestamo=?, hora_prestamo=?, fecha_devolucion_estimada=?,
           fecha_devolucion_real=?, estado=?, observaciones=?,
           updated_at=datetime('now','localtime')
           WHERE id=?""",
        (bien_id_val, equipo_descripcion, funcionario, entregado_por.strip() or None,
         fecha_prestamo, hora_prestamo or None, fecha_devolucion_estimada or None,
         fecha_devolucion_real or None, estado, observaciones.strip() or None,
         reg_id),
    )
    conn.commit()
    conn.close()
    registrar_log(user, "editar", _MOD, f"Préstamo ID {reg_id}")
    return RedirectResponse(f"/equipos/{reg_id}?msg=actualizado", status_code=303)


@router.post("/{reg_id}/devolver")
async def devolver(request: Request, reg_id: int):
    """Acción rápida: marca el préstamo como Devuelto con fecha de hoy."""
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/equipos/{reg_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute(
        """UPDATE prestamos_equipos SET estado='Devuelto',
           fecha_devolucion_real=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (date.today().isoformat(), reg_id),
    )
    conn.commit()
    conn.close()
    registrar_log(user, "editar", _MOD, f"Préstamo ID {reg_id} marcado como devuelto")
    return RedirectResponse(f"/equipos/{reg_id}?msg=devuelto", status_code=303)


@router.post("/{reg_id}/eliminar")
async def eliminar(request: Request, reg_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/equipos/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("DELETE FROM prestamos_equipos WHERE id = ?", (reg_id,))
    conn.commit()
    conn.close()
    registrar_log(user, "eliminar", _MOD, f"Préstamo ID {reg_id}")
    return RedirectResponse("/equipos/?msg=eliminado", status_code=303)


# ═══════════════════════ REPORTE DE BIENES MUEBLES OCDI ══════════════════════

@router.get("/bienes/lista", response_class=HTMLResponse)
async def bienes_lista(
    request: Request,
    responsable: str = "",
    categoria: str = "",
    q: str = "",
    page: int = 1,
    msg: str = "",
):
    conn = get_db()
    where, params = ["1=1"], []
    if responsable:
        where.append("nombre_responsable = ?")
        params.append(responsable)
    if categoria:
        where.append("descripcion_elemento = ?")
        params.append(categoria)
    if q:
        where.append("""(UPPER(descripcion_elemento) LIKE ? OR UPPER(descripcion_detallada) LIKE ?
                          OR UPPER(marca) LIKE ? OR UPPER(modelo) LIKE ? OR UPPER(numero_serial) LIKE ?
                          OR UPPER(nombre_responsable) LIKE ? OR numero_placa_fisica LIKE ?)""")
        like = f"%{q.upper()}%"
        params += [like, like, like, like, like, like, f"%{q}%"]

    cond = " AND ".join(where)
    total = conn.execute(f"SELECT COUNT(*) FROM bienes_muebles WHERE {cond}", params).fetchone()[0]
    total_pages = max(1, (total + _PAGE_SIZE - 1) // _PAGE_SIZE)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * _PAGE_SIZE
    rows = conn.execute(
        f"SELECT * FROM bienes_muebles WHERE {cond} ORDER BY descripcion_elemento, marca LIMIT ? OFFSET ?",
        params + [_PAGE_SIZE, offset],
    ).fetchall()

    responsables_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT nombre_responsable FROM bienes_muebles WHERE nombre_responsable IS NOT NULL ORDER BY nombre_responsable"
    ).fetchall()]
    categorias_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT descripcion_elemento FROM bienes_muebles WHERE descripcion_elemento IS NOT NULL ORDER BY descripcion_elemento"
    ).fetchall()]
    conn.close()

    return templates.TemplateResponse("bienes_lista.html", tpl(request, _MOD,
        rows=[dict(r) for r in rows], total=total, page=page, total_pages=total_pages,
        responsable=responsable, categoria=categoria, q=q,
        responsables_bd=responsables_bd, categorias_bd=categorias_bd,
        msg=msg, active="bienes_lista",
    ))


@router.get("/bienes/importar", response_class=HTMLResponse)
async def bienes_importar_form(request: Request, msg: str = ""):
    user = getattr(request.state, "user", None)
    if not _pi(user, _MOD):
        return RedirectResponse("/equipos/bienes/lista?msg=sin_permiso", status_code=303)
    conn = get_db()
    total_bd = conn.execute("SELECT COUNT(*) FROM bienes_muebles").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("bienes_importar.html", tpl(request, _MOD,
        total_bd=total_bd, msg=msg, active="bienes_importar",
    ))


# Encabezados esperados del Excel oficial "REPORTE BIENES MUEBLES OCDI"
_HEADER_CAMPO = {
    "ID_PLACA": "id_placa",
    "NUMERO PLACA FÍSICA": "numero_placa_fisica",
    "NUMERO PLACA FISICA": "numero_placa_fisica",
    "MARCA DEL ELEMENTO": "marca",
    "MODELO DEL ELEMENTO": "modelo",
    "NÚMERO SERIAL": "numero_serial",
    "NUMERO SERIAL": "numero_serial",
    "ID DEL ELEMENTO": "id_elemento",
    "DESCRIPCION DEL ELEMENTO": "descripcion_elemento",
    "DESCRIPCIÓN DEL ELEMENTO": "descripcion_elemento",
    "DESCRIPCION DETALLADA DEL ELEMENTO": "descripcion_detallada",
    "DESCRIPCIÓN DETALLADA DEL ELEMENTO": "descripcion_detallada",
    "INTERNO FUNCIONARIO": "interno_funcionario",
    "NOMBRES RESPONSABLE": "nombre_responsable",
    "NUMERO DEL INGRESO": "numero_ingreso",
    "FECHA INGRESO": "fecha_ingreso",
    "FECHA SERVICIO": "fecha_servicio",
    "NO. IDENTIFICACION FUNCIONARIO": "identificacion_funcionario",
    "NO IDENTIFICACION FUNCIONARIO": "identificacion_funcionario",
    "CANTIDAD_VIDA_UTIL": "cantidad_vida_util",
    "NUMERO_CONTRATO": "numero_contrato",
    "PROVEEDOR": "proveedor",
}


@router.post("/bienes/importar")
async def bienes_importar_post(request: Request, archivo: UploadFile = File(...)):
    user = getattr(request.state, "user", None)
    if not _pi(user, _MOD):
        return RedirectResponse("/equipos/bienes/lista?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/equipos/bienes/importar?msg=error_openpyxl", status_code=303)

    contenido = await archivo.read()
    if not contenido:
        return RedirectResponse("/equipos/bienes/importar?msg=error_vacio", status_code=303)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception:
        return RedirectResponse("/equipos/bienes/importar?msg=error_archivo", status_code=303)

    header_row = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
    col_campo = {ci: _HEADER_CAMPO[h] for ci, h in enumerate(header_row) if h in _HEADER_CAMPO}
    if not col_campo:
        return RedirectResponse("/equipos/bienes/importar?msg=error_hoja", status_code=303)

    conn = get_db()
    count = 0
    try:
        conn.execute("DELETE FROM bienes_muebles")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in row if v is not None):
                continue
            datos = {}
            for ci, campo in col_campo.items():
                if ci >= len(row):
                    continue
                val = row[ci]
                if campo in ("fecha_ingreso", "fecha_servicio"):
                    datos[campo] = _fecha(val)
                elif campo == "cantidad_vida_util":
                    try:
                        datos[campo] = int(val) if val not in (None, "") else None
                    except (ValueError, TypeError):
                        datos[campo] = None
                else:
                    datos[campo] = _v(val)
            if not any(datos.values()):
                continue
            campos_ins = list(datos.keys())
            conn.execute(
                f"INSERT INTO bienes_muebles ({', '.join(campos_ins)}) VALUES ({', '.join(['?']*len(campos_ins))})",
                [datos[c] for c in campos_ins],
            )
            count += 1
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return RedirectResponse("/equipos/bienes/importar?msg=error_import", status_code=303)
    conn.close()
    registrar_log(user, "importar", _MOD, f"{count} bienes importados")
    return RedirectResponse(f"/equipos/bienes/importar?msg=importado_{count}", status_code=303)


@router.get("/bienes/exportar")
async def bienes_exportar():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/equipos/bienes/lista?msg=error_openpyxl", status_code=303)

    conn = get_db()
    rows = conn.execute("SELECT * FROM bienes_muebles ORDER BY descripcion_elemento, marca").fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BIENES MUEBLES OCDI"

    headers = [
        "ID_PLACA", "NUMERO PLACA FÍSICA", "MARCA DEL ELEMENTO", "MODELO DEL ELEMENTO",
        "NÚMERO SERIAL", "ID DEL ELEMENTO", "DESCRIPCION DEL ELEMENTO",
        "DESCRIPCION DETALLADA DEL ELEMENTO", "INTERNO FUNCIONARIO", "NOMBRES RESPONSABLE",
        "NUMERO DEL INGRESO", "FECHA INGRESO", "FECHA SERVICIO",
        "NO. IDENTIFICACION FUNCIONARIO", "CANTIDAD_VIDA_UTIL", "NUMERO_CONTRATO", "PROVEEDOR",
    ]
    campos = [
        "id_placa", "numero_placa_fisica", "marca", "modelo", "numero_serial", "id_elemento",
        "descripcion_elemento", "descripcion_detallada", "interno_funcionario", "nombre_responsable",
        "numero_ingreso", "fecha_ingreso", "fecha_servicio", "identificacion_funcionario",
        "cantidad_vida_util", "numero_contrato", "proveedor",
    ]
    header_fill = PatternFill("solid", fgColor="1B4F8A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        d = dict(r)
        ws.append([d.get(c) for c in campos])

    ws.freeze_panes = "A2"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Bienes_Muebles_OCDI.xlsx"},
    )
