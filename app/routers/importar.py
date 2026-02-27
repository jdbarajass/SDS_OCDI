from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from datetime import datetime
import io

from app.database import get_db, row_to_dict

router = APIRouter()
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))


def _fecha(valor) -> str | None:
    """Convierte un valor de celda Excel a string ISO (YYYY-MM-DD) o None."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        # Descartar fechas artefacto de Excel (año < 1950)
        if valor.year < 1950:
            return None
        return valor.strftime("%Y-%m-%d")
    s = str(valor).strip()
    if not s or s == "0":
        return None
    # Intentar parsear si viene como texto
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None  # formato no reconocido → descartar (evita guardar "#VALUE!" u otros errores de Excel)


def _texto(valor) -> str | None:
    if valor is None:
        return None
    s = str(valor).strip()
    return s if s else None


def _entero(valor) -> int | None:
    if valor is None:
        return None
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return None


def _mapear_fila(row) -> dict:
    """
    Mapea una fila de la hoja ENCABEZADO del Excel a los campos de la BD.
    Índices basados en el análisis del archivo original (1-indexed).
    """
    def c(idx):
        """Obtiene valor de columna por índice 1-based."""
        try:
            return row[idx - 1]
        except IndexError:
            return None

    return {
        "n_expediente":               _texto(c(1)),
        "anio":                       _entero(c(2)),
        "mes":                        _texto(c(3)),
        "origen_proceso":             _texto(c(4)),
        "n_radicado":                 _texto(c(5)),
        "fecha_radicado":             _fecha(c(6)),
        "fecha_siias":                _fecha(c(7)),
        "ingreso_siias":              _texto(c(8)) or "NO",
        "ingreso_siad":               _texto(c(9)) or "NO",
        "fecha_ingreso_siad":         _fecha(c(10)),
        "ingreso_sid4":               _texto(c(11)) or "NO",
        "nombre_abogado":             _texto(c(12)),
        "impedimento":                _texto(c(13)) or "NO",
        "investigado":                _texto(c(14)),
        # Col 15: FECHA APERTURA INVESTIGACION (puede ser de indagación o investigación)
        # Col 16: ETAPA
        "etapa":                      _texto(c(16)),
        "perfil_indagado":            _texto(c(17)),
        "entidad_origen":             _texto(c(18)),
        "quejoso":                    _texto(c(19)),
        "asunto":                     _texto(c(20)),
        "tipologia":                  _texto(c(21)),
        "descripcion_tipologia":      _texto(c(22)),
        "relacionado_siniestro":      _texto(c(23)) or "NO",
        "responsable_siniestro":      _texto(c(24)),
        "relacionado_acoso":          _texto(c(25)) or "NO",
        "responsable_acoso":          _texto(c(26)),
        "relacionado_corrupcion":     _texto(c(27)) or "NO",
        "valores_institucionales":    _texto(c(28)),
        "fecha_hechos":               _texto(c(29)),
        # Bloque Indagación Previa
        "fecha_apertura_indagacion":  _fecha(c(30)),
        "numero_auto_apertura_ind":   _texto(c(31)),
        "fecha_auto_apertura_ind":    _fecha(c(32)),
        "plazo_ind":                  _entero(c(33)) or 180,
        "fecha_vencimiento_ind":      _fecha(c(34)),
        # Col 35: ALERTAS (calculado, se omite)
        "numero_auto_traslado_ind":   _texto(c(36)),
        "fecha_auto_traslado_ind":    _fecha(c(37)),
        "numero_auto_archivo_ind":    _texto(c(38)),
        "fecha_auto_archivo_ind":     _fecha(c(39)),
        # Bloque Investigación Disciplinaria
        "fecha_apertura_investigacion": _fecha(c(40)),
        "numero_auto_apertura_inv":   _texto(c(41)),
        "fecha_auto_apertura_inv":    _fecha(c(42)),
        # Col 43: FECHA_VENCIMIENTO_INV
        "fecha_vencimiento_inv":      _fecha(c(43)),
        # Col 44: ALERTAS (calculado, se omite)
        "plazo_inv":                  _entero(c(45)) or 180,
        "numero_auto_traslado_inv":   _texto(c(46)),
        "fecha_auto_traslado_inv":    _fecha(c(47)),
        "numero_auto_archivo_inv":    _texto(c(48)),
        "fecha_auto_archivo_inv":     _fecha(c(49)),
        # Cierre
        "estado_proceso":             _texto(c(50)),
        "observaciones_finales":      _texto(c(51)),
    }


@router.get("/importar", response_class=HTMLResponse)
async def importar_form(request: Request, msg: str = ""):
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("importar.html", {
        "request": request,
        "active": "importar",
        "resultado": None,
        "total_bd": total,
        "msg": msg,
    })


@router.post("/importar/limpiar-bd")
async def limpiar_base_datos():
    conn = get_db()
    conn.execute("DELETE FROM actuaciones")
    conn.execute("DELETE FROM escaneos")
    conn.execute("DELETE FROM expedientes")
    # Resetear los autoincrement para que los IDs empiecen desde 1
    conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('expedientes','escaneos','actuaciones')")
    conn.commit()
    conn.close()
    return RedirectResponse("/importar?msg=bd_limpiada", status_code=303)


@router.post("/importar", response_class=HTMLResponse)
async def importar_excel(request: Request, archivo: UploadFile = File(...)):
    import openpyxl

    resultado = {
        "insertados": 0,
        "omitidos": 0,
        "errores": [],
        "hoja_usada": "",
    }

    try:
        contenido = await archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

        # Buscar hoja de datos:
        # 1) Primero "ENCABEZADO" (formato estándar del sistema)
        # 2) Si no, la primera hoja cuya celda A1 contenga "EXPEDIENTE"
        # 3) Como último recurso, la primera hoja
        hoja_nombre = None
        if "ENCABEZADO" in wb.sheetnames:
            hoja_nombre = "ENCABEZADO"
        else:
            for nombre in wb.sheetnames:
                primera_celda = wb[nombre].cell(1, 1).value
                if primera_celda and "EXPEDIENTE" in str(primera_celda).upper():
                    hoja_nombre = nombre
                    break
            if not hoja_nombre:
                hoja_nombre = wb.sheetnames[0]

        ws = wb[hoja_nombre]
        resultado["hoja_usada"] = hoja_nombre

        conn = get_db()

        # Saltar fila 1 (encabezados), procesar desde fila 2
        for fila_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas completamente vacías
            if all(v is None for v in row):
                continue

            try:
                datos = _mapear_fila(list(row))

                if not datos.get("n_expediente"):
                    resultado["omitidos"] += 1
                    continue

                # Verificar si ya existe (por n_expediente + anio)
                existe = conn.execute(
                    "SELECT id FROM expedientes WHERE n_expediente = ? AND anio = ?",
                    (datos["n_expediente"], datos["anio"]),
                ).fetchone()

                if existe:
                    resultado["omitidos"] += 1
                    continue

                conn.execute("""
                    INSERT INTO expedientes (
                        n_expediente, anio, mes, origen_proceso, n_radicado,
                        fecha_radicado, fecha_siias, ingreso_siias, ingreso_siad,
                        fecha_ingreso_siad, ingreso_sid4,
                        nombre_abogado, impedimento, investigado, perfil_indagado,
                        entidad_origen, quejoso,
                        asunto, tipologia, descripcion_tipologia,
                        relacionado_siniestro, responsable_siniestro,
                        relacionado_acoso, responsable_acoso,
                        relacionado_corrupcion, valores_institucionales, fecha_hechos,
                        fecha_apertura_indagacion, numero_auto_apertura_ind,
                        fecha_auto_apertura_ind, plazo_ind, fecha_vencimiento_ind,
                        numero_auto_traslado_ind, fecha_auto_traslado_ind,
                        numero_auto_archivo_ind, fecha_auto_archivo_ind,
                        fecha_apertura_investigacion, numero_auto_apertura_inv,
                        fecha_auto_apertura_inv, plazo_inv, fecha_vencimiento_inv,
                        numero_auto_traslado_inv, fecha_auto_traslado_inv,
                        numero_auto_archivo_inv, fecha_auto_archivo_inv,
                        etapa, estado_proceso, observaciones_finales,
                        created_by
                    ) VALUES (
                        :n_expediente, :anio, :mes, :origen_proceso, :n_radicado,
                        :fecha_radicado, :fecha_siias, :ingreso_siias, :ingreso_siad,
                        :fecha_ingreso_siad, :ingreso_sid4,
                        :nombre_abogado, :impedimento, :investigado, :perfil_indagado,
                        :entidad_origen, :quejoso,
                        :asunto, :tipologia, :descripcion_tipologia,
                        :relacionado_siniestro, :responsable_siniestro,
                        :relacionado_acoso, :responsable_acoso,
                        :relacionado_corrupcion, :valores_institucionales, :fecha_hechos,
                        :fecha_apertura_indagacion, :numero_auto_apertura_ind,
                        :fecha_auto_apertura_ind, :plazo_ind, :fecha_vencimiento_ind,
                        :numero_auto_traslado_ind, :fecha_auto_traslado_ind,
                        :numero_auto_archivo_ind, :fecha_auto_archivo_ind,
                        :fecha_apertura_investigacion, :numero_auto_apertura_inv,
                        :fecha_auto_apertura_inv, :plazo_inv, :fecha_vencimiento_inv,
                        :numero_auto_traslado_inv, :fecha_auto_traslado_inv,
                        :numero_auto_archivo_inv, :fecha_auto_archivo_inv,
                        :etapa, :estado_proceso, :observaciones_finales,
                        'Importación Excel'
                    )
                """, datos)

                resultado["insertados"] += 1

            except Exception as e:
                resultado["errores"].append(f"Fila {fila_num}: {str(e)}")

        conn.commit()
        conn.close()

    except Exception as e:
        resultado["errores"].append(f"Error al leer el archivo: {str(e)}")

    conn2 = get_db()
    total = conn2.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]
    conn2.close()
    return templates.TemplateResponse("importar.html", {
        "request": request,
        "active": "importar",
        "resultado": resultado,
        "total_bd": total,
        "msg": "",
    })
