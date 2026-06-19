from fastapi import APIRouter, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date, datetime as _dt
from calendar import monthrange
from typing import Optional
import json
import io
import sqlite3

from app.database import get_db, calcular_alerta, row_to_dict
from app.auth_utils import puede_escribir as _pw, puede_importar as _pi, registrar_log

_MOD = "expedientes"
_ROOT = Path(__file__).parent.parent.parent

router = APIRouter()
templates = make_templates(str(Path(__file__).parent.parent / "templates"))

# ── Datos estáticos ────────────────────────────────────────────────────────────

MESES = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
         "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

MEDIOS_INGRESO = ["SDQS", "CORREO ELECTRONICO", "OFICIO", "VERBAL", "RADICADO", "OTRO"]

ABOGADOS = [
    "MABEL GICELLA HURTADO SANCHEZ",
    "RODOLFO CARRILLO QUINTERO",
    "DAVID FELIPE MORALES NOGUERA",
    "CARLOS ALFONSO PARRA MALAVER",
    "CESAR IVAN RODRIGUEZ DAMIAN",
    "MARA LUCIA UCROS MERLANO",
    "JANIK HERNANDO DE LA HOZ RIOS",
]

TIPOS_EXPEDIENTE = ["FISICO", "ELECTRONICO"]

PERFILES_INVESTIGADO = [
    "ALMACENISTA GENERAL", "ASESOR", "AUXILIAR ADMINISTRATIVO", "AUXILIAR ÁREA SALUD",
    "AUXILIAR DE SERVICIOS GENERALES", "CONDUCTOR", "DIRECTOR", "JEFE DE OFICINA",
    "MÉDICO GENERAL", "PROFESIONAL ESPECIALIZADO", "PROFESIONAL ESPECIALIZADO ÁREA SALUD",
    "PROFESIONAL UNIVERSITARIO", "PROFESIONAL UNIVERSITARIO ÁREA SALUD", "SECRETARIO",
    "SECRETARIO EJECUTIVO", "SUBDIRECTOR", "SUBSECRETARIO DE DESPACHO",
    "TÉCNICO ÁREA SALUD", "TÉCNICO OPERATIVO", "TESORERO GENERAL",
]

TIEMPOS_PRORROGA = ["1", "3", "6"]

ETAPAS = ["INDAGACIÓN PREVIA", "INVESTIGACIÓN DISCIPLINARIA"]

ESTADOS = [
    "ABIERTO", "AUTO DE ARCHIVO", "ACUMULADO", "INCORPORADO",
    "TRASLADADO - PERSONERIA", "TRASLADADO - PROCURADURIA",
    "TRASLADADO - ALCALDIA",
    "TRASLADADO - SUBRED CENTRO ORIENTE E.S.E",
    "TRASLADADO - SUBRED NORTE E.S.E",
    "TRASLADADO - SUBRED SUR E.S.E",
    "TRASLADADO - SUBRED SUR OCCIDENTE E.S.E",
    "PLIEGO DE CARGOS",
]

# Estados que cierran el expediente: a partir de aquí los plazos dejan de "correr".
# Fuente única de verdad — usada por _enriquecer() para que TODAS las vistas
# (lista, dashboard, detalle, exportar-filtrado) coincidan en qué cuenta como "vencido".
ESTADOS_CERRADOS = {"AUTO DE ARCHIVO", "ACUMULADO", "INCORPORADO"}

ENTIDADES_SUGERIDAS = [
    "SDQS", "PERSONERIA DE BOGOTA", "PERSONA PARTICULAR", "ANONIMO", "CONTRALORIA",
]

VALORES_SUGERIDOS = [
    "HONESTIDAD", "RESPETO", "COMPROMISO", "DILIGENCIA",
    "JUSTICIA", "ALTRUISMO", "TODOS",
]

# ── Caché de JSON ──────────────────────────────────────────────────────────────

_tipologias_cache = None
_entidades_cache = None


def _get_tipologias():
    global _tipologias_cache
    if _tipologias_cache is None:
        try:
            path = _ROOT / "Tipologias_Json.txt"
            with open(path, "r", encoding="utf-8") as f:
                _tipologias_cache = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            _tipologias_cache = []
    return _tipologias_cache


def _get_entidades():
    global _entidades_cache
    if _entidades_cache is None:
        try:
            path = _ROOT / "EntidadesDependencias_Json.txt"
            with open(path, "r", encoding="utf-8") as f:
                _entidades_cache = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            _entidades_cache = []
    return _entidades_cache


# ── Helpers de fecha y semáforo ────────────────────────────────────────────────

def _safe_date(s) -> date | None:
    """Acepta 'YYYY-MM-DD' y también 'YYYY-MM-DD HH:MM:SS' (legacy de importaciones anteriores)."""
    if not s:
        return None
    try:
        return date.fromisoformat(str(s).strip()[:10])
    except ValueError:
        return None


def _add_months(d: date, months: int) -> date:
    m = d.month - 1 + months
    year = d.year + m // 12
    month = m % 12 + 1
    day = min(d.day, monthrange(year, month)[1])
    return date(year, month, day)


def _add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return date(d.year + years, 2, 28)


def _parse_flexible_date(s: str) -> date | None:
    if not s:
        return None
    s = s.strip()
    if len(s) == 4 and s.isdigit():
        try:
            return date(int(s), 1, 1)
        except ValueError:
            return None
    elif len(s) == 7 and s[4] == "-":
        try:
            return date(int(s[:4]), int(s[5:]), 1)
        except ValueError:
            return None
    else:
        try:
            return date.fromisoformat(s)
        except ValueError:
            return None


def _enriquecer(exp: dict) -> dict:
    # VENCIMIENTO ETAPA INDAGACIÓN = 6 meses desde FECHA DEL AUTO
    d = _safe_date(exp.get("fecha_auto_apertura_ind"))
    fv_ind = _add_months(d, 6).isoformat() if d else None
    exp["fecha_vencimiento_ind"] = fv_ind
    exp["alerta_ind"] = calcular_alerta(fv_ind)

    # PRESCRIPCION = 5 años desde FECHA DE LOS HECHOS
    d_hechos = _parse_flexible_date(exp.get("fecha_hechos") or "")
    fv_presc = _add_years(d_hechos, 5).isoformat() if d_hechos else None
    exp["fecha_prescripcion"] = fv_presc
    exp["alerta_prescripcion"] = calcular_alerta(fv_presc)

    # VENCIMIENTO ETAPA INVESTIGACIÓN = 6 meses desde FECHA APERTURA INVESTIGACION
    d = _safe_date(exp.get("fecha_apertura_investigacion"))
    fv_inv = _add_months(d, 6).isoformat() if d else None
    exp["fecha_vencimiento_inv"] = fv_inv
    exp["alerta_inv"] = calcular_alerta(fv_inv)

    # VENCIMIENTO PRORROGA = FECHA DE PRORROGA + TIEMPO PRORROGA meses
    fv_prorr = None
    d_prorr = _safe_date(exp.get("fecha_prorroga"))
    if d_prorr and exp.get("tiempo_prorroga"):
        try:
            fv_prorr = _add_months(d_prorr, int(exp["tiempo_prorroga"])).isoformat()
        except (ValueError, TypeError):
            pass
    exp["fecha_vencimiento_prorroga"] = fv_prorr
    exp["alerta_prorroga"] = calcular_alerta(fv_prorr)

    # Si el expediente ya está cerrado, el plazo dejó de correr: no debe contar
    # como "vencido" ni "próximo" en ninguna vista (lista, dashboard, export).
    if (exp.get("estado_proceso") or "").strip().upper() in ESTADOS_CERRADOS:
        _cerrado = {"dias": None, "clase": "sin-plazo", "texto": "Cerrado — plazo no aplica"}
        exp["alerta_ind"] = dict(_cerrado)
        exp["alerta_inv"] = dict(_cerrado)
        exp["alerta_prescripcion"] = dict(_cerrado)
        exp["alerta_prorroga"] = dict(_cerrado)

    return exp


def _limpiar(v):
    if v is None or str(v).strip() == "":
        return None
    return str(v).strip()


def _next_n_expediente(conn) -> str:
    row = conn.execute("SELECT MAX(CAST(n_expediente AS INTEGER)) FROM expedientes").fetchone()
    siguiente = (row[0] or 0) + 1
    return f"{siguiente:03d}"


def _abogados_activos() -> list[str]:
    """Lee los abogados desde la tabla usuarios (rol='abogado'), fuente de verdad
    administrada en Usuarios y Permisos. ABOGADOS queda como respaldo estático
    solo por si la tabla aún no tiene abogados registrados (instalación nueva)."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT nombre_completo FROM usuarios WHERE rol='abogado' AND activo=1 ORDER BY nombre_completo"
        ).fetchall()
    finally:
        conn.close()
    return [r[0] for r in rows] or ABOGADOS


def _ctx_base():
    return {
        "meses": MESES,
        "medios_ingreso": MEDIOS_INGRESO,
        "abogados": _abogados_activos(),
        "tipos_expediente": TIPOS_EXPEDIENTE,
        "perfiles_investigado": PERFILES_INVESTIGADO,
        "tiempos_prorroga": TIEMPOS_PRORROGA,
        "etapas": ETAPAS,
        "estados": ESTADOS,
        "entidades_sugeridas": ENTIDADES_SUGERIDAS,
        "valores_sugeridos": VALORES_SUGERIDOS,
        "tipologias_json": json.dumps(_get_tipologias(), ensure_ascii=False),
        "entidades_json": json.dumps(_get_entidades(), ensure_ascii=False),
    }


# ── Listado ────────────────────────────────────────────────────────────────────

@router.get("/expedientes", response_class=HTMLResponse)
async def lista_expedientes(
    request: Request,
    q: str = "",
    anio: str = "",
    mes: str = "",
    abogado: str = "",
    etapa: str = "",
    estado: str = "",
    alerta: str = "",
    page: int = 1,
    por_pagina: int = 50,
    msg: str = "",
):
    conn = get_db()
    filtros = []
    params = []

    if q:
        filtros.append("""(
            n_expediente LIKE ? OR CAST(n_expediente AS INTEGER) = ?
            OR nombre_investigado LIKE ? OR asunto LIKE ?
            OR n_radicado LIKE ? OR quejoso LIKE ?
            OR entidad_origen LIKE ?
        )""")
        q_int = int(q.strip()) if q.strip().isdigit() else -1
        params += [f"%{q}%", q_int, f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"]
    if anio:
        filtros.append("anio = ?")
        params.append(int(anio))
    if mes:
        filtros.append("mes = ?")
        params.append(mes)
    if abogado:
        filtros.append("abogado_asignado LIKE ?")
        params.append(f"%{abogado}%")
    if etapa:
        filtros.append("etapa_actual = ?")
        params.append(etapa)
    if estado:
        filtros.append("estado_proceso = ?")
        params.append(estado)

    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    total = conn.execute(f"SELECT COUNT(*) FROM expedientes {where}", params).fetchone()[0]
    total_pages = max(1, (total + por_pagina - 1) // por_pagina)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * por_pagina

    rows_raw = conn.execute(
        f"SELECT * FROM expedientes {where} ORDER BY CAST(n_expediente AS INTEGER) DESC LIMIT ? OFFSET ?",
        params + [por_pagina, offset],
    ).fetchall()

    anios_list = [r[0] for r in conn.execute(
        "SELECT DISTINCT anio FROM expedientes WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()]
    abogados_list = [r[0] for r in conn.execute(
        "SELECT DISTINCT abogado_asignado FROM expedientes WHERE abogado_asignado IS NOT NULL ORDER BY abogado_asignado"
    ).fetchall()]
    conn.close()

    rows = [_enriquecer(dict(r)) for r in rows_raw]

    # Filtro por alerta (post-enriquecimiento)
    if alerta:
        rows = [r for r in rows if (
            r["alerta_ind"]["clase"] == alerta or
            r["alerta_inv"]["clase"] == alerta or
            r["alerta_prescripcion"]["clase"] == alerta
        )]

    return templates.TemplateResponse("lista.html", {
        "request": request,
        "rows": rows,
        "total": total,
        "page": page,
        "total_pages": total_pages,
        "por_pagina": por_pagina,
        "q": q, "anio": anio, "mes": mes,
        "abogado": abogado, "etapa": etapa,
        "estado": estado, "alerta": alerta,
        "anios_list": anios_list,
        "abogados_list": abogados_list,
        "etapas": ETAPAS,
        "estados": ESTADOS,
        "meses": MESES,
        "msg": msg,
        "active": "lista",
    })


# ── Nuevo expediente ───────────────────────────────────────────────────────────

@router.get("/expediente/nuevo", response_class=HTMLResponse)
async def nuevo_form(request: Request):
    conn = get_db()
    proximo = _next_n_expediente(conn)
    conn.close()
    ctx = _ctx_base()
    ctx.update({
        "request": request,
        "r": {},
        "proximo_n": proximo,
        "modo": "nuevo",
        "active": "nuevo",
    })
    return templates.TemplateResponse("form.html", ctx)


@router.post("/expediente/nuevo", response_class=HTMLResponse)
async def nuevo_post(request: Request):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/expedientes?msg=sin_permiso", status_code=303)

    form = await request.form()

    def f(key):
        return _limpiar(form.get(key))

    conn = get_db()
    campos = [
        "n_expediente", "anio", "mes", "medio_ingreso", "n_radicado",
        "fecha_radicado", "abogado_asignado", "entidad_origen", "quejoso",
        "asunto", "impedimento", "fecha_apertura_expediente",
        "numero_auto_apertura_ind", "fecha_auto_apertura_ind", "tipo_expediente",
        "tipologia", "relacionado_siniestro", "responsable_siniestro",
        "relacionado_maltrato", "relacionado_corrupcion", "valores_institucionales",
        "fecha_hechos_obs", "fecha_hechos",
        "fecha_ultima_act_indagacion", "numero_auto_ultima_act_ind",
        "fecha_apertura_investigacion", "numero_auto_apertura_inv",
        "nombre_investigado", "cedula", "perfil_investigado", "area_origen_investigado",
        "fecha_prorroga", "numero_auto_prorroga", "tiempo_prorroga",
        "fecha_ultima_act_investigacion", "numero_auto_ultima_act_inv",
        "numero_auto_traslado", "fecha_auto_traslado",
        "numero_auto_acumulacion", "fecha_auto_acumulacion", "expediente_acumula",
        "fecha_auto_archivo", "numero_auto_archivo",
        "fecha_auto_pliego_cargos", "numero_auto_pliego_cargos",
        "etapa_actual", "estado_proceso", "observaciones",
    ]
    vals = []
    for c in campos:
        v = f(c)
        if c == "anio" and v:
            try:
                v = int(v)
            except ValueError:
                v = None
        vals.append(v)

    n_exp = f("n_expediente") or ""
    anio_val = f("anio") or ""
    try:
        conn.execute(
            f"INSERT INTO expedientes ({', '.join(campos)}, created_by) VALUES ({', '.join(['?']*len(campos))}, ?)",
            vals + [user.get("nombre_completo") if user else None],
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return RedirectResponse(f"/expedientes?msg=duplicado_{n_exp}_{anio_val}", status_code=303)
    conn.close()
    registrar_log(user, "CREAR", _MOD, f"Expediente {n_exp}")
    return RedirectResponse(f"/expedientes?msg=creado", status_code=303)


# ── Detalle ────────────────────────────────────────────────────────────────────

@router.get("/expediente/{exp_id}", response_class=HTMLResponse)
async def detalle(request: Request, exp_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM expedientes WHERE id = ?", (exp_id,)).fetchone()
    conn.close()
    if not row:
        return RedirectResponse("/expedientes?msg=no_encontrado", status_code=303)
    exp = _enriquecer(dict(row))
    ctx = _ctx_base()
    ctx.update({
        "request": request,
        "r": exp,
        "modo": "ver",
        "active": "lista",
    })
    return templates.TemplateResponse("form.html", ctx)


# ── Editar ─────────────────────────────────────────────────────────────────────

@router.get("/expediente/{exp_id}/editar", response_class=HTMLResponse)
async def editar_form(request: Request, exp_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM expedientes WHERE id = ?", (exp_id,)).fetchone()
    conn.close()
    if not row:
        return RedirectResponse("/expedientes?msg=no_encontrado", status_code=303)
    exp = _enriquecer(dict(row))
    ctx = _ctx_base()
    ctx.update({
        "request": request,
        "r": exp,
        "modo": "editar",
        "active": "lista",
    })
    return templates.TemplateResponse("form.html", ctx)


@router.post("/expediente/{exp_id}/editar", response_class=HTMLResponse)
async def editar_post(request: Request, exp_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/expediente/{exp_id}?msg=sin_permiso", status_code=303)

    form = await request.form()

    def f(key):
        return _limpiar(form.get(key))

    campos = [
        "n_expediente", "anio", "mes", "medio_ingreso", "n_radicado",
        "fecha_radicado", "abogado_asignado", "entidad_origen", "quejoso",
        "asunto", "impedimento", "fecha_apertura_expediente",
        "numero_auto_apertura_ind", "fecha_auto_apertura_ind", "tipo_expediente",
        "tipologia", "relacionado_siniestro", "responsable_siniestro",
        "relacionado_maltrato", "relacionado_corrupcion", "valores_institucionales",
        "fecha_hechos_obs", "fecha_hechos",
        "fecha_ultima_act_indagacion", "numero_auto_ultima_act_ind",
        "fecha_apertura_investigacion", "numero_auto_apertura_inv",
        "nombre_investigado", "cedula", "perfil_investigado", "area_origen_investigado",
        "fecha_prorroga", "numero_auto_prorroga", "tiempo_prorroga",
        "fecha_ultima_act_investigacion", "numero_auto_ultima_act_inv",
        "numero_auto_traslado", "fecha_auto_traslado",
        "numero_auto_acumulacion", "fecha_auto_acumulacion", "expediente_acumula",
        "fecha_auto_archivo", "numero_auto_archivo",
        "fecha_auto_pliego_cargos", "numero_auto_pliego_cargos",
        "etapa_actual", "estado_proceso", "observaciones",
    ]
    set_clause = ", ".join(f"{c} = ?" for c in campos)
    vals = []
    for c in campos:
        v = f(c)
        if c == "anio" and v:
            try:
                v = int(v)
            except ValueError:
                v = None
        vals.append(v)

    n_exp = f("n_expediente") or str(exp_id)
    anio_val = f("anio") or ""
    conn = get_db()
    try:
        conn.execute(
            f"UPDATE expedientes SET {set_clause}, updated_at = datetime('now','localtime') WHERE id = ?",
            vals + [exp_id],
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return RedirectResponse(f"/expedientes?msg=duplicado_{n_exp}_{anio_val}", status_code=303)
    conn.close()
    registrar_log(user, "EDITAR", _MOD, f"Expediente {n_exp}")
    return RedirectResponse(f"/expedientes?msg=actualizado", status_code=303)


# ── Eliminar ───────────────────────────────────────────────────────────────────

@router.post("/expediente/{exp_id}/eliminar")
async def eliminar(request: Request, exp_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/expedientes?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute("SELECT n_expediente FROM expedientes WHERE id = ?", (exp_id,)).fetchone()
    n_exp = row["n_expediente"] if row else str(exp_id)
    conn.execute("DELETE FROM expedientes WHERE id = ?", (exp_id,))
    conn.commit()
    conn.close()
    registrar_log(user, "ELIMINAR", _MOD, f"Expediente {n_exp}")
    return RedirectResponse(f"/expedientes?msg=eliminado_{n_exp}", status_code=303)


# ── Exportar — Página de personalización ──────────────────────────────────────

@router.get("/exportar-filtrado", response_class=HTMLResponse)
async def exportar_filtrado_page(request: Request):
    conn = get_db()
    anios_list = [r[0] for r in conn.execute(
        "SELECT DISTINCT anio FROM expedientes WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()]
    total_preview = conn.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]
    conn.close()
    filtros = {
        "anios": [], "abogados": [], "etapas": [], "estados": [],
        "fecha_desde": "", "fecha_hasta": "",
        "solo_vencidos": False, "proximos_30": False, "proximos_60": False,
        "bloques_off": [],
    }
    return templates.TemplateResponse("exportar_filtrado.html", {
        "request": request,
        "active": "exportar",
        "anios": anios_list,
        "abogados": _abogados_activos(),
        "etapas": ETAPAS,
        "estados": ESTADOS,
        "filtros": filtros,
        "total_preview": total_preview,
    })


# ── Exportar — Descarga Excel ──────────────────────────────────────────────────

@router.get("/exportar-filtrado/descargar")
async def exportar_descargar(
    request: Request,
    q: str = "",
    anio: str = "",
    mes: str = "",
    abogado: str = "",
    etapa: str = "",
    estado: str = "",
    solo_vencidos: str = "",
    proximos_30: str = "",
    proximos_60: str = "",
    fecha_desde: str = "",
    fecha_hasta: str = "",
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/expedientes?msg=error_openpyxl")

    anios_sel    = request.query_params.getlist("anios")
    abogados_sel = request.query_params.getlist("abogados")
    etapas_sel   = request.query_params.getlist("etapas")
    estados_sel  = request.query_params.getlist("estados")
    bloques      = request.query_params.getlist("bloques")
    if not bloques:
        bloques = ["identificacion", "partes", "asunto", "indagacion", "investigacion", "cierre"]

    conn = get_db()
    filtros_sql, params = [], []

    # Filtros simples (compat con lista.html y dashboard)
    if q:
        filtros_sql.append("(n_expediente LIKE ? OR nombre_investigado LIKE ? OR asunto LIKE ? OR n_radicado LIKE ?)")
        params += [f"%{q}%"] * 4
    if anio:
        filtros_sql.append("anio = ?"); params.append(int(anio))
    if mes:
        filtros_sql.append("mes = ?"); params.append(mes)
    if abogado:
        filtros_sql.append("abogado_asignado LIKE ?"); params.append(f"%{abogado}%")
    if etapa:
        filtros_sql.append("etapa_actual = ?"); params.append(etapa)
    if estado:
        filtros_sql.append("estado_proceso = ?"); params.append(estado)

    # Filtros multi-select (formulario personalizado)
    if anios_sel:
        filtros_sql.append(f"anio IN ({','.join(['?']*len(anios_sel))})")
        params += [int(a) for a in anios_sel]
    if abogados_sel:
        filtros_sql.append(f"abogado_asignado IN ({','.join(['?']*len(abogados_sel))})")
        params += abogados_sel
    if etapas_sel:
        filtros_sql.append(f"etapa_actual IN ({','.join(['?']*len(etapas_sel))})")
        params += etapas_sel
    if estados_sel:
        filtros_sql.append(f"estado_proceso IN ({','.join(['?']*len(estados_sel))})")
        params += estados_sel
    if fecha_desde:
        filtros_sql.append("fecha_radicado >= ?"); params.append(fecha_desde)
    if fecha_hasta:
        filtros_sql.append("fecha_radicado <= ?"); params.append(fecha_hasta)

    where = ("WHERE " + " AND ".join(filtros_sql)) if filtros_sql else ""
    rows = conn.execute(
        f"SELECT * FROM expedientes {where} ORDER BY anio, CAST(n_expediente AS INTEGER)",
        params,
    ).fetchall()
    conn.close()

    datos = [_enriquecer(dict(r)) for r in rows]

    # Filtros de alerta — misma fuente de verdad que la pantalla de Lista
    # (clases calculadas por _enriquecer/calcular_alerta, que ya excluye
    # expedientes cerrados y cubre indagación + investigación + prescripción).
    def _alertas(d):
        return (d["alerta_ind"], d["alerta_inv"], d["alerta_prescripcion"])

    if solo_vencidos:
        datos = [d for d in datos if any(a["clase"] == "vencido" for a in _alertas(d))]
    if proximos_30:
        datos = [d for d in datos if any(
            a["dias"] is not None and 0 <= a["dias"] <= 30 for a in _alertas(d)
        )]
    if proximos_60:
        datos = [d for d in datos if any(
            a["dias"] is not None and 0 <= a["dias"] <= 60 for a in _alertas(d)
        )]

    # Columnas según bloques seleccionados
    BLOQUES_DEF = {
        "identificacion": (
            ["N. EXPEDIENTE","AÑO","MES","MEDIO DE INGRESO","N. RADICADO",
             "FECHA RADICADO","ABOGADO ASIGNADO","IMPEDIMENTO","TIPO EXPEDIENTE"],
            ["n_expediente","anio","mes","medio_ingreso","n_radicado",
             "fecha_radicado","abogado_asignado","impedimento","tipo_expediente"],
        ),
        "partes": (
            ["NOMBRE INVESTIGADO","CÉDULA","PERFIL INVESTIGADO","ÁREA ORIGEN INVESTIGADO",
             "ENTIDAD ORIGEN","QUEJOSO"],
            ["nombre_investigado","cedula","perfil_investigado","area_origen_investigado",
             "entidad_origen","quejoso"],
        ),
        "asunto": (
            ["ASUNTO","TIPOLOGÍA","VALORES INSTITUCIONALES","FECHA HECHOS (OBS.)","FECHA HECHOS",
             "PRESCRIPCIÓN (calc.)","REL. SINIESTRO","RESP. SINIESTRO",
             "REL. MALTRATO/ACOSO","REL. CORRUPCIÓN"],
            ["asunto","tipologia","valores_institucionales","fecha_hechos_obs","fecha_hechos",
             "fecha_prescripcion","relacionado_siniestro","responsable_siniestro",
             "relacionado_maltrato","relacionado_corrupcion"],
        ),
        "indagacion": (
            ["F. APERTURA EXPEDIENTE","N. AUTO APERTURA IND.","F. AUTO APERTURA IND.",
             "VENCIMIENTO IND. (calc.)","F. ÚLTIMA ACT. IND.","N. AUTO ÚLTIMA ACT. IND."],
            ["fecha_apertura_expediente","numero_auto_apertura_ind","fecha_auto_apertura_ind",
             "fecha_vencimiento_ind","fecha_ultima_act_indagacion","numero_auto_ultima_act_ind"],
        ),
        "investigacion": (
            ["F. APERTURA INV.","N. AUTO APERTURA INV.","VENCIMIENTO INV. (calc.)",
             "F. ÚLTIMA ACT. INV.","N. AUTO ÚLTIMA ACT. INV.",
             "F. PRÓRROGA","N. AUTO PRÓRROGA","TIEMPO PRÓRROGA (meses)","VENCIMIENTO PRÓRROGA (calc.)"],
            ["fecha_apertura_investigacion","numero_auto_apertura_inv","fecha_vencimiento_inv",
             "fecha_ultima_act_investigacion","numero_auto_ultima_act_inv",
             "fecha_prorroga","numero_auto_prorroga","tiempo_prorroga","fecha_vencimiento_prorroga"],
        ),
        "cierre": (
            ["N. AUTO TRASLADO","F. AUTO TRASLADO","N. AUTO ACUMULACIÓN","F. AUTO ACUMULACIÓN",
             "EXP. ACUMULA","F. AUTO ARCHIVO","N. AUTO ARCHIVO",
             "F. AUTO PLIEGO CARGOS","N. AUTO PLIEGO CARGOS",
             "ETAPA ACTUAL","ESTADO DEL PROCESO","OBSERVACIONES"],
            ["numero_auto_traslado","fecha_auto_traslado","numero_auto_acumulacion","fecha_auto_acumulacion",
             "expediente_acumula","fecha_auto_archivo","numero_auto_archivo",
             "fecha_auto_pliego_cargos","numero_auto_pliego_cargos",
             "etapa_actual","estado_proceso","observaciones"],
        ),
    }

    headers_out, campos_out = [], []
    for bloque in ["identificacion","partes","asunto","indagacion","investigacion","cierre"]:
        if bloque in bloques:
            h, c = BLOQUES_DEF[bloque]
            headers_out += h
            campos_out  += c

    if not headers_out:
        headers_out = ["N. EXPEDIENTE","AÑO","ETAPA ACTUAL","ESTADO DEL PROCESO"]
        campos_out  = ["n_expediente","anio","etapa_actual","estado_proceso"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base Expedientes"
    h_font   = Font(bold=True, color="FFFFFF", size=10)
    fill_h   = PatternFill("solid", fgColor="1B4F8A")
    alt_fill = PatternFill("solid", fgColor="EBF1F8")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, h in enumerate(headers_out, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = fill_h; cell.font = h_font; cell.alignment = center
    ws.row_dimensions[1].height = 42

    for ri, d in enumerate(datos, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, campo in enumerate(campos_out, 1):
            cell = ws.cell(row=ri, column=ci, value=d.get(campo))
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    hoy_str = date.today().strftime("%Y%m%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=BaseExpedientes_{hoy_str}.xlsx"},
    )


# ── Importar Excel ─────────────────────────────────────────────────────────────

@router.get("/importar", response_class=HTMLResponse)
async def importar_form(request: Request, msg: str = ""):
    user = getattr(request.state, "user", None)
    if not _pi(user, _MOD):
        return RedirectResponse("/expedientes?msg=sin_permiso", status_code=303)
    conn = get_db()
    total_bd = conn.execute("SELECT COUNT(*) FROM expedientes").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("importar.html", {
        "request": request,
        "active": "importar",
        "msg": msg,
        "total_bd": total_bd,
    })


@router.post("/importar", response_class=HTMLResponse)
async def importar_post(request: Request):
    from fastapi import UploadFile, File
    user = getattr(request.state, "user", None)
    if not _pi(user, _MOD):
        return RedirectResponse("/expedientes?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/importar?msg=error_openpyxl", status_code=303)

    form = await request.form()
    archivo = form.get("archivo")
    if not archivo or not archivo.filename:
        return RedirectResponse("/importar?msg=error_vacio", status_code=303)

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        return RedirectResponse("/importar?msg=error_archivo", status_code=303)

    # Mapeo encabezado Excel → campo BD.
    # Cubre el formato del backup general Y el del exportador individual (bloques filtrados).
    # Los encabezados calculados (PRESCRIPCIÓN, VENCIMIENTO calc.) no están en el mapa → se ignoran.
    HEADER_CAMPO = {
        "N. EXPEDIENTE": "n_expediente",
        "AÑO": "anio",
        "MES": "mes",
        "MEDIO DE INGRESO": "medio_ingreso",
        "N. RADICADO": "n_radicado",
        "FECHA RADICADO": "fecha_radicado",
        "ABOGADO ASIGNADO": "abogado_asignado",
        "ENTIDAD ORIGEN": "entidad_origen",
        "QUEJOSO": "quejoso",
        "ASUNTO": "asunto",
        "IMPEDIMENTO": "impedimento",
        "FECHA APERTURA EXPEDIENTE": "fecha_apertura_expediente",
        "F. APERTURA EXPEDIENTE": "fecha_apertura_expediente",
        "N. AUTO APERTURA INDAGACIÓN": "numero_auto_apertura_ind",
        "N. AUTO APERTURA IND.": "numero_auto_apertura_ind",
        "FECHA AUTO APERTURA IND.": "fecha_auto_apertura_ind",
        "F. AUTO APERTURA IND.": "fecha_auto_apertura_ind",
        "TIPO EXPEDIENTE": "tipo_expediente",
        "TIPOLOGIA": "tipologia",
        "TIPOLOGÍA": "tipologia",
        "RELACIONADO SINIESTRO": "relacionado_siniestro",
        "REL. SINIESTRO": "relacionado_siniestro",
        "RESPONSABLE SINIESTRO": "responsable_siniestro",
        "RESP. SINIESTRO": "responsable_siniestro",
        "RELACIONADO MALTRATO": "relacionado_maltrato",
        "REL. MALTRATO/ACOSO": "relacionado_maltrato",
        "RELACIONADO CORRUPCIÓN": "relacionado_corrupcion",
        "RELACIONADO CORRUPCION": "relacionado_corrupcion",
        "REL. CORRUPCIÓN": "relacionado_corrupcion",
        "VALORES INSTITUCIONALES": "valores_institucionales",
        "FECHA HECHOS (OBSERVACIONES)": "fecha_hechos_obs",
        "FECHA HECHOS (OBS.)": "fecha_hechos_obs",
        "FECHA HECHOS": "fecha_hechos",
        "F. ÚLTIMA ACTUACIÓN INDAGACIÓN": "fecha_ultima_act_indagacion",
        "F. ÚLTIMA ACT. IND.": "fecha_ultima_act_indagacion",
        "N. AUTO ÚLTIMA ACT. IND.": "numero_auto_ultima_act_ind",
        "F. APERTURA INVESTIGACIÓN": "fecha_apertura_investigacion",
        "F. APERTURA INV.": "fecha_apertura_investigacion",
        "N. AUTO APERTURA INV.": "numero_auto_apertura_inv",
        "F. ÚLTIMA ACTUACIÓN INVESTIGACIÓN": "fecha_ultima_act_investigacion",
        "F. ÚLTIMA ACT. INV.": "fecha_ultima_act_investigacion",
        "N. AUTO ÚLTIMA ACT. INV.": "numero_auto_ultima_act_inv",
        "NOMBRE INVESTIGADO": "nombre_investigado",
        "CÉDULA": "cedula",
        "CEDULA": "cedula",
        "PERFIL INVESTIGADO": "perfil_investigado",
        "ÁREA ORIGEN INVESTIGADO": "area_origen_investigado",
        "AREA ORIGEN INVESTIGADO": "area_origen_investigado",
        "FECHA PRÓRROGA": "fecha_prorroga",
        "F. PRÓRROGA": "fecha_prorroga",
        "N. AUTO PRÓRROGA": "numero_auto_prorroga",
        "TIEMPO PRÓRROGA": "tiempo_prorroga",
        "TIEMPO PRÓRROGA (MESES)": "tiempo_prorroga",
        "N. AUTO TRASLADO": "numero_auto_traslado",
        "F. AUTO TRASLADO": "fecha_auto_traslado",
        "N. AUTO ACUMULACIÓN": "numero_auto_acumulacion",
        "F. AUTO ACUMULACIÓN": "fecha_auto_acumulacion",
        "EXPEDIENTE ACUMULA": "expediente_acumula",
        "EXP. ACUMULA": "expediente_acumula",
        "F. AUTO ARCHIVO": "fecha_auto_archivo",
        "N. AUTO ARCHIVO": "numero_auto_archivo",
        "F. AUTO PLIEGO CARGOS": "fecha_auto_pliego_cargos",
        "N. AUTO PLIEGO CARGOS": "numero_auto_pliego_cargos",
        "ETAPA ACTUAL": "etapa_actual",
        "ESTADO DEL PROCESO": "estado_proceso",
        "OBSERVACIONES": "observaciones",
        "CREADO POR": "created_by",
        "FECHA CREACIÓN": "created_at",
        "ÚLTIMA ACTUALIZACIÓN": "updated_at",
    }

    DB_CAMPOS = set(HEADER_CAMPO.values())

    def _v(val):
        if val is None:
            return None
        # openpyxl devuelve datetime para celdas formateadas como fecha
        if isinstance(val, _dt):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, date):
            return val.isoformat()
        s = str(val).strip()
        return None if s in ("", "nan", "None", "#VALUE!", "#N/A", "#REF!", "—") else s

    ws_name = "Base Expedientes"
    if ws_name not in wb.sheetnames:
        return RedirectResponse("/importar?msg=error_hoja", status_code=303)

    ws = wb[ws_name]

    # Construir mapa col_index → campo usando la fila de encabezados
    header_row = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
    col_campo = {}
    for ci, h in enumerate(header_row):
        campo = HEADER_CAMPO.get(h)
        if campo and campo in DB_CAMPOS:
            col_campo[ci] = campo

    conn = get_db()

    # Respaldo de Seguimiento Mensual ANTES del wipe: el DELETE de abajo dispara
    # ON DELETE CASCADE sobre seguimiento_mensual (FK a expedientes.id). Como el
    # reimport asigna IDs nuevos, sin este respaldo se pierde TODO el historial
    # de seguimiento de TODOS los expedientes en cada reimportación.
    seguimiento_backup = [dict(r) for r in conn.execute("""
        SELECT e.n_expediente, e.anio AS exp_anio,
               sm.anio AS sm_anio, sm.mes, sm.descripcion, sm.created_by
        FROM seguimiento_mensual sm
        JOIN expedientes e ON e.id = sm.expediente_id
    """).fetchall()]

    conn.execute("DELETE FROM expedientes")
    count = 0
    nuevo_id_por_clave: dict[tuple, int] = {}
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in row if v is not None):
                continue
            datos = {campo: _v(row[ci]) if ci < len(row) else None
                     for ci, campo in col_campo.items()}
            if not datos.get("n_expediente"):
                continue
            campos_ins = list(datos.keys())
            cur = conn.execute(
                f"INSERT INTO expedientes ({', '.join(campos_ins)}) VALUES ({', '.join(['?']*len(campos_ins))})",
                [datos[c] for c in campos_ins],
            )
            nuevo_id_por_clave[(datos.get("n_expediente"), datos.get("anio"))] = cur.lastrowid
            count += 1

        # Restaurar seguimiento mensual para los expedientes que sigan presentes
        # en el Excel reimportado (mismo n_expediente + año), re-vinculado al nuevo id.
        restaurados = 0
        for sb in seguimiento_backup:
            nuevo_id = nuevo_id_por_clave.get((sb["n_expediente"], sb["exp_anio"]))
            if nuevo_id:
                conn.execute(
                    """INSERT INTO seguimiento_mensual
                           (expediente_id, anio, mes, descripcion, created_by)
                       VALUES (?,?,?,?,?)
                       ON CONFLICT(expediente_id, anio, mes) DO UPDATE SET
                           descripcion = excluded.descripcion,
                           created_by  = excluded.created_by""",
                    (nuevo_id, sb["sm_anio"], sb["mes"], sb["descripcion"], sb["created_by"]),
                )
                restaurados += 1

        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return RedirectResponse("/importar?msg=error_import", status_code=303)
    conn.close()
    registrar_log(user, "IMPORTAR", _MOD, f"{count} expedientes, {restaurados} seguimientos preservados")
    return RedirectResponse(f"/expedientes?msg=importado_{count}", status_code=303)


@router.get("/autos", response_class=HTMLResponse)
async def autos_redirect(request: Request):
    return RedirectResponse("/control-autos/", status_code=302)
