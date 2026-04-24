from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date, timedelta
import io

from urllib.parse import quote_plus as _quote_plus

from app.database import get_db
from app.auth_utils import tpl, puede_escribir as _pw, registrar_log

_MOD = "correspondencia"

router = APIRouter(prefix="/correspondencia")
templates = make_templates(str(Path(__file__).parent.parent / "templates"))
templates.env.filters["quote_plus"] = _quote_plus

MESES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

TIPOS_RESPUESTA = [
    "RESPUESTA",
    "TRASLADO",
    "ANEXO EXPEDIENTE",
    "ANEXO AL EXPEDIENTE",
    "DEVOLUCION",
    "INFORMATIVO",
    "REUNION",
    "APERTURA EXPEDIENTE",
    "AUTO INHIBITORIO",
    "ANTECEDENTES",
    "RESPUESTA CORREO ELECTRONICO",
]

TERMINOS_DIAS = [3, 5, 10, 15, 30]

ABOGADOS_RESPONSABLES = [
    "ANDRES EDUARDO SANDOVAL MAYORGA",
    "CARLOS ALFONSO PARRA MALAVER",
    "CESAR IVAN RODRIGUEZ DAMIAN",
    "DAVID FELIPE MORALES NOGUERA",
    "JANIK HERNANDO DE LA HOZ RIOS",
    "JOSE DE JESUS BARAJAS SOTELO",
    "LUNA GICELL GUZMAN YATE",
    "MABEL GICELLA HURTADO SANCHEZ",
    "MAGDA XIMENA PAREDES LIEVANO",
    "MARA LUCIA UCROS MERLANO",
    "MARTHA PATRICIA AÑEZ MAESTRE",
    "RODOLFO CARRILLO QUINTERO",
]

TIPOS_REQUERIMIENTO = [
    "DERECHO DE PETICION",
    "TUTELA",
    "PROPOSICION DEL CONSEJO",
    "REQUERIMIENTO ENTES DE CONTROL",
    "COMUNICACION INTERNA",
    "COMUNICACION EXTERNA",
]

# Mapa de limpieza de nombres para importación histórica
RESPONSABLE_MAP = {
    "ANDRES SANDOVAL": "ANDRES EDUARDO SANDOVAL MAYORGA",
    "ANDRES EDUARDO SANDOVAL": "ANDRES EDUARDO SANDOVAL MAYORGA",
    "CARLOS PARRA": "CARLOS ALFONSO PARRA MALAVER",
    "CARLOS ALFONSO PARRA": "CARLOS ALFONSO PARRA MALAVER",
    "CESAR IVAN": "CESAR IVAN RODRIGUEZ DAMIAN",
    "CESAR IVAN RODRIGUEZ": "CESAR IVAN RODRIGUEZ DAMIAN",
    "CESAR RODRIGUEZ": "CESAR IVAN RODRIGUEZ DAMIAN",
    "DAVID FELIPE  MORALES": "DAVID FELIPE MORALES NOGUERA",
    "DAVID FELIPE MORALES": "DAVID FELIPE MORALES NOGUERA",
    "DAVID MORALES": "DAVID FELIPE MORALES NOGUERA",
    "DE LA HOZ": "JANIK HERNANDO DE LA HOZ RIOS",
    "JANIK DE LA HOZ": "JANIK HERNANDO DE LA HOZ RIOS",
    "JANIK HERNANDO DE LA HOZ": "JANIK HERNANDO DE LA HOZ RIOS",
    "JOSE BARAJAS": "JOSE DE JESUS BARAJAS SOTELO",
    "LUNA GUZMAN": "LUNA GICELL GUZMAN YATE",
    "LUNA GICELL GUZMAN": "LUNA GICELL GUZMAN YATE",
    "MABEL HURTADO": "MABEL GICELLA HURTADO SANCHEZ",
    "MABEL GICELLA HURTADO": "MABEL GICELLA HURTADO SANCHEZ",
    "GICELLA HURTADO": "MABEL GICELLA HURTADO SANCHEZ",
    "MABEL GICELA HURTADO SANCHEZ": "MABEL GICELLA HURTADO SANCHEZ",
    "MAGDA PAREDES": "MAGDA XIMENA PAREDES LIEVANO",
    "MAGDA XIMENA PAREDES": "MAGDA XIMENA PAREDES LIEVANO",
    "MARA UCROS": "MARA LUCIA UCROS MERLANO",
    "MARA LUCIA UCROS": "MARA LUCIA UCROS MERLANO",
    "MARTHA AÑEZ": "MARTHA PATRICIA AÑEZ MAESTRE",
    "MARTHA PATRICIA AÑEZ": "MARTHA PATRICIA AÑEZ MAESTRE",
    "RODOLFO CARRILLO": "RODOLFO CARRILLO QUINTERO",
    "PROFESIONALES": "TODOS LOS PROFESIONALES",
    "TODOS LO PROFESIONALES": "TODOS LOS PROFESIONALES",
    "TODOS LOS PROFESIONALES": "TODOS LOS PROFESIONALES",
}


def _v(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "None", "#VALUE!", "#N/A", "#REF!", "—"):
        return None
    return s


def _clean_responsable(nombre: str | None) -> str | None:
    if not nombre:
        return None
    n = str(nombre).strip().upper()
    return RESPONSABLE_MAP.get(n, n.title() if n.islower() else n)


def _anios_disponibles():
    return list(range(2024, date.today().year + 3))


def _get_catalogos(conn):
    tipos_doc = [r[0] for r in conn.execute(
        "SELECT nombre FROM corr_tipos_documento ORDER BY nombre"
    ).fetchall()]
    return ABOGADOS_RESPONSABLES, tipos_doc


# ── Días hábiles Colombia ──────────────────────────────────────────────────────

def _easter(year: int) -> date:
    """Gauss algorithm for Easter Sunday."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _next_monday(d: date) -> date:
    """Return d if Monday, else advance to next Monday."""
    days_ahead = (7 - d.weekday()) % 7
    return d if days_ahead == 0 else d + timedelta(days=days_ahead)


def _festivos_colombia(year: int) -> set:
    festivos = set()
    for m, day in [(1, 1), (5, 1), (7, 20), (8, 7), (12, 8), (12, 25)]:
        festivos.add(date(year, m, day))
    for m, day in [(1, 6), (3, 19), (6, 29), (8, 15), (10, 12), (11, 1), (11, 11)]:
        festivos.add(_next_monday(date(year, m, day)))
    easter = _easter(year)
    festivos.add(easter - timedelta(days=3))   # Jueves Santo
    festivos.add(easter - timedelta(days=2))   # Viernes Santo
    festivos.add(_next_monday(easter + timedelta(days=39)))   # Ascensión
    festivos.add(_next_monday(easter + timedelta(days=60)))   # Corpus Christi
    festivos.add(_next_monday(easter + timedelta(days=68)))   # Sagrado Corazón
    return festivos


def _add_dias_habiles(inicio: date, dias: int) -> date:
    festivos = _festivos_colombia(inicio.year) | _festivos_colombia(inicio.year + 1)
    current = inicio
    count = 0
    while count < dias:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in festivos:
            count += 1
    return current


def _subtract_dias_habiles(fin: date, dias: int) -> date:
    """Resta `dias` días hábiles hacia atrás desde `fin`."""
    festivos = _festivos_colombia(fin.year) | _festivos_colombia(fin.year - 1)
    current = fin
    count = 0
    while count < dias:
        current -= timedelta(days=1)
        if current.weekday() < 5 and current not in festivos:
            count += 1
    return current


def _calcular_semaforo_row(r: dict) -> dict:
    _ANEXO_VALS = {"ANEXO EXPEDIENTE", "ANEXO AL EXPEDIENTE"}
    r["dias_restantes"] = None
    r["fecha_vencimiento"] = None        # fecha legal real: fecha_ingreso + termino días hábiles
    r["fecha_termino_respuesta"] = None  # fecha de revisión: 2 días hábiles antes del vencimiento
    r["pendiente"] = not bool(r.get("fecha_radicado_salida"))

    if r.get("tipo_respuesta") and r["tipo_respuesta"].strip().upper() in _ANEXO_VALS:
        r["semaforo"] = "verde"
        r["dias_transcurridos"] = None
        return r

    if r.get("fecha_radicado_salida"):
        r["semaforo"] = "respondido"
        if r.get("fecha_ingreso"):
            try:
                fi = date.fromisoformat(r["fecha_ingreso"][:10])
                fs = date.fromisoformat(r["fecha_radicado_salida"][:10])
                r["dias_transcurridos"] = (fs - fi).days
            except Exception:
                pass
        return r

    if not r.get("fecha_ingreso"):
        r["semaforo"] = None
        r["dias_transcurridos"] = None
        return r

    if r.get("termino_dias"):
        try:
            fi = date.fromisoformat(r["fecha_ingreso"][:10])
            termino = int(r["termino_dias"])
            # Fecha legal de vencimiento: fecha_ingreso + N días hábiles
            fecha_venc = _add_dias_habiles(fi, termino)
            # Fecha de revisión: 2 días hábiles antes del vencimiento (nunca cae en finde/festivo)
            fecha_rev = _subtract_dias_habiles(fecha_venc, 2)
            r["fecha_vencimiento"] = fecha_venc.isoformat()
            r["fecha_termino_respuesta"] = fecha_rev.isoformat()
            dias_restantes = (fecha_rev - date.today()).days
            r["dias_restantes"] = dias_restantes
            r["dias_transcurridos"] = None
            if dias_restantes >= 2:
                r["semaforo"] = "verde"
            elif dias_restantes >= 0:
                r["semaforo"] = "amarilla"
            else:
                r["semaforo"] = "roja"
        except Exception:
            r["semaforo"] = None
        return r

    # Sin término definido: usar días transcurridos desde ingreso
    try:
        fi = date.fromisoformat(r["fecha_ingreso"][:10])
        dias = (date.today() - fi).days
        r["dias_transcurridos"] = dias
        if dias <= 5:
            r["semaforo"] = "verde"
        elif dias <= 8:
            r["semaforo"] = "amarilla"
        else:
            r["semaforo"] = "roja"
    except Exception:
        r["semaforo"] = None

    return r


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    conn = get_db()

    total = conn.execute("SELECT COUNT(*) FROM correspondencia").fetchone()[0]

    stats = conn.execute("""
        SELECT
            SUM(CASE WHEN fecha_radicado_salida IS NOT NULL AND fecha_radicado_salida != '' THEN 1 ELSE 0 END) AS respondidos,
            SUM(CASE WHEN UPPER(TRIM(tipo_respuesta)) IN ('ANEXO EXPEDIENTE','ANEXO AL EXPEDIENTE') THEN 1
                     WHEN (fecha_radicado_salida IS NULL OR fecha_radicado_salida = '') AND fecha_ingreso IS NOT NULL
                     AND (tipo_respuesta IS NULL OR UPPER(TRIM(tipo_respuesta)) NOT IN ('ANEXO EXPEDIENTE','ANEXO AL EXPEDIENTE'))
                     AND CAST(julianday('now','localtime') - julianday(substr(fecha_ingreso,1,10)) AS INTEGER) <= 5 THEN 1 ELSE 0 END) AS verde,
            SUM(CASE WHEN (tipo_respuesta IS NULL OR UPPER(TRIM(tipo_respuesta)) NOT IN ('ANEXO EXPEDIENTE','ANEXO AL EXPEDIENTE'))
                     AND (fecha_radicado_salida IS NULL OR fecha_radicado_salida = '') AND fecha_ingreso IS NOT NULL
                     AND CAST(julianday('now','localtime') - julianday(substr(fecha_ingreso,1,10)) AS INTEGER) BETWEEN 6 AND 8 THEN 1 ELSE 0 END) AS amarilla,
            SUM(CASE WHEN (tipo_respuesta IS NULL OR UPPER(TRIM(tipo_respuesta)) NOT IN ('ANEXO EXPEDIENTE','ANEXO AL EXPEDIENTE'))
                     AND (fecha_radicado_salida IS NULL OR fecha_radicado_salida = '') AND fecha_ingreso IS NOT NULL
                     AND CAST(julianday('now','localtime') - julianday(substr(fecha_ingreso,1,10)) AS INTEGER) >= 9 THEN 1 ELSE 0 END) AS roja
        FROM correspondencia
    """).fetchone()

    por_responsable = conn.execute("""
        SELECT responsable, COUNT(*) cant,
               SUM(CASE WHEN fecha_radicado_salida IS NOT NULL AND fecha_radicado_salida != '' THEN 1 ELSE 0 END) respondidos
        FROM correspondencia WHERE responsable IS NOT NULL
        GROUP BY responsable ORDER BY cant DESC LIMIT 15
    """).fetchall()

    por_mes = conn.execute("""
        SELECT mes, COUNT(*) cant FROM correspondencia
        WHERE mes IS NOT NULL GROUP BY mes
        ORDER BY CASE mes
            WHEN 'ENERO' THEN 1 WHEN 'FEBRERO' THEN 2 WHEN 'MARZO' THEN 3
            WHEN 'ABRIL' THEN 4 WHEN 'MAYO' THEN 5 WHEN 'JUNIO' THEN 6
            WHEN 'JULIO' THEN 7 WHEN 'AGOSTO' THEN 8 WHEN 'SEPTIEMBRE' THEN 9
            WHEN 'OCTUBRE' THEN 10 WHEN 'NOVIEMBRE' THEN 11 WHEN 'DICIEMBRE' THEN 12
            ELSE 99 END
    """).fetchall()

    criticos = conn.execute("""
        SELECT c.id, c.n_radicado, c.responsable, c.asunto, c.mes, c.fecha_ingreso,
               CAST(julianday('now','localtime') - julianday(substr(c.fecha_ingreso,1,10)) AS INTEGER) AS dias_transcurridos
        FROM correspondencia c
        WHERE (c.fecha_radicado_salida IS NULL OR c.fecha_radicado_salida = '')
        AND (c.tipo_respuesta IS NULL OR UPPER(TRIM(c.tipo_respuesta)) NOT IN ('ANEXO EXPEDIENTE','ANEXO AL EXPEDIENTE'))
        AND c.fecha_ingreso IS NOT NULL
        ORDER BY dias_transcurridos DESC LIMIT 20
    """).fetchall()

    conn.close()

    return templates.TemplateResponse("corr_dashboard.html", {
        "request": request,
        "active": "corr_dashboard",
        "total": total,
        "stats": dict(stats) if stats else {},
        "por_responsable": [dict(r) for r in por_responsable],
        "por_mes": [dict(r) for r in por_mes],
        "criticos": [dict(r) for r in criticos],
    })


# ── LISTA ──────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def lista(
    request: Request,
    q: str = "",
    semaforo: str = "",
    responsable: str = "",
    mes: str = "",
    anio: str = "",
    page: int = 1,
    por_pagina: int = 25,
    msg: str = "",
):
    conn = get_db()
    responsables, tipos_doc = _get_catalogos(conn)

    filtros = ["1=1"]
    params: list = []

    if q.strip():
        filtros.append("(c.n_radicado LIKE ? OR c.origen LIKE ? OR c.asunto LIKE ? OR c.caso_bmp LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"]
    if responsable.strip():
        filtros.append("c.responsable = ?")
        params.append(responsable.strip())
    if mes.strip():
        filtros.append("c.mes = ?")
        params.append(mes.strip())
    if anio.strip():
        filtros.append("c.anio = ?")
        params.append(int(anio.strip()))

    where = " AND ".join(filtros)

    # Fetch all qualifying rows — semaphore filter applied in Python
    rows_raw = conn.execute(f"""
        SELECT c.*,
               GROUP_CONCAT(rs.radicado, ' | ') AS radicados_salida,
               GROUP_CONCAT(COALESCE(rs.url, ''), ' | ') AS radicados_urls
        FROM correspondencia c
        LEFT JOIN correspondencia_radicados_salida rs ON rs.correspondencia_id = c.id
        WHERE {where}
        GROUP BY c.id
        ORDER BY c.fecha_ingreso DESC
    """, params).fetchall()

    anios_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT anio FROM correspondencia WHERE anio IS NOT NULL ORDER BY anio DESC"
    ).fetchall()]
    conn.close()

    # Compute semaphore and extract first radicado URL for list display
    all_rows = []
    for r in rows_raw:
        d = _calcular_semaforo_row(dict(r))
        urls_str = d.get("radicados_urls") or ""
        urls = [u.strip() for u in urls_str.split(" | ")] if urls_str.strip() else []
        d["primer_url_salida"] = next((u for u in urls if u), None)
        all_rows.append(d)

    # Apply semaphore filter in Python
    if semaforo == "pendiente":
        all_rows = [r for r in all_rows if r.get("pendiente")]
    elif semaforo:
        all_rows = [r for r in all_rows if r.get("semaforo") == semaforo]

    total = len(all_rows)
    total_pages = max(1, (total + por_pagina - 1) // por_pagina)
    offset = (page - 1) * por_pagina
    rows = all_rows[offset:offset + por_pagina]

    return templates.TemplateResponse("corr_lista.html", tpl(request, _MOD,
        active="corr_lista",
        rows=rows, total=total, page=page, total_pages=total_pages,
        por_pagina=por_pagina, q=q, semaforo=semaforo, responsable=responsable,
        mes=mes, anio=anio, responsables=responsables, meses=MESES,
        anios=anios_bd, msg=msg,
        back_url=request.url.path + ("?" + str(request.url.query) if request.url.query else ""),
    ))


# ── NUEVO ──────────────────────────────────────────────────────────────────────

@router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_form(request: Request):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/correspondencia/?msg=sin_permiso", status_code=303)
    conn = get_db()
    responsables, tipos_doc = _get_catalogos(conn)
    conn.close()
    return templates.TemplateResponse("corr_form.html", tpl(request, _MOD,
        active="corr_nuevo", modo="nuevo", reg={}, radicados_salida=[],
        responsables=responsables, tipos_doc=tipos_doc,
        tipos_respuesta=TIPOS_RESPUESTA, terminos_dias=TERMINOS_DIAS,
        tipos_requerimiento=TIPOS_REQUERIMIENTO, meses=MESES,
        anios=_anios_disponibles(),
    ))


@router.post("/nuevo")
async def nuevo_post(
    request: Request,
    anio: int = Form(None),
    mes: str = Form(""),
    fecha_ingreso: str = Form(""),
    n_radicado: str = Form(""),
    origen: str = Form(""),
    asunto: str = Form(""),
    tipo_documento: str = Form(""),
    responsable: str = Form(""),
    caso_bmp: str = Form(""),
    fecha_radicado_salida: str = Form(""),
    tipo_respuesta: str = Form(""),
    tramite_salida: str = Form(""),
    correo_remitente: str = Form(""),
    sinproc_personeria: str = Form(""),
    tipo_requerimiento: str = Form(""),
    termino_dias: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/correspondencia/?msg=sin_permiso", status_code=303)
    conn = get_db()
    termino_val = int(termino_dias) if termino_dias.strip() else None
    cur = conn.execute("""
        INSERT INTO correspondencia
        (anio, mes, fecha_ingreso, n_radicado, origen, asunto, tipo_documento,
         responsable, caso_bmp, fecha_radicado_salida, tipo_respuesta, tramite_salida,
         correo_remitente, sinproc_personeria, tipo_requerimiento, termino_dias)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [
        anio, _v(mes), _v(fecha_ingreso), _v(n_radicado), _v(origen), _v(asunto),
        _v(tipo_documento), _v(responsable), _v(caso_bmp),
        _v(fecha_radicado_salida), _v(tipo_respuesta), _v(tramite_salida),
        _v(correo_remitente), _v(sinproc_personeria), _v(tipo_requerimiento), termino_val,
    ])
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    registrar_log(user, "crear", _MOD, f"Oficio #{new_id} — {_v(n_radicado)}",
                  request.client.host if request.client else None)
    return RedirectResponse(f"/correspondencia/{new_id}/editar?msg=creado", status_code=303)


# ── EXPORTAR EXCEL ─────────────────────────────────────────────────────────────

@router.get("/exportar")
async def exportar():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/correspondencia/?msg=error_openpyxl")

    conn = get_db()
    rows = conn.execute("""
        SELECT c.*,
               GROUP_CONCAT(rs.radicado, ' | ') AS radicados_salida,
               GROUP_CONCAT(COALESCE(rs.url, ''), ' | ') AS radicados_urls
        FROM correspondencia c
        LEFT JOIN correspondencia_radicados_salida rs ON rs.correspondencia_id = c.id
        GROUP BY c.id
        ORDER BY c.fecha_ingreso DESC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CORRESPONDENCIA"

    h_fill = PatternFill("solid", fgColor="1B4F8A")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="EBF1F8")
    link_font = Font(color="0563C1", underline="single", size=10)

    headers = [
        "AÑO", "MES", "FECHA INGRESO DE OFICIO", "N. RADICADOS",
        "ENTIDAD", "CORREO REMITENTE", "ASUNTO", "NUMERO SINPROC PERSONERIA",
        "TIPO DE REQUERIMIENTO", "TERMINO (DIAS)", "TIPO DE DOCUMENTO",
        "RESPONSABLE", "CASO BMP", "N RADICADO SALIDA",
        "FECHA RADICADO DE SALIDA", "TIPO DE RESPUESTA", "OBSERVACIONES",
        "FECHA TERMINO DE RESPUESTA PETICION", "DÍAS TRANSCURRIDOS",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = center
    ws.row_dimensions[1].height = 36

    for ri, row in enumerate(rows, 2):
        d = _calcular_semaforo_row(dict(row))
        fill = alt_fill if ri % 2 == 0 else None

        urls_str = d.get("radicados_urls") or ""
        urls_list = [u.strip() for u in urls_str.split(" | ")] if urls_str.strip() else []
        first_url = next((u for u in urls_list if u), None)

        vals = [
            d.get("anio"),
            d.get("mes"),
            d.get("fecha_ingreso")[:10] if d.get("fecha_ingreso") else None,
            d.get("n_radicado"),
            d.get("origen"),
            d.get("correo_remitente"),
            d.get("asunto"),
            d.get("sinproc_personeria"),
            d.get("tipo_requerimiento"),
            d.get("termino_dias"),
            d.get("tipo_documento"),
            d.get("responsable"),
            d.get("caso_bmp"),
            d.get("radicados_salida"),           # col 14 — N RADICADO SALIDA
            d.get("fecha_radicado_salida")[:10] if d.get("fecha_radicado_salida") else None,
            d.get("tipo_respuesta"),
            d.get("tramite_salida"),
            d.get("fecha_termino_respuesta"),     # col 18 — FECHA TERMINO (calculated)
            d.get("dias_transcurridos"),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="center", wrap_text=ci in (5, 7))
            if fill:
                cell.fill = fill

        # Add hyperlink on N RADICADO SALIDA (column 14) when URL exists
        if first_url and d.get("radicados_salida"):
            rad_cell = ws.cell(row=ri, column=14)
            rad_cell.hyperlink = first_url
            rad_cell.font = link_font

    col_widths = [6, 12, 20, 18, 30, 30, 40, 20, 40, 10, 18, 28, 10, 22, 20, 25, 30, 28, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = date.today().strftime("%Y%m%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Correspondencia_{hoy}.xlsx"},
    )


# ── IMPORTAR ───────────────────────────────────────────────────────────────────

@router.get("/importar", response_class=HTMLResponse)
async def importar_form(request: Request, msg: str = ""):
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) FROM correspondencia").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("corr_importar.html", {
        "request": request,
        "active": "corr_importar",
        "msg": msg,
        "total_actual": total,
    })


@router.post("/importar")
async def importar_post(request: Request, archivo: UploadFile = File(...)):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/correspondencia/importar?msg=sin_permiso", status_code=303)
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/correspondencia/importar?msg=error_openpyxl", status_code=303)

    contenido = await archivo.read()
    if not contenido:
        return RedirectResponse("/correspondencia/importar?msg=error_vacio", status_code=303)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        return RedirectResponse("/correspondencia/importar?msg=error_archivo", status_code=303)

    MESES_HOJAS = ("ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                   "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE")

    hojas_originales = [
        sn for sn in wb.sheetnames
        if any(sn.upper().startswith(m) for m in MESES_HOJAS)
    ]
    es_formato_original = len(hojas_originales) > 0

    def _norm_fecha(f):
        if not f:
            return None
        s = str(f).strip()
        if len(s) >= 10:
            return s[:10]
        return s

    conn = get_db()
    insertados = 0
    try:
        conn.execute("DELETE FROM correspondencia_radicados_salida")
        conn.execute("DELETE FROM correspondencia")

        if es_formato_original:
            # Formato original: iterar todas las hojas de meses
            for nombre_hoja in hojas_originales:
                ws = wb[nombre_hoja]
                for row in ws.iter_rows(min_row=6, values_only=True):
                    if not any(v for v in row if v is not None):
                        continue
                    if _v(row[0]) and str(row[0]).strip().upper() in ("MES", "MES "):
                        continue

                    mes_val    = _v(row[0])
                    fi_val     = _norm_fecha(_v(row[1]))
                    rad_val    = _v(row[2])
                    orig_val   = _v(row[3])
                    asunto_val = _v(row[4])
                    tipo_d_val = _v(row[5])
                    resp_val   = _v(row[6])
                    bmp_val    = _v(row[7])
                    rad_sal    = _v(row[10]) if len(row) > 10 else None
                    fsal_val   = _norm_fecha(_v(row[11]) if len(row) > 11 else None)
                    tipor_val  = _v(row[12]) if len(row) > 12 else None
                    tram_val   = _v(row[13]) if len(row) > 13 else None

                    anio_val = None
                    if fi_val:
                        try:
                            anio_val = int(fi_val[:4])
                        except Exception:
                            pass

                    if mes_val:
                        mes_val = mes_val.strip().upper()
                    resp_val = _clean_responsable(resp_val)

                    if not fi_val and not rad_val:
                        continue

                    cur = conn.execute("""
                        INSERT INTO correspondencia
                        (anio, mes, fecha_ingreso, n_radicado, origen, asunto, tipo_documento,
                         responsable, caso_bmp, fecha_radicado_salida, tipo_respuesta, tramite_salida)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """, [anio_val, mes_val, fi_val, rad_val, orig_val, asunto_val,
                          tipo_d_val, resp_val, bmp_val, fsal_val, tipor_val, tram_val])
                    cid = cur.lastrowid
                    insertados += 1

                    if rad_sal:
                        for r in str(rad_sal).split("|"):
                            r = r.strip()
                            if r:
                                conn.execute(
                                    "INSERT INTO correspondencia_radicados_salida (correspondencia_id, radicado) VALUES (?,?)",
                                    (cid, r),
                                )

        else:
            # Formato exportado: hoja "CORRESPONDENCIA" con headers en fila 1
            ws = None
            for nombre in ["CORRESPONDENCIA", "CORR", "2026"]:
                if nombre in wb.sheetnames:
                    ws = wb[nombre]
                    break
            if ws is None:
                ws = wb.active

            # Build header→column-index map (handles any column order, old or new format)
            header_row_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]
            header_row = [str(c.value or "").strip().upper() for c in header_row_cells]
            hmap = {h: i for i, h in enumerate(header_row)}

            def _hi(name, *alts):
                idx = hmap.get(name)
                if idx is None:
                    for alt in alts:
                        idx = hmap.get(alt)
                        if idx is not None:
                            break
                return idx if idx is not None else -1

            i_correo   = _hi("CORREO REMITENTE")
            i_asunto   = _hi("ASUNTO", "ASUNTO AGILSALUD")
            i_sinproc  = _hi("NUMERO SINPROC PERSONERIA", "SINPROC PERSONERIA")
            i_tiporeq  = _hi("TIPO DE REQUERIMIENTO")
            i_termino  = _hi("TERMINO (DIAS)")
            i_tipodoc  = _hi("TIPO DE DOCUMENTO")
            i_resp     = _hi("RESPONSABLE")
            i_bmp      = _hi("CASO BMP")
            i_radsal   = _hi("N RADICADO SALIDA")
            i_urlsal   = _hi("URL RADICADO SALIDA")   # legacy v1 column (may be absent)
            i_fechasal = _hi("FECHA RADICADO DE SALIDA")
            i_tipor    = _hi("TIPO DE RESPUESTA")
            i_observ   = _hi("OBSERVACIONES")

            def _cv(cells, idx):
                if idx < 0 or idx >= len(cells):
                    return None
                return _v(cells[idx].value)

            for row_cells in ws.iter_rows(min_row=2):
                if not any(c.value for c in row_cells):
                    continue
                try:
                    anio_val = int(row_cells[0].value)
                except (ValueError, TypeError):
                    continue

                mes_val     = _cv(row_cells, 1)
                fi_val      = _norm_fecha(_cv(row_cells, 2))
                rad_val     = _cv(row_cells, 3)
                orig_val    = _cv(row_cells, 4)
                correo_val  = _cv(row_cells, i_correo)
                asunto_val  = _cv(row_cells, i_asunto)
                sinproc_val = _cv(row_cells, i_sinproc)
                tiporeq_val = _cv(row_cells, i_tiporeq)
                termino_str = _cv(row_cells, i_termino)
                tipo_d_val  = _cv(row_cells, i_tipodoc)
                resp_val    = _cv(row_cells, i_resp)
                bmp_val     = _cv(row_cells, i_bmp)
                rad_sal     = _cv(row_cells, i_radsal)
                fsal_val    = _norm_fecha(_cv(row_cells, i_fechasal))
                tipor_val   = _cv(row_cells, i_tipor)
                tram_val    = _cv(row_cells, i_observ)

                # Read hyperlink URL from N RADICADO SALIDA cell; fallback to legacy URL column
                url_sal = None
                if i_radsal >= 0 and i_radsal < len(row_cells):
                    hl = row_cells[i_radsal].hyperlink
                    if hl:
                        url_sal = hl.target
                if not url_sal and i_urlsal >= 0:
                    url_sal = _cv(row_cells, i_urlsal)

                termino_val = None
                if termino_str:
                    try:
                        termino_val = int(float(termino_str))
                    except Exception:
                        pass

                if mes_val:
                    mes_val = mes_val.strip().upper()
                resp_val = _clean_responsable(resp_val)

                cur = conn.execute("""
                    INSERT INTO correspondencia
                    (anio, mes, fecha_ingreso, n_radicado, origen, asunto, tipo_documento,
                     responsable, caso_bmp, fecha_radicado_salida, tipo_respuesta, tramite_salida,
                     correo_remitente, sinproc_personeria, tipo_requerimiento, termino_dias)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, [anio_val, mes_val, fi_val, rad_val, orig_val, asunto_val,
                      tipo_d_val, resp_val, bmp_val, fsal_val, tipor_val, tram_val,
                      correo_val, sinproc_val, tiporeq_val, termino_val])
                cid = cur.lastrowid
                insertados += 1

                if rad_sal:
                    radicados_list = [r.strip() for r in str(rad_sal).split("|") if r.strip()]
                    for idx_r, r in enumerate(radicados_list):
                        u = url_sal if idx_r == 0 else None
                        conn.execute(
                            "INSERT INTO correspondencia_radicados_salida (correspondencia_id, radicado, url) VALUES (?,?,?)",
                            (cid, r, u),
                        )

        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return RedirectResponse(f"/correspondencia/importar?msg=error_import", status_code=303)

    conn.close()
    registrar_log(user, "importar", _MOD, f"{insertados} registros importados",
                  request.client.host if request.client else None)
    return RedirectResponse(f"/correspondencia/importar?msg=ok_{insertados}", status_code=303)


# ── CONFIGURAR CATÁLOGOS ───────────────────────────────────────────────────────

@router.get("/configurar", response_class=HTMLResponse)
async def configurar(request: Request, msg: str = ""):
    conn = get_db()
    responsables = conn.execute(
        "SELECT id, nombre FROM corr_responsables ORDER BY nombre"
    ).fetchall()
    tipos_doc = conn.execute(
        "SELECT id, nombre FROM corr_tipos_documento ORDER BY nombre"
    ).fetchall()
    conn.close()
    return templates.TemplateResponse("corr_configurar.html", {
        "request": request,
        "active": "corr_configurar",
        "responsables": [dict(r) for r in responsables],
        "tipos_doc": [dict(r) for r in tipos_doc],
        "tipos_respuesta": TIPOS_RESPUESTA,
        "msg": msg,
    })


@router.post("/configurar/responsable/nuevo")
async def responsable_nuevo(nombre: str = Form(...)):
    nombre = nombre.strip().upper()
    if not nombre:
        return RedirectResponse("/correspondencia/configurar?msg=vacio", status_code=303)
    conn = get_db()
    try:
        conn.execute("INSERT INTO corr_responsables (nombre) VALUES (?)", (nombre,))
        conn.commit()
    except Exception:
        conn.close()
        return RedirectResponse("/correspondencia/configurar?msg=duplicado", status_code=303)
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


@router.post("/configurar/responsable/{rid}/editar")
async def responsable_editar(rid: int, nombre: str = Form(...)):
    nombre = nombre.strip().upper()
    if not nombre:
        return RedirectResponse("/correspondencia/configurar?msg=vacio", status_code=303)
    conn = get_db()
    conn.execute("UPDATE corr_responsables SET nombre=? WHERE id=?", (nombre, rid))
    conn.commit()
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


@router.post("/configurar/responsable/{rid}/eliminar")
async def responsable_eliminar(rid: int):
    conn = get_db()
    conn.execute("DELETE FROM corr_responsables WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


@router.post("/configurar/tipo_doc/nuevo")
async def tipo_doc_nuevo(nombre: str = Form(...)):
    nombre = nombre.strip().upper()
    if not nombre:
        return RedirectResponse("/correspondencia/configurar?msg=vacio", status_code=303)
    conn = get_db()
    try:
        conn.execute("INSERT INTO corr_tipos_documento (nombre) VALUES (?)", (nombre,))
        conn.commit()
    except Exception:
        conn.close()
        return RedirectResponse("/correspondencia/configurar?msg=duplicado", status_code=303)
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


@router.post("/configurar/tipo_doc/{tid}/editar")
async def tipo_doc_editar(tid: int, nombre: str = Form(...)):
    nombre = nombre.strip().upper()
    conn = get_db()
    conn.execute("UPDATE corr_tipos_documento SET nombre=? WHERE id=?", (nombre, tid))
    conn.commit()
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


@router.post("/configurar/tipo_doc/{tid}/eliminar")
async def tipo_doc_eliminar(tid: int):
    conn = get_db()
    conn.execute("DELETE FROM corr_tipos_documento WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return RedirectResponse("/correspondencia/configurar?msg=ok", status_code=303)


# ── IMPORTAR DESDE AGIL SALUD (Documentos.xlsx) ───────────────────────────────

_AGILSALUD_DESTINATARIOS = {
    "MARTHA PATRICIA AÑEZ MAESTRE",
    "MABEL GICELA HURTADO SANCHEZ",
}

_MES_NOMBRES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}


@router.get("/importar-agilsalud", response_class=HTMLResponse)
async def importar_agilsalud_form(request: Request, msg: str = ""):
    return templates.TemplateResponse("corr_importar_agilsalud.html", {
        "request": request,
        "active": "corr_importar_agilsalud",
        "msg": msg,
        "preview": None,
    })


@router.post("/importar-agilsalud/preview", response_class=HTMLResponse)
async def importar_agilsalud_preview(request: Request, archivo: UploadFile = File(...)):
    try:
        import openpyxl
    except ImportError:
        return templates.TemplateResponse("corr_importar_agilsalud.html", {
            "request": request, "active": "corr_importar_agilsalud",
            "msg": "error_openpyxl", "preview": None,
        })

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception:
        return templates.TemplateResponse("corr_importar_agilsalud.html", {
            "request": request, "active": "corr_importar_agilsalud",
            "msg": "error_archivo", "preview": None,
        })

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        destinatario = str(row[4] or "").strip().upper() if len(row) > 4 else ""
        if destinatario not in _AGILSALUD_DESTINATARIOS:
            continue

        n_radicado = str(row[0] or "").strip() if len(row) > 0 else ""
        origen = str(row[6] or "").strip() if len(row) > 6 else ""
        correo_remitente = str(row[8] or "").strip() if len(row) > 8 else ""
        fecha_raw = row[10] if len(row) > 10 else None
        asunto = str(row[12] or "").strip() if len(row) > 12 else ""

        fecha_ingreso = ""
        mes = ""
        anio = ""
        if fecha_raw:
            try:
                if hasattr(fecha_raw, "strftime"):
                    fecha_ingreso = fecha_raw.strftime("%Y-%m-%d")
                    mes = _MES_NOMBRES.get(fecha_raw.month, "")
                    anio = fecha_raw.year
                else:
                    s = str(fecha_raw).strip()
                    fecha_ingreso = s[:10]
                    from datetime import datetime as _dt
                    dt = _dt.fromisoformat(fecha_ingreso)
                    mes = _MES_NOMBRES.get(dt.month, "")
                    anio = dt.year
            except Exception:
                pass

        resp_map = {
            "MARTHA PATRICIA AÑEZ MAESTRE": "MARTHA PATRICIA AÑEZ MAESTRE",
            "MABEL GICELA HURTADO SANCHEZ": "MABEL GICELLA HURTADO",
        }
        responsable = resp_map.get(destinatario, destinatario.title())

        filas.append({
            "n_radicado": n_radicado,
            "responsable": responsable,
            "origen": origen,
            "correo_remitente": correo_remitente,
            "fecha_ingreso": fecha_ingreso,
            "mes": mes,
            "anio": anio,
            "asunto": asunto,
        })

    if not filas:
        return templates.TemplateResponse("corr_importar_agilsalud.html", {
            "request": request, "active": "corr_importar_agilsalud",
            "msg": "error_vacio", "preview": None,
        })

    import json
    preview_json = json.dumps(filas, ensure_ascii=False)
    return templates.TemplateResponse("corr_importar_agilsalud.html", {
        "request": request,
        "active": "corr_importar_agilsalud",
        "msg": "",
        "preview": filas,
        "preview_json": preview_json,
    })


@router.post("/importar-agilsalud/confirmar")
async def importar_agilsalud_confirmar(request: Request, datos_json: str = Form(...)):
    import json
    try:
        filas = json.loads(datos_json)
    except Exception:
        return RedirectResponse("/correspondencia/importar-agilsalud?msg=error_import", status_code=303)

    conn = get_db()
    try:
        insertados = 0
        for f in filas:
            conn.execute(
                """INSERT INTO correspondencia
                   (anio, mes, fecha_ingreso, n_radicado, origen, asunto,
                    responsable, correo_remitente)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (f.get("anio") or None, f.get("mes") or None,
                 f.get("fecha_ingreso") or None, f.get("n_radicado") or None,
                 f.get("origen") or None, f.get("asunto") or None,
                 f.get("responsable") or None, f.get("correo_remitente") or None),
            )
            insertados += 1
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return RedirectResponse("/correspondencia/importar-agilsalud?msg=error_import", status_code=303)
    conn.close()
    return RedirectResponse(f"/correspondencia/importar-agilsalud?msg=ok_{insertados}", status_code=303)


# ── VER / EDITAR / ELIMINAR ────────────────────────────────────────────────────

@router.get("/{reg_id}", response_class=HTMLResponse)
async def ver(request: Request, reg_id: int, back: str = ""):
    conn = get_db()
    row = conn.execute("SELECT * FROM correspondencia WHERE id=?", (reg_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/correspondencia/?msg=no_encontrado")
    radicados = conn.execute(
        "SELECT * FROM correspondencia_radicados_salida WHERE correspondencia_id=? ORDER BY id",
        (reg_id,),
    ).fetchall()
    conn.close()
    reg = _calcular_semaforo_row(dict(row))
    return templates.TemplateResponse("corr_detalle.html", tpl(request, _MOD,
        active="corr_lista", reg=reg,
        radicados=[dict(r) for r in radicados],
        back=back or "/correspondencia/",
    ))


@router.get("/{reg_id}/editar", response_class=HTMLResponse)
async def editar_form(request: Request, reg_id: int, msg: str = "", back: str = ""):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/correspondencia/{reg_id}?msg=sin_permiso", status_code=303)
    conn = get_db()
    row = conn.execute("SELECT * FROM correspondencia WHERE id=?", (reg_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/correspondencia/?msg=no_encontrado")
    radicados = conn.execute(
        "SELECT * FROM correspondencia_radicados_salida WHERE correspondencia_id=? ORDER BY id",
        (reg_id,),
    ).fetchall()
    responsables, tipos_doc = _get_catalogos(conn)
    conn.close()
    return templates.TemplateResponse("corr_form.html", tpl(request, _MOD,
        active="corr_lista", modo="editar", reg=dict(row),
        radicados_salida=[dict(r) for r in radicados],
        responsables=responsables, tipos_doc=tipos_doc,
        tipos_respuesta=TIPOS_RESPUESTA, terminos_dias=TERMINOS_DIAS,
        tipos_requerimiento=TIPOS_REQUERIMIENTO, meses=MESES,
        anios=_anios_disponibles(), msg=msg,
        back=back or "/correspondencia/",
    ))


@router.post("/{reg_id}/editar")
async def editar_post(
    request: Request,
    reg_id: int,
    anio: int = Form(None),
    mes: str = Form(""),
    fecha_ingreso: str = Form(""),
    n_radicado: str = Form(""),
    origen: str = Form(""),
    asunto: str = Form(""),
    tipo_documento: str = Form(""),
    responsable: str = Form(""),
    caso_bmp: str = Form(""),
    fecha_radicado_salida: str = Form(""),
    tipo_respuesta: str = Form(""),
    tramite_salida: str = Form(""),
    correo_remitente: str = Form(""),
    sinproc_personeria: str = Form(""),
    tipo_requerimiento: str = Form(""),
    termino_dias: str = Form(""),
):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse(f"/correspondencia/{reg_id}?msg=sin_permiso", status_code=303)
    termino_val = int(termino_dias) if termino_dias.strip() else None
    conn = get_db()
    conn.execute("""
        UPDATE correspondencia SET
        anio=?, mes=?, fecha_ingreso=?, n_radicado=?, origen=?, asunto=?,
        tipo_documento=?, responsable=?, caso_bmp=?, fecha_radicado_salida=?,
        tipo_respuesta=?, tramite_salida=?, correo_remitente=?,
        sinproc_personeria=?, tipo_requerimiento=?, termino_dias=?,
        updated_at=datetime('now','localtime')
        WHERE id=?
    """, [
        anio, _v(mes), _v(fecha_ingreso), _v(n_radicado), _v(origen), _v(asunto),
        _v(tipo_documento), _v(responsable), _v(caso_bmp),
        _v(fecha_radicado_salida), _v(tipo_respuesta), _v(tramite_salida),
        _v(correo_remitente), _v(sinproc_personeria), _v(tipo_requerimiento),
        termino_val, reg_id,
    ])
    conn.commit()
    conn.close()
    registrar_log(user, "editar", _MOD, f"Oficio #{reg_id} — {_v(n_radicado)}",
                  request.client.host if request.client else None)
    return RedirectResponse(f"/correspondencia/{reg_id}/editar?msg=actualizado", status_code=303)


@router.post("/{reg_id}/eliminar")
async def eliminar(request: Request, reg_id: int):
    user = getattr(request.state, "user", None)
    if not _pw(user, _MOD):
        return RedirectResponse("/correspondencia/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("DELETE FROM correspondencia WHERE id=?", (reg_id,))
    conn.commit()
    conn.close()
    registrar_log(user, "eliminar", _MOD, f"Oficio #{reg_id}",
                  request.client.host if request.client else None)
    return RedirectResponse("/correspondencia/?msg=eliminado", status_code=303)


# ── RADICADOS DE SALIDA ────────────────────────────────────────────────────────

@router.post("/{reg_id}/radicado_salida/nuevo")
async def radicado_nuevo(reg_id: int, radicado: str = Form(...), url: str = Form("")):
    r = radicado.strip()
    u = url.strip() or None
    if r:
        conn = get_db()
        conn.execute(
            "INSERT INTO correspondencia_radicados_salida (correspondencia_id, radicado, url) VALUES (?,?,?)",
            (reg_id, r, u),
        )
        conn.commit()
        conn.close()
    return RedirectResponse(f"/correspondencia/{reg_id}/editar?msg=rad_ok", status_code=303)


@router.post("/radicado_salida/{rad_id}/eliminar")
async def radicado_eliminar(rad_id: int):
    conn = get_db()
    row = conn.execute(
        "SELECT correspondencia_id FROM correspondencia_radicados_salida WHERE id=?", (rad_id,)
    ).fetchone()
    conn.execute("DELETE FROM correspondencia_radicados_salida WHERE id=?", (rad_id,))
    conn.commit()
    reg_id = row["correspondencia_id"] if row else None
    conn.close()
    if reg_id:
        return RedirectResponse(f"/correspondencia/{reg_id}/editar?msg=rad_eliminado", status_code=303)
    return RedirectResponse("/correspondencia/", status_code=303)
