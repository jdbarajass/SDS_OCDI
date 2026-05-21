from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from pathlib import Path
from app.template_utils import make_templates
from datetime import date, datetime
import io

from app.database import get_db, row_to_dict, get_personal_oficina
from app.auth_utils import tpl, puede_escribir as _pw, puede_importar as _pi, registrar_log

_MOD = "sdqs"

router = APIRouter(prefix="/sdqs")
templates = make_templates(str(Path(__file__).parent.parent / "templates"))

MESES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

ESTADOS_PROCESO = [
    "INDAGACION PREVIA",
    "INVESTIGACION DISCIPLINARIA",
    "INHIBITORIO",
]

ABOGADOS = [
    "CARLOS ALFONSO PARRA MALAVER",
    "CESAR IVAN RODRIGUEZ DAMIAN",
    "DAVID FELIPE MORALES NOGUERA",
    "JANIK HERNANDO DE LA HOZ RIOS",
    "MABEL GICELLA HURTADO SANCHEZ",
    "MARA LUCIA UCROS MERLANO",
    "RODOLFO CARRILLO QUINTERO",
    "ANDRES EDUARDO SANDOVAL MAYORGA",
    "MAGDA XIMENA PAREDES LIEVANO",
    "LUNA GICELL GUZMAN YATE",
    "MARTHA PATRICIA AÑEZ MAESTRE",
]

RESP_MAP = {
    "CESAR RODRIGUEZ":       "CESAR IVAN RODRIGUEZ DAMIAN",
    "DAVID MORALES":         "DAVID FELIPE MORALES NOGUERA",
    "MARA OCRUS":            "MARA LUCIA UCROS MERLANO",
    "MARA UCROS":            "MARA LUCIA UCROS MERLANO",
    "MABEL GICELLA HURTADO": "MABEL GICELLA HURTADO SANCHEZ",
    "CARLOS PARRA":          "CARLOS ALFONSO PARRA MALAVER",
    "ANDRES SANDOVAL":       "ANDRES EDUARDO SANDOVAL MAYORGA",
}

_PAGE_SIZE = 25


def _str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return s if s.upper() != "NAN" else ""


def _calcular_semaforo_sdqs(reg: dict) -> dict:
    """
    Calcula estado_dias (total días del plazo) y semaforo_sdqs ('verde'/'amarillo'/'rojo').
    Verde  = primera mitad del plazo aún no cumplida.
    Amarillo = segunda mitad del plazo (pero > 2 días restantes).
    Rojo   = 2 días o menos hasta el vencimiento (o ya vencido).
    """
    fa = reg.get("fecha_asignacion")
    fv = reg.get("fecha_vencimiento")
    reg["estado_dias"] = None
    reg["semaforo_sdqs"] = None
    if not fa or not fv:
        return reg
    try:
        fa_d = datetime.fromisoformat(str(fa)[:10]).date()
        fv_d = datetime.fromisoformat(str(fv)[:10]).date()
        total_dias = (fv_d - fa_d).days
        if total_dias <= 0:
            return reg
        hoy = date.today()
        dias_transcurridos = (hoy - fa_d).days
        dias_restantes = (fv_d - hoy).days
        reg["estado_dias"] = total_dias
        if dias_restantes <= 2:
            reg["semaforo_sdqs"] = "rojo"
        elif dias_transcurridos >= total_dias // 2:
            reg["semaforo_sdqs"] = "amarillo"
        else:
            reg["semaforo_sdqs"] = "verde"
    except Exception:
        pass
    return reg


# ── Lista ─────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def lista(
    request: Request,
    mes: str = "",
    competencia_ocdi: str = "",
    responsable: str = "",
    q: str = "",
    page: int = 1,
    msg: str = "",
):
    conn = get_db()
    where, params = ["1=1"], []
    if mes:
        where.append("UPPER(mes) = ?")
        params.append(mes.upper())
    if competencia_ocdi:
        where.append("UPPER(competencia_ocdi) = ?")
        params.append(competencia_ocdi.upper())
    if responsable:
        where.append("UPPER(responsable) = ?")
        params.append(responsable.upper())
    if q:
        where.append("(UPPER(sdqs) LIKE ? OR UPPER(quejoso) LIKE ? OR UPPER(tema) LIKE ?)")
        like = f"%{q.upper()}%"
        params += [like, like, like]

    sql_base = f"FROM sdqs WHERE {' AND '.join(where)}"
    total = conn.execute(f"SELECT COUNT(*) {sql_base}", params).fetchone()[0]
    offset = (page - 1) * _PAGE_SIZE
    rows = conn.execute(
        f"SELECT * {sql_base} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [_PAGE_SIZE, offset],
    ).fetchall()

    meses_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT mes FROM sdqs WHERE mes IS NOT NULL ORDER BY mes"
    ).fetchall()]
    responsables_bd = [r[0] for r in conn.execute(
        "SELECT DISTINCT responsable FROM sdqs WHERE responsable IS NOT NULL AND responsable != '' ORDER BY responsable"
    ).fetchall()]
    conn.close()

    total_pages = max(1, (total + _PAGE_SIZE - 1) // _PAGE_SIZE)
    registros = [_calcular_semaforo_sdqs(row_to_dict(r)) for r in rows]

    return templates.TemplateResponse("sdqs_lista.html", tpl(request, _MOD,
        registros=registros,
        total=total,
        page=page,
        total_pages=total_pages,
        meses_bd=meses_bd,
        responsables_bd=responsables_bd,
        fil_mes=mes,
        fil_competencia=competencia_ocdi,
        fil_responsable=responsable,
        fil_q=q,
        msg=msg,
    ))


# ── Nuevo ─────────────────────────────────────────────────────────────────────

@router.get("/nuevo", response_class=HTMLResponse)
async def nuevo_get(request: Request):
    conn = get_db()
    abogados = get_personal_oficina(conn)
    conn.close()
    return templates.TemplateResponse("sdqs_form.html", tpl(request, _MOD,
        modo="nuevo",
        registro={},
        meses=MESES,
        abogados=abogados,
        estados=ESTADOS_PROCESO,
    ))


@router.post("/nuevo")
async def nuevo_post(
    request: Request,
    mes: str = Form(""),
    fecha_asignacion: str = Form(""),
    sdqs_num: str = Form("", alias="sdqs"),
    url_sdqs: str = Form(""),
    fecha_vencimiento: str = Form(""),
    quejoso: str = Form(""),
    correo: str = Form(""),
    tema: str = Form(""),
    competencia_ocdi: str = Form("NO"),
    bpm: str = Form(""),
    responsable: str = Form(""),
    rad_salida: str = Form(""),
    url_rad_salida: str = Form(""),
    fecha_respuesta: str = Form(""),
    observaciones: str = Form(""),
    estado_proceso: str = Form(""),
    hecho_corrupto: str = Form(""),
    valor_institucional: str = Form(""),
    tipologia: str = Form(""),
):
    user = request.state.user
    if not _pw(user, _MOD):
        return RedirectResponse("/sdqs/?msg=sin_permiso", status_code=303)

    obligatorios = [mes, fecha_asignacion, sdqs_num, quejoso, correo, tema, competencia_ocdi, observaciones]
    if any(not v.strip() for v in obligatorios):
        return RedirectResponse("/sdqs/nuevo?msg=error_obligatorios", status_code=303)

    conn = get_db()
    conn.execute(
        """INSERT INTO sdqs
           (mes, fecha_asignacion, sdqs, url_sdqs, fecha_vencimiento, quejoso, correo, tema,
            competencia_ocdi, bpm, responsable, rad_salida, url_rad_salida,
            fecha_respuesta, observaciones,
            estado_proceso, hecho_corrupto, valor_institucional, tipologia, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (mes.upper(), fecha_asignacion, sdqs_num.upper(),
         url_sdqs or None,
         fecha_vencimiento or None,
         quejoso.upper(), correo, tema.upper(), competencia_ocdi.upper(),
         bpm or None, responsable or None, rad_salida or None, url_rad_salida or None,
         fecha_respuesta or None,
         observaciones, estado_proceso or None, hecho_corrupto or None,
         valor_institucional or None, tipologia or None,
         user.get("nombre_completo") if user else None),
    )
    conn.commit()
    conn.close()
    registrar_log(user, "crear", _MOD, f"SDQS: {sdqs_num}")
    return RedirectResponse("/sdqs/?msg=creado", status_code=303)


# ── Exportar Excel ────────────────────────────────────────────────────────────

@router.get("/exportar")
async def exportar(
    request: Request,
    mes: str = "",
    competencia_ocdi: str = "",
    responsable: str = "",
    q: str = "",
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/sdqs/?msg=error_archivo", status_code=303)

    conn = get_db()
    where, params = ["1=1"], []
    if mes:
        where.append("UPPER(mes) = ?")
        params.append(mes.upper())
    if competencia_ocdi:
        where.append("UPPER(competencia_ocdi) = ?")
        params.append(competencia_ocdi.upper())
    if responsable:
        where.append("UPPER(responsable) = ?")
        params.append(responsable.upper())
    if q:
        where.append("(UPPER(sdqs) LIKE ? OR UPPER(quejoso) LIKE ? OR UPPER(tema) LIKE ?)")
        like = f"%{q.upper()}%"
        params += [like, like, like]

    rows = conn.execute(
        f"SELECT * FROM sdqs WHERE {' AND '.join(where)} ORDER BY fecha_asignacion, id",
        params,
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SDQS"

    headers = [
        "MES", "FECHA ASIGNACION", "SDQS", "FECHA VENCIMIENTO", "ESTADO DIAS",
        "QUEJOSO", "CORREO", "TEMA",
        "COMPETENCIA OCDI", "BPM", "RESPONSABLE", "RAD SALIDA",
        "FECHA RESPUESTA", "OBSERVACIONES", "ESTADO PROCESO",
        "HECHO CORRUPTO", "VALOR INSTITUCIONAL", "TIPOLOGIA", "URL SDQS",
    ]
    header_fill = PatternFill("solid", fgColor="1B4F8A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    SEM_COLORS = {"verde": "D4EDDA", "amarillo": "FFF3CD", "rojo": "F8D7DA"}
    link_font = Font(color="0563C1", underline="single")

    for r in rows:
        d = _calcular_semaforo_sdqs(row_to_dict(r))
        ws.append([
            d.get("mes"), d.get("fecha_asignacion"), d.get("sdqs"),
            d.get("fecha_vencimiento"), d.get("estado_dias"),
            d.get("quejoso"), d.get("correo"), d.get("tema"),
            d.get("competencia_ocdi"), d.get("bpm"), d.get("responsable"),
            d.get("rad_salida"), d.get("fecha_respuesta"), d.get("observaciones"),
            d.get("estado_proceso"), d.get("hecho_corrupto"),
            d.get("valor_institucional"), d.get("tipologia"), d.get("url_sdqs"),
        ])
        ri = ws.max_row
        sem = d.get("semaforo_sdqs")
        if sem and sem in SEM_COLORS:
            ws.cell(ri, 5).fill = PatternFill("solid", fgColor=SEM_COLORS[sem])
        # Hipervínculo en columna SDQS (col 3)
        url_sdqs = d.get("url_sdqs")
        if url_sdqs and d.get("sdqs"):
            sdqs_cell = ws.cell(ri, 3)
            sdqs_cell.hyperlink = url_sdqs
            sdqs_cell.font = link_font
        # Hipervínculo en columna RAD SALIDA (col 12)
        url = d.get("url_rad_salida")
        if url and d.get("rad_salida"):
            rad_cell = ws.cell(ri, 12)
            rad_cell.hyperlink = url
            rad_cell.font = link_font

    col_widths = [10, 16, 14, 16, 12, 28, 28, 50, 14, 14, 28, 16, 16, 40, 22, 22, 22, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SDQS_export.xlsx"},
    )


# ── Importar ──────────────────────────────────────────────────────────────────

@router.get("/importar", response_class=HTMLResponse)
async def importar_get(request: Request, msg: str = ""):
    user = request.state.user
    if not _pi(user, _MOD):
        return RedirectResponse("/sdqs/?msg=sin_permiso", status_code=303)
    conn = get_db()
    total_bd = conn.execute("SELECT COUNT(*) FROM sdqs").fetchone()[0]
    conn.close()
    return templates.TemplateResponse("sdqs_importar.html", tpl(request, _MOD,
        total_bd=total_bd,
        msg=msg,
    ))


@router.post("/importar")
async def importar_post(request: Request, archivo: UploadFile = File(...)):
    user = request.state.user
    if not _pi(user, _MOD):
        return RedirectResponse("/sdqs/?msg=sin_permiso", status_code=303)

    contenido = await archivo.read()
    count, errors = _importar_excel_sdqs(contenido)
    if errors:
        return RedirectResponse("/sdqs/importar?msg=error_archivo", status_code=303)
    registrar_log(user, "importar", _MOD, f"{count} registros importados")
    return RedirectResponse(f"/sdqs/importar?msg=importado_{count}", status_code=303)


@router.post("/limpiar")
async def limpiar(request: Request):
    user = request.state.user
    if not _pi(user, _MOD):
        return RedirectResponse("/sdqs/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("DELETE FROM sdqs")
    conn.commit()
    conn.close()
    registrar_log(user, "limpiar", _MOD, "Tabla SDQS borrada")
    return RedirectResponse("/sdqs/importar?msg=importado_0", status_code=303)


# ── Ver ───────────────────────────────────────────────────────────────────────

@router.get("/{id}", response_class=HTMLResponse)
async def ver(request: Request, id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM sdqs WHERE id = ?", (id,)).fetchone()
    abogados = get_personal_oficina(conn)
    conn.close()
    if not row:
        return RedirectResponse("/sdqs/?msg=no_encontrado", status_code=303)
    registro = _calcular_semaforo_sdqs(row_to_dict(row))
    return templates.TemplateResponse("sdqs_form.html", tpl(request, _MOD,
        modo="ver",
        registro=registro,
        meses=MESES,
        abogados=abogados,
        estados=ESTADOS_PROCESO,
    ))


# ── Editar ────────────────────────────────────────────────────────────────────

@router.get("/{id}/editar", response_class=HTMLResponse)
async def editar_get(request: Request, id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM sdqs WHERE id = ?", (id,)).fetchone()
    abogados = get_personal_oficina(conn)
    conn.close()
    if not row:
        return RedirectResponse("/sdqs/?msg=no_encontrado", status_code=303)
    return templates.TemplateResponse("sdqs_form.html", tpl(request, _MOD,
        modo="editar",
        registro=row_to_dict(row),
        meses=MESES,
        abogados=abogados,
        estados=ESTADOS_PROCESO,
    ))


@router.post("/{id}/editar")
async def editar_post(
    request: Request,
    id: int,
    mes: str = Form(""),
    fecha_asignacion: str = Form(""),
    sdqs_num: str = Form("", alias="sdqs"),
    url_sdqs: str = Form(""),
    fecha_vencimiento: str = Form(""),
    quejoso: str = Form(""),
    correo: str = Form(""),
    tema: str = Form(""),
    competencia_ocdi: str = Form("NO"),
    bpm: str = Form(""),
    responsable: str = Form(""),
    rad_salida: str = Form(""),
    url_rad_salida: str = Form(""),
    fecha_respuesta: str = Form(""),
    observaciones: str = Form(""),
    estado_proceso: str = Form(""),
    hecho_corrupto: str = Form(""),
    valor_institucional: str = Form(""),
    tipologia: str = Form(""),
):
    user = request.state.user
    if not _pw(user, _MOD):
        return RedirectResponse(f"/sdqs/?msg=sin_permiso", status_code=303)

    conn = get_db()
    conn.execute(
        """UPDATE sdqs SET
           mes=?, fecha_asignacion=?, sdqs=?, url_sdqs=?, fecha_vencimiento=?,
           quejoso=?, correo=?, tema=?, competencia_ocdi=?,
           bpm=?, responsable=?, rad_salida=?, url_rad_salida=?,
           fecha_respuesta=?,
           observaciones=?, estado_proceso=?, hecho_corrupto=?,
           valor_institucional=?, tipologia=?,
           updated_at=datetime('now','localtime')
           WHERE id=?""",
        (mes.upper(), fecha_asignacion, sdqs_num.upper(),
         url_sdqs or None,
         fecha_vencimiento or None,
         quejoso.upper(), correo, tema.upper(), competencia_ocdi.upper(),
         bpm or None, responsable or None, rad_salida or None, url_rad_salida or None,
         fecha_respuesta or None,
         observaciones, estado_proceso or None, hecho_corrupto or None,
         valor_institucional or None, tipologia or None, id),
    )
    conn.commit()
    conn.close()
    registrar_log(user, "editar", _MOD, f"ID: {id}")
    return RedirectResponse(f"/sdqs/{id}?msg=actualizado", status_code=303)


# ── Eliminar ──────────────────────────────────────────────────────────────────

@router.post("/{id}/eliminar")
async def eliminar(request: Request, id: int):
    user = request.state.user
    if not _pw(user, _MOD):
        return RedirectResponse("/sdqs/?msg=sin_permiso", status_code=303)
    conn = get_db()
    conn.execute("DELETE FROM sdqs WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    registrar_log(user, "eliminar", _MOD, f"ID: {id}")
    return RedirectResponse("/sdqs/?msg=eliminado", status_code=303)


# ── Helper de importación Excel ───────────────────────────────────────────────

def _importar_excel_sdqs(archivo_bytes: bytes):
    """
    Columnas del Excel BASE SDQS_Actualizada.xlsx (18 cols):
    0  MES
    1  FECHA ASIGNACION
    2  SDQS
    3  FECHA VENCIMIENTO
    4  ESTADO DIAS  (ignorado — se calcula)
    5  QUEJOSO
    6  CORREO
    7  TEMA
    8  COMPETENCIA OCDI
    9  BPM
    10 RESPONSABLE
    11 RAD SALIDA
    12 FECHA RESPUESTA
    13 OBSERVACIONES
    14 ESTADO PROCESO
    15 HECHO CORRUPTO
    16 VALOR INSTITUCIONAL
    17 TIPOLOGIA
    """
    try:
        import openpyxl
    except ImportError:
        return 0, ["openpyxl no instalado"]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), data_only=True, keep_links=True)
        ws = wb.active
    except Exception as e:
        return 0, [str(e)]

    # Detectar posiciones de columnas por encabezado (tolerante a reordenamientos)
    header_row = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
    COL = {h: i for i, h in enumerate(header_row)}

    def _ci(nombres):
        for n in nombres:
            if n in COL:
                return COL[n]
        return None

    i_mes    = _ci(["MES"])
    i_fasig  = _ci(["FECHA ASIGNACION", "FECHA ASIGNACIÓN", "FECHA RADICADO"])
    i_sdqs   = _ci(["SDQS"])
    i_urlsdqs = _ci(["URL SDQS"])
    i_fvenc  = _ci(["FECHA VENCIMINETO", "FECHA VENCIMIENTO"])
    i_quej   = _ci(["QUEJOSO"])
    i_correo = _ci(["CORREO"])
    i_tema   = _ci(["TEMA"])
    i_comp   = _ci(["COMPETENCIA OCDI"])
    i_bpm    = _ci(["BPM"])
    i_resp   = _ci(["RESPONSABLE"])
    i_rsal   = _ci(["RAD SALIDA"])
    i_frsp   = _ci(["FECHA RESPUESTA"])
    i_obs    = _ci(["OBSERVACIONES"])
    i_est    = _ci(["ESTADO PROCESO"])
    i_hech   = _ci(["HECHO CORRUPTO"])
    i_vinst  = _ci(["VALOR INSTITUCIONAL"])
    i_tipo   = _ci(["TIPOLOGIA"])

    def _v(row, idx):
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        s = str(val).strip()
        return None if s.upper() in ("NAN", "NONE", "") else s

    def _fecha(row, idx):
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        try:
            if isinstance(val, (date, datetime)):
                return val.isoformat()[:10]
            s = str(val).strip()
            if not s or s.upper() in ("NAN", "NONE"):
                return None
            return s[:10]
        except Exception:
            return None

    conn = get_db()
    count = 0
    errors = []

    for row_cells in ws.iter_rows(min_row=2, values_only=False):
        row = [c.value for c in row_cells]
        if all(v is None for v in row):
            continue
        try:
            sdqs_num = (_v(row, i_sdqs) or "").upper()
            if not sdqs_num:
                continue

            mes       = (_v(row, i_mes) or "").upper()
            fa        = _fecha(row, i_fasig)
            fv        = _fecha(row, i_fvenc)
            quejoso   = (_v(row, i_quej) or "").upper()
            correo    = _v(row, i_correo) or ""
            tema      = (_v(row, i_tema) or "").upper()
            comp_raw  = (_v(row, i_comp) or "NO").upper()
            comp      = "SI" if comp_raw in ("SI", "SÍ", "S", "1") else "NO"
            bpm       = _v(row, i_bpm)
            resp_raw  = (_v(row, i_resp) or "").upper()
            rad_sal   = _v(row, i_rsal)
            f_resp    = _fecha(row, i_frsp)
            obs       = _v(row, i_obs) or ""
            estado    = (_v(row, i_est) or "").upper() or None
            hecho     = _v(row, i_hech)
            valor_inst = _v(row, i_vinst)
            tipologia  = _v(row, i_tipo)

            responsable = RESP_MAP.get(resp_raw, resp_raw) or None
            if responsable == "":
                responsable = None

            # Leer hipervínculo de la celda SDQS
            url_sdqs_val = None
            if i_sdqs is not None and i_sdqs < len(row_cells):
                hl = row_cells[i_sdqs].hyperlink
                if hl:
                    url_sdqs_val = getattr(hl, "target", None)
            # Leer columna URL SDQS (texto plano) si existe y no se obtuvo de hipervínculo
            if not url_sdqs_val:
                url_sdqs_val = _v(row, i_urlsdqs) if i_urlsdqs is not None else None

            # Leer hipervínculo de la celda RAD SALIDA
            url_rad_sal = None
            if i_rsal is not None and i_rsal < len(row_cells):
                hl = row_cells[i_rsal].hyperlink
                if hl:
                    url_rad_sal = getattr(hl, "target", None)

            conn.execute(
                """INSERT INTO sdqs
                   (mes, fecha_asignacion, sdqs, url_sdqs, fecha_vencimiento, quejoso, correo, tema,
                    competencia_ocdi, bpm, responsable, rad_salida, url_rad_salida,
                    fecha_respuesta,
                    observaciones, estado_proceso, hecho_corrupto, valor_institucional,
                    tipologia, created_by)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(sdqs) DO UPDATE SET
                       mes=excluded.mes,
                       fecha_asignacion=excluded.fecha_asignacion,
                       url_sdqs=excluded.url_sdqs,
                       fecha_vencimiento=excluded.fecha_vencimiento,
                       quejoso=excluded.quejoso, correo=excluded.correo,
                       tema=excluded.tema, competencia_ocdi=excluded.competencia_ocdi,
                       bpm=excluded.bpm, responsable=excluded.responsable,
                       rad_salida=excluded.rad_salida,
                       url_rad_salida=excluded.url_rad_salida,
                       fecha_respuesta=excluded.fecha_respuesta,
                       observaciones=excluded.observaciones, estado_proceso=excluded.estado_proceso,
                       hecho_corrupto=excluded.hecho_corrupto,
                       valor_institucional=excluded.valor_institucional,
                       tipologia=excluded.tipologia,
                       updated_at=datetime('now','localtime')""",
                (mes, fa, sdqs_num, url_sdqs_val, fv, quejoso, correo, tema, comp,
                 bpm, responsable, rad_sal, url_rad_sal, f_resp, obs,
                 estado if estado else None, hecho, valor_inst, tipologia, "IMPORTACION"),
            )
            count += 1
        except Exception as e:
            errors.append(str(e))

    conn.commit()
    conn.close()
    return count, errors
